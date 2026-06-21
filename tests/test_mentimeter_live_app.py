from datetime import timedelta
from io import BytesIO

from openpyxl import load_workbook

from mentimeter_live_app.app import build_session_report, create_app, socketio
from mentimeter_live_app.models import db, Participant, Question, Response, Session, SessionRun, utcnow


class TestConfig:
    TESTING = True
    SECRET_KEY = "test"
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    MENTI_SEED_DEMO = True


def build_app():
    return create_app(
        {
            "TESTING": True,
            "SECRET_KEY": "test",
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "MENTI_SEED_DEMO": True,
        }
    )


def build_protected_app():
    return create_app(
        {
            "TESTING": True,
            "SECRET_KEY": "test",
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "MENTI_SEED_DEMO": True,
            "MENTI_ADMIN_PIN": "2468",
        }
    )


def build_password_app():
    return create_app(
        {
            "TESTING": True,
            "SECRET_KEY": "test",
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "MENTI_SEED_DEMO": True,
            "MENTI_ADMIN_USERNAME": "presentador",
            "MENTI_ADMIN_PASSWORD": "secreto",
        }
    )


def build_limited_app():
    return create_app(
        {
            "TESTING": True,
            "SECRET_KEY": "test",
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "MENTI_SEED_DEMO": True,
            "MENTI_RESPONSE_RATE_LIMIT": 1,
            "MENTI_RESPONSE_RATE_WINDOW": 60,
            "MAX_CONTENT_LENGTH": 512,
        }
    )


def test_demo_session_and_public_pages_exist():
    app = build_app()
    client = app.test_client()

    assert client.get("/admin").status_code == 200
    assert client.get("/join").status_code == 200

    payload = client.get("/api/sessions/123456").get_json()["session"]
    assert payload["code"] == "123456"
    assert len(payload["questions"]) == 4
    assert payload["questions"][0]["type"] == "content_slide"
    assert {question["type"] for question in payload["questions"]} == {"content_slide", "multiple_choice", "word_cloud", "quiz"}


def test_presenter_surfaces_mount_results_inside_slide_canvas():
    app = build_app()
    client = app.test_client()

    for path in ["/admin?code=123456", "/present/123456"]:
        html = client.get(path).get_data(as_text=True)
        canvas_index = html.index("data-slide-canvas")
        results_index = html.index("data-slide-results-stage")
        article_end = html.index("</article>", canvas_index)

        assert "live-results-strip" not in html
        assert canvas_index < results_index < article_end


def test_interactive_slide_layout_blocks_render_inside_slide_canvas():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Bloques editables"}).get_json()["session"]
    client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "multiple_choice",
            "title": "Pregunta movible",
            "prompt": "Elige una opcion",
            "options": ["A", "B"],
            "config": {
                "layout_blocks": {
                    "question": {"x": 5, "y": 10, "w": 80, "h": 22},
                    "activity": {"x": 6, "y": 40, "w": 35, "h": 45},
                    "results": {"x": 48, "y": 40, "w": 45, "h": 45},
                }
            },
        },
    )

    for path in [f"/admin?code={session['code']}", f"/present/{session['code']}"]:
        html = client.get(path).get_data(as_text=True)
        canvas_index = html.index("data-slide-canvas")
        article_end = html.index("</article>", canvas_index)

        for block_id in ["question", "activity", "results"]:
            block_index = html.index(f'data-layout-block-id="{block_id}"')
            assert canvas_index < block_index < article_end
        assert html.index("slide-layout-block-results") < html.index("data-slide-results-stage")


def test_admin_initial_canvas_marks_text_click_targets():
    app = build_app()
    client = app.test_client()

    content_html = client.get("/admin?code=123456").get_data(as_text=True)
    canvas_index = content_html.index("data-slide-canvas")
    article_end = content_html.index("</article>", canvas_index)
    content_canvas = content_html[canvas_index:article_end]
    assert 'data-text-box-id="title"' in content_canvas
    assert 'data-text-box-id="body"' in content_canvas

    session = client.post("/api/sessions", json={"title": "Targets iniciales"}).get_json()["session"]
    client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "multiple_choice",
            "title": "Pregunta",
            "prompt": "Elige",
            "options": ["A", "B"],
        },
    )
    interactive_html = client.get(f"/admin?code={session['code']}").get_data(as_text=True)
    canvas_index = interactive_html.index("data-slide-canvas")
    article_end = interactive_html.index("</article>", canvas_index)
    interactive_canvas = interactive_html[canvas_index:article_end]
    assert 'data-text-target="title"' in interactive_canvas
    assert 'data-text-target="prompt"' in interactive_canvas
    assert 'data-text-target="option:0"' in interactive_canvas


def test_optional_admin_pin_protects_presenter_surfaces_but_not_audience():
    app = build_protected_app()
    client = app.test_client()

    assert client.get("/admin").status_code == 302
    assert client.get("/api/sessions").status_code == 401
    socket_client = socketio.test_client(app, flask_test_client=client)
    control = socket_client.emit("presenter_control", {"code": "123456", "action": "start"}, callback=True)
    assert control == {"ok": False, "error": "No autorizado."}
    socket_client.disconnect()
    assert client.get("/join").status_code == 200
    assert client.get("/s/123456").status_code == 200
    assert client.get("/api/sessions/123456").status_code == 200

    bad_login = client.post("/admin-login", data={"pin": "0000"})
    assert bad_login.status_code == 401

    good_login = client.post("/admin-login?next=/admin", data={"pin": "2468"})
    assert good_login.status_code == 302
    assert good_login.headers["Location"].endswith("/admin")
    assert client.get("/admin").status_code == 200
    assert client.get("/api/sessions").status_code == 200


def test_admin_username_password_can_protect_presenter_surfaces():
    app = build_password_app()
    client = app.test_client()

    assert client.get("/admin").status_code == 302
    login_page = client.get("/admin-login")
    assert login_page.status_code == 200
    assert b'name="username"' in login_page.data
    assert b'name="password"' in login_page.data

    bad_login = client.post("/admin-login", data={"username": "presentador", "password": "mal"})
    assert bad_login.status_code == 401

    good_login = client.post(
        "/admin-login?next=/admin",
        data={"username": "presentador", "password": "secreto"},
    )
    assert good_login.status_code == 302
    assert good_login.headers["Location"].endswith("/admin")
    assert client.get("/admin").status_code == 200
    assert client.get("/api/sessions").status_code == 200


def test_presenter_can_create_edit_delete_and_reorder_questions():
    app = build_app()
    client = app.test_client()

    created = client.post("/api/sessions", json={"title": "Taller ciudadano"})
    assert created.status_code == 201
    session = created.get_json()["session"]
    assert len(session["code"]) == 6

    first = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "multiple_choice",
            "title": "Prioridad",
            "prompt": "Elige una",
            "options": ["A", "B"],
        },
    ).get_json()["question"]
    second = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "scale",
            "title": "Calificacion",
            "prompt": "Califica",
            "config": {"min": 1, "max": 5},
        },
    ).get_json()["question"]

    edited = client.patch(
        f"/api/sessions/{session['code']}/questions/{first['id']}",
        json={"title": "Prioridad actualizada"},
    ).get_json()["question"]
    assert edited["title"] == "Prioridad actualizada"

    reordered = client.post(
        f"/api/sessions/{session['code']}/questions/reorder",
        json={"question_ids": [second["id"], first["id"]]},
    ).get_json()["session"]
    assert [question["id"] for question in reordered["questions"]] == [second["id"], first["id"]]

    deleted = client.delete(f"/api/sessions/{session['code']}/questions/{first['id']}").get_json()["session"]
    assert [question["id"] for question in deleted["questions"]] == [second["id"]]


def test_templates_duplicate_theme_exports_and_insights_are_available():
    app = build_app()
    client = app.test_client()

    templates = client.get("/api/question-templates").get_json()["templates"]
    assert templates

    session = client.post("/api/sessions", json={"title": "Taller con plantillas"}).get_json()["session"]
    renamed = client.patch(f"/api/sessions/{session['code']}", json={"title": "Presentacion editada", "theme": "ocean"}).get_json()["session"]
    assert renamed["title"] == "Presentacion editada"
    assert renamed["theme"] == "ocean"

    created = client.post(f"/api/sessions/{session['code']}/questions", json=templates[0]["payload"]).get_json()["question"]
    duplicated = client.post(f"/api/sessions/{session['code']}/questions/{created['id']}/duplicate").get_json()["session"]
    assert len(duplicated["questions"]) == 2

    themed = client.post(f"/api/sessions/{session['code']}/control", json={"action": "set_theme", "theme": "contrast"}).get_json()["session"]
    assert themed["theme"] == "contrast"

    insights = client.get(f"/api/sessions/{session['code']}/insights").get_json()["insights"]
    assert insights["question_count"] == 2
    assert client.get(f"/api/sessions/{session['code']}/export.csv").status_code == 200
    assert client.get(f"/api/sessions/{session['code']}/export.xlsx").status_code == 200
    assert client.get(f"/api/sessions/{session['code']}/export.pdf").status_code == 200
    assert client.get("/healthz").get_json()["ok"] is True


def test_all_required_question_types_can_be_created():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Tipos requeridos"}).get_json()["session"]
    code = session["code"]

    examples = [
        {
            "type": "content_slide",
            "title": "Portada",
            "prompt": "",
            "config": {"layout": "qr", "body": "Escanea el codigo.", "show_qr": True},
        },
        {"type": "multiple_choice", "title": "Opcion", "prompt": "Elige", "options": ["A", "B"]},
        {"type": "word_cloud", "title": "Nube", "prompt": "Una palabra"},
        {"type": "scale", "title": "Escala", "prompt": "Califica", "config": {"min": 1, "max": 10}},
        {"type": "open_text", "title": "Abierta", "prompt": "Opina"},
        {"type": "ranking", "title": "Ranking", "prompt": "Ordena", "options": ["A", "B", "C"]},
        {
            "type": "quiz",
            "title": "Quiz",
            "prompt": "Correcta",
            "options": ["A", "B"],
            "correct_option_labels": ["B"],
            "config": {"timer_seconds": 20, "points": 50},
        },
    ]

    created_types = []
    for payload in examples:
        response = client.post(f"/api/sessions/{code}/questions", json=payload)
        assert response.status_code == 201
        created_types.append(response.get_json()["question"]["type"])

    assert created_types == [item["type"] for item in examples]


def test_interactive_question_config_forces_results_inside_slide():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Contrato visual"}).get_json()["session"]
    code = session["code"]

    multiple = client.post(
        f"/api/sessions/{code}/questions",
        json={
            "type": "multiple_choice",
            "title": "Prioridad",
            "prompt": "Elige una",
            "options": ["A", "B"],
            "config": {
                "result_placement": "below",
                "show_results": False,
                "result_layout": "list",
                "layout_blocks": {
                    "question": {"x": -8, "y": 120, "w": 8, "h": 4, "z": 200},
                    "activity": {"x": 30.5, "y": 44.25, "w": 36, "h": 32, "z": 6},
                    "legacy": {"x": 1, "y": 1, "w": 1, "h": 1},
                },
            },
        },
    ).get_json()["question"]
    assert multiple["config"]["result_placement"] == "slide"
    assert multiple["config"]["show_results"] is False
    assert multiple["config"]["result_layout"] == "list"
    assert set(multiple["config"]["layout_blocks"]) == {"question", "activity", "results"}
    assert multiple["config"]["layout_blocks"]["question"] == {"id": "question", "x": 0, "y": 90.0, "w": 12, "h": 10, "z": 100}
    assert multiple["config"]["layout_blocks"]["activity"]["x"] == 30.5
    assert multiple["config"]["layout_blocks"]["activity"]["y"] == 44.25

    cloud = client.post(
        f"/api/sessions/{code}/questions",
        json={
            "type": "word_cloud",
            "title": "Ideas",
            "prompt": "Una palabra",
            "config": {"result_placement": "footer", "result_layout": "footer"},
        },
    ).get_json()["question"]
    assert cloud["config"]["result_placement"] == "slide"
    assert cloud["config"]["show_results"] is True
    assert cloud["config"]["result_layout"] == "cloud"

    with app.app_context():
        question = db.session.get(Question, multiple["id"])
        question.config_json = {"result_placement": "below"}
        db.session.commit()

    refreshed = client.get(f"/api/sessions/{code}").get_json()["session"]
    refreshed_multiple = next(question for question in refreshed["questions"] if question["id"] == multiple["id"])
    assert refreshed_multiple["config"]["result_placement"] == "slide"
    assert refreshed_multiple["config"]["show_results"] is True
    assert refreshed_multiple["config"]["result_layout"] == "chart"
    assert refreshed_multiple["config"]["layout_blocks"]["question"]["w"] == 86
    assert refreshed_multiple["config"]["layout_blocks"]["results"]["x"] == 53


def test_interactive_question_config_preserves_text_styles():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Estilos de texto"}).get_json()["session"]
    code = session["code"]

    question = client.post(
        f"/api/sessions/{code}/questions",
        json={
            "type": "multiple_choice",
            "title": "Titulo",
            "prompt": "Pregunta",
            "options": ["A", "B"],
            "config": {
                "text_styles": {
                    "title": {
                        "font_size": 44,
                        "font_weight": 800,
                        "color": "#2563eb",
                        "background": "transparent",
                        "align": "center",
                        "auto_fit": False,
                    },
                    "prompt": {
                        "font_size": 22,
                        "font_weight": 400,
                        "color": "#334155",
                        "background": "#ffffff",
                        "align": "left",
                        "auto_fit": True,
                    },
                    "option:0": {
                        "font_size": 18,
                        "font_weight": 800,
                        "color": "#647c3d",
                        "background": "transparent",
                        "align": "right",
                        "auto_fit": True,
                    },
                    "bad key !": {"font_size": 999, "color": "bad"},
                }
            },
        },
    ).get_json()["question"]

    styles = question["config"]["text_styles"]
    assert styles["title"]["font_size"] == 44
    assert styles["title"]["font_weight"] == 800
    assert styles["title"]["auto_fit"] is False
    assert styles["title"]["align"] == "center"
    assert styles["prompt"]["background"] == "#ffffff"
    assert styles["option:0"]["color"] == "#647c3d"
    assert styles["option:0"]["align"] == "right"
    assert styles["badkey"]["font_size"] == 120
    assert styles["badkey"]["color"] == "#17212f"


def test_content_slide_is_saved_but_does_not_accept_responses():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Presentacion con contenido"}).get_json()["session"]
    slide = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "content_slide",
            "title": "Bienvenida",
            "prompt": "",
            "config": {"layout": "instructions", "body": "Lee las instrucciones.", "show_qr": False},
        },
    ).get_json()["question"]
    assert slide["config"]["layout"] == "instructions"
    assert slide["config"]["body"] == "Lee las instrucciones."

    client.post(f"/api/sessions/{session['code']}/control", json={"action": "start"})
    response = client.post(f"/api/sessions/{session['code']}/questions/{slide['id']}/responses", json={"text": "hola"})
    assert response.status_code == 400
    assert "no acepta respuestas" in response.get_json()["error"]


def test_content_slide_text_boxes_are_saved_sanitized_and_synced():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Editor visual"}).get_json()["session"]

    response = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "content_slide",
            "title": "Portada",
            "prompt": "",
            "config": {
                "layout": "title",
                "body": "Texto anterior",
                "text_boxes": [
                    {
                        "id": "title",
                        "text": "Contexto del SGC",
                        "x": -10,
                        "y": 120,
                        "w": 2,
                        "h": 140,
                        "font_size": 160,
                        "font_weight": "bold",
                        "color": "#2563EB",
                        "background": "transparent",
                        "align": "center",
                        "auto_fit": False,
                        "z": 5,
                    },
                    {
                        "id": "body",
                        "text": "Partes interesadas y alcance",
                        "x": 18.5,
                        "y": 42,
                        "w": 52,
                        "h": 18,
                        "font_size": 28,
                        "font_weight": 400,
                        "color": "#334155",
                        "background": "#FFFFFF",
                        "align": "right",
                    },
                ],
            },
        },
    )

    assert response.status_code == 201
    config = response.get_json()["question"]["config"]
    title_box, body_box = config["text_boxes"]
    assert title_box["x"] == 0
    assert title_box["y"] == 100
    assert title_box["w"] == 5
    assert title_box["h"] == 100
    assert title_box["font_size"] == 120
    assert title_box["font_weight"] == 800
    assert title_box["color"] == "#2563eb"
    assert title_box["auto_fit"] is False
    assert body_box["background"] == "#ffffff"
    assert body_box["align"] == "right"
    assert config["body"] == "Partes interesadas y alcance"


def test_content_slide_text_boxes_reject_invalid_color_and_text_limits():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Validacion visual"}).get_json()["session"]
    base_payload = {
        "type": "content_slide",
        "title": "Portada",
        "prompt": "",
        "config": {
            "layout": "title",
            "text_boxes": [
                {"id": "title", "text": "Titulo", "color": "#17212f"},
            ],
        },
    }

    bad_color = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={**base_payload, "config": {"layout": "title", "text_boxes": [{"text": "Titulo", "color": "blue"}]}},
    )
    assert bad_color.status_code == 400
    assert "Color invalido" in bad_color.get_json()["error"]

    bad_text = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            **base_payload,
            "config": {"layout": "title", "text_boxes": [{"text": "x" * 1201, "color": "#17212f"}]},
        },
    )
    assert bad_text.status_code == 400


def test_duplicate_content_slide_preserves_text_boxes_and_drops_runtime_state():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Duplicado visual"}).get_json()["session"]
    created = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "content_slide",
            "title": "Portada",
            "prompt": "",
            "config": {
                "layout": "title",
                "text_boxes": [
                    {"id": "title", "text": "Slide editable", "color": "#17212f", "font_size": 48},
                    {"id": "body", "text": "Cuerpo editable", "color": "#334155", "font_size": 24},
                ],
            },
        },
    ).get_json()["question"]

    with app.app_context():
        question = db.session.get(Question, created["id"])
        config = dict(question.config_json or {})
        config["timer_started_at"] = "2026-06-21T12:00:00"
        question.config_json = config
        db.session.commit()

    duplicate = client.post(f"/api/sessions/{session['code']}/questions/{created['id']}/duplicate").get_json()["question"]
    assert duplicate["config"]["text_boxes"][0]["text"] == "Slide editable"
    assert duplicate["config"]["text_boxes"][1]["text"] == "Cuerpo editable"
    assert "timer_started_at" not in duplicate["config"]


def test_single_choice_vote_is_replaced_for_same_participant():
    app = build_app()
    client = app.test_client()
    session = client.get("/api/sessions/123456").get_json()["session"]
    question = next(question for question in session["questions"] if question["type"] == "multiple_choice")
    question_index = [item["id"] for item in session["questions"]].index(question["id"])
    first_option, second_option = question["options"][:2]

    assert client.post("/api/sessions/123456/control", json={"action": "start"}).status_code == 200
    assert client.post("/api/sessions/123456/control", json={"action": "go_to_slide", "index": question_index}).status_code == 200
    assert client.post(
        f"/api/sessions/123456/questions/{question['id']}/responses",
        json={"option_id": first_option["id"]},
    ).status_code == 200
    payload = client.post(
        f"/api/sessions/123456/questions/{question['id']}/responses",
        json={"option_id": second_option["id"]},
    ).get_json()
    assert payload["ok"] is True

    with app.app_context():
        db_session = Session.query.filter_by(code="123456").first()
        responses = Response.query.filter_by(session_id=db_session.id, question_id=question["id"]).all()
        assert len(responses) == 1
        assert responses[0].payload_json["option_id"] == second_option["id"]

    counts = {item["id"]: item["count"] for item in payload["results"]["options"]}
    assert counts[first_option["id"]] == 0
    assert counts[second_option["id"]] == 1


def test_session_runs_preserve_responses_across_reset_for_same_participant():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Historial persistente"}).get_json()["session"]
    question = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "multiple_choice",
            "title": "Prioridad",
            "prompt": "Elige una",
            "options": ["Primera", "Segunda"],
        },
    ).get_json()["question"]
    first_option, second_option = question["options"]

    started = client.post(f"/api/sessions/{session['code']}/control", json={"action": "start"}).get_json()["session"]
    assert started["active_run_id"]
    assert len(started["runs"]) == 1
    first_run_id = started["active_run_id"]
    assert client.post(
        f"/api/sessions/{session['code']}/questions/{question['id']}/responses",
        json={"option_id": first_option["id"]},
    ).status_code == 200

    reset = client.post(f"/api/sessions/{session['code']}/control", json={"action": "reset"}).get_json()["session"]
    assert reset["status"] == "draft"
    assert reset["active_run_id"] is None
    assert reset["runs"][0]["status"] == "closed"
    assert reset["runs"][0]["response_count"] == 1
    reset_question = next(item for item in reset["questions"] if item["id"] == question["id"])
    assert reset_question["results"]["total"] == 0

    restarted = client.post(f"/api/sessions/{session['code']}/control", json={"action": "start"}).get_json()["session"]
    second_run_id = restarted["active_run_id"]
    assert second_run_id != first_run_id
    second_submit = client.post(
        f"/api/sessions/{session['code']}/questions/{question['id']}/responses",
        json={"option_id": second_option["id"]},
    ).get_json()
    assert second_submit["results"]["total"] == 1

    with app.app_context():
        db_session = Session.query.filter_by(code=session["code"]).first()
        responses = Response.query.filter_by(session_id=db_session.id, question_id=question["id"]).order_by(Response.id).all()
        assert len(responses) == 2
        assert {response.run_id for response in responses} == {first_run_id, second_run_id}
        assert {response.payload_json["option_id"] for response in responses} == {first_option["id"], second_option["id"]}
        first_run = db.session.get(SessionRun, first_run_id)
        assert first_run.summary_json["response_count"] == 1


def test_run_exports_xlsx_pdf_and_report_are_interpreted_by_run():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Reporte por ejecucion"}).get_json()["session"]
    choice = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "multiple_choice",
            "title": "Votacion",
            "prompt": "Elige",
            "options": ["A", "B"],
        },
    ).get_json()["question"]
    cloud = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={"type": "word_cloud", "title": "Nube", "prompt": "Una palabra"},
    ).get_json()["question"]

    started = client.post(f"/api/sessions/{session['code']}/control", json={"action": "start"}).get_json()["session"]
    run_id = started["active_run_id"]
    client.post(
        f"/api/sessions/{session['code']}/questions/{choice['id']}/responses",
        json={"option_id": choice["options"][0]["id"]},
    )
    client.post(f"/api/sessions/{session['code']}/control", json={"action": "go_to_slide", "index": 1})
    client.post(f"/api/sessions/{session['code']}/questions/{cloud['id']}/responses", json={"text": "calidad"})
    closed = client.post(f"/api/sessions/{session['code']}/control", json={"action": "close"}).get_json()["session"]
    assert closed["runs"][0]["status"] == "closed"

    with app.app_context():
        db_session = Session.query.filter_by(code=session["code"]).first()
        report = build_session_report(db_session, db.session.get(SessionRun, run_id))
        assert report["run_label"] == "Ejecucion 1"
        assert report["response_count"] == 2
        assert report["responses"][0]["participant"] == "Participante 1"
        assert any(row["metric"] == "A" and row["count"] == 1 for row in report["slides"][0]["result_rows"])

    xlsx = client.get(f"/api/sessions/{session['code']}/runs/{run_id}/export.xlsx")
    assert xlsx.status_code == 200
    workbook = load_workbook(BytesIO(xlsx.data), data_only=True)
    assert {"Resumen", "Diapositivas", "Resultados", "Respuestas", "Opcion multiple", "Nube"}.issubset(set(workbook.sheetnames))
    assert workbook["Resumen"]["A1"].value == "Campo"
    assert any(row[3] == "A" and row[5] == 1 for row in workbook["Resultados"].iter_rows(min_row=2, values_only=True))

    pdf = client.get(f"/api/sessions/{session['code']}/runs/{run_id}/export.pdf")
    assert pdf.status_code == 200
    assert pdf.data.startswith(b"%PDF")


def test_legacy_responses_without_run_are_reported_as_historical():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Legacy"}).get_json()["session"]
    question = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={"type": "open_text", "title": "Comentario", "prompt": "Opina"},
    ).get_json()["question"]

    with app.app_context():
        db_session = Session.query.filter_by(code=session["code"]).first()
        participant = Participant(session=db_session, token="legacy-token", connected=False)
        db.session.add(participant)
        db.session.flush()
        response = Response(
            session=db_session,
            question=db.session.get(Question, question["id"]),
            participant=participant,
            response_key="default",
            payload_json={"text": "respuesta anterior", "status": "approved"},
        )
        db.session.add(response)
        db.session.commit()
        report = build_session_report(db_session)
        assert report["run_label"] == "Historico sin ejecucion"
        assert report["response_count"] == 1
        assert report["responses"][0]["participant"] == "Participante 1"

    workbook = load_workbook(BytesIO(client.get(f"/api/sessions/{session['code']}/export.xlsx").data), data_only=True)
    assert workbook["Resumen"]["B4"].value == "Historico sin ejecucion"


def test_public_response_rate_limit_and_payload_size_are_enforced():
    app = build_limited_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Sesion limitada"}).get_json()["session"]
    question = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "multiple_choice",
            "title": "Voto",
            "prompt": "Elige",
            "options": ["A", "B"],
        },
    ).get_json()["question"]
    option_id = question["options"][0]["id"]
    client.post(f"/api/sessions/{session['code']}/control", json={"action": "start"})

    first = client.post(f"/api/sessions/{session['code']}/questions/{question['id']}/responses", json={"option_id": option_id})
    assert first.status_code == 200

    second = client.post(f"/api/sessions/{session['code']}/questions/{question['id']}/responses", json={"option_id": option_id})
    assert second.status_code == 429
    assert second.headers["Retry-After"]

    oversized = client.post(
        f"/api/sessions/{session['code']}/questions/{question['id']}/responses",
        data="x" * 800,
        content_type="application/json",
    )
    assert oversized.status_code == 413

    socket_client = socketio.test_client(app, flask_test_client=client)
    socket_response = socket_client.emit(
        "submit_response",
        {"code": session["code"], "question_id": question["id"], "payload": {"text": "x" * 800}},
        callback=True,
    )
    assert socket_response == {"ok": False, "error": "Payload demasiado grande."}
    socket_client.disconnect()


def test_manual_moderation_hides_then_reveals_open_text_results():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions", json={"title": "Moderacion"}).get_json()["session"]
    question = client.post(
        f"/api/sessions/{session['code']}/questions",
        json={
            "type": "word_cloud",
            "title": "Ideas",
            "prompt": "Una palabra",
            "config": {"moderation": "manual"},
        },
    ).get_json()["question"]
    client.post(f"/api/sessions/{session['code']}/control", json={"action": "start"})

    submitted = client.post(
        f"/api/sessions/{session['code']}/questions/{question['id']}/responses",
        json={"text": "integridad"},
    ).get_json()
    assert submitted["results"]["words"] == []

    state = client.get(f"/api/sessions/{session['code']}").get_json()["session"]
    pending = state["questions"][0]["pending_responses"][0]
    approved = client.post(
        f"/api/sessions/{session['code']}/questions/{question['id']}/responses/{pending['id']}/moderate",
        json={"action": "approve"},
    ).get_json()
    assert approved["results"]["words"][0]["text"] == "integridad"


def test_quiz_timer_is_enforced_by_server_state():
    app = build_app()
    client = app.test_client()
    session = client.get("/api/sessions/123456").get_json()["session"]
    quiz = next(question for question in session["questions"] if question["type"] == "quiz")
    quiz_index = [question["id"] for question in session["questions"]].index(quiz["id"])
    client.post("/api/sessions/123456/control", json={"action": "start"})
    client.post("/api/sessions/123456/control", json={"action": "go_to_slide", "index": quiz_index})

    with app.app_context():
        question = db.session.get(Question, quiz["id"])
        config = dict(question.config_json)
        config["timer_started_at"] = (utcnow() - timedelta(seconds=config["timer_seconds"] + 5)).isoformat()
        question.config_json = config
        db.session.commit()

    refreshed = client.get("/api/sessions/123456").get_json()["session"]
    refreshed_quiz = next(question for question in refreshed["questions"] if question["id"] == quiz["id"])
    assert refreshed_quiz["is_open"] is False
    assert refreshed_quiz["timer"]["remaining"] == 0


def test_socketio_join_navigation_and_live_results():
    app = build_app()
    client = app.test_client()
    session = client.post("/api/sessions/123456/control", json={"action": "start"}).get_json()["session"]
    word_question = next(question for question in session["questions"] if question["type"] == "word_cloud")
    word_index = [question["id"] for question in session["questions"]].index(word_question["id"])

    socket_client = socketio.test_client(app)
    join_ack = socket_client.emit("join_session", {"code": "123456"}, callback=True)
    assert join_ack["ok"] is True

    next_ack = socket_client.emit("presenter_control", {"code": "123456", "action": "go_to_slide", "index": word_index}, callback=True)
    assert next_ack["ok"] is True
    assert next_ack["session"]["active_question_id"] == word_question["id"]

    submit_ack = socket_client.emit(
        "submit_response",
        {
            "code": "123456",
            "question_id": word_question["id"],
            "participant_token": join_ack["participant_token"],
            "payload": {"text": "transparencia"},
        },
        callback=True,
    )
    assert submit_ack["ok"] is True
    assert submit_ack["results"]["words"][0]["text"] == "transparencia"
    received = socket_client.get_received()
    assert any(item["name"] == "results_updated" for item in received)

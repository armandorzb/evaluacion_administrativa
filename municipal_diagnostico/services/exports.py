from __future__ import annotations

import csv
import math
import zipfile
from datetime import UTC, datetime
from io import BytesIO, StringIO
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.shapes import Circle, Drawing, Line, Polygon, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from municipal_diagnostico.services.analytics import build_evaluation_report_detail, summarize_period
from municipal_diagnostico.services.campaign_analytics import build_assignment_report_detail


EXCEL_THEME = {
    "navy": "163042",
    "navy_soft": "EEF2F5",
    "sage": "728B50",
    "sage_soft": "EEF3E5",
    "line": "D9E0E5",
    "muted": "516673",
    "low": "B13A3F",
    "medium": "C56E14",
    "high": "B39B26",
    "optimal": "5D7F48",
    "white": "FFFFFF",
}

PDF_THEME = {
    "navy": colors.HexColor("#163042"),
    "navy_soft": colors.HexColor("#EEF2F5"),
    "sage": colors.HexColor("#728B50"),
    "sage_soft": colors.HexColor("#EEF3E5"),
    "line": colors.HexColor("#D9E0E5"),
    "muted": colors.HexColor("#516673"),
    "low": colors.HexColor("#B13A3F"),
    "medium": colors.HexColor("#C56E14"),
    "high": colors.HexColor("#B39B26"),
    "optimal": colors.HexColor("#5D7F48"),
}

WORD_THEME = {
    "navy": "163042",
    "line": "D9E0E5",
    "muted": "516673",
    "low": "B13A3F",
    "medium": "C56E14",
    "high": "B39B26",
    "optimal": "5D7F48",
    "sage": "728B50",
    "sage_soft": "EEF3E5",
    "white": "FFFFFF",
}

PDF_FOOTER_LEGEND = "Contrato de prestación de servicios CAAS/ORD/72/2025/PSP-122-2026"

THIN_SIDE = Side(style="thin", color=EXCEL_THEME["line"])
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_evaluation_pdf(evaluacion) -> BytesIO:
    detail = build_evaluation_report_detail(evaluacion)
    return _build_detail_pdf(
        detail,
        title="Reporte ejecutivo de evaluación",
        subtitle_lines=[
            f"Dependencia: {evaluacion.dependencia.nombre}",
            f"Periodo: {evaluacion.periodo.nombre}",
            f"Estado: {detail['summary']['state_label']}",
            f"Cuestionario: {detail['questionnaire_name']} | Ejes: {detail['total_axes']} | Evidencias activas: {detail['total_evidence']}",
        ],
        preliminary_message="Resultado preliminar. Esta evaluación sigue en seguimiento operativo y todavía no forma parte del universo oficial cerrado.",
        cover_highlights=[
            evaluacion.dependencia.nombre,
            f"Cuestionario: {detail['questionnaire_name']}",
        ],
        cover_detail_lines=[],
    )


def build_evaluation_excel(evaluacion) -> BytesIO:
    detail = build_evaluation_report_detail(evaluacion)
    return _build_detail_excel(
        detail,
        title="Reporte ejecutivo de evaluacion",
        subtitle=f"{evaluacion.dependencia.nombre} | {evaluacion.periodo.nombre}",
    )


def build_evaluation_word(evaluacion) -> BytesIO:
    detail = build_evaluation_report_detail(evaluacion)
    return _build_detail_word(
        detail,
        title="Reporte ejecutivo de evaluacion",
        subtitle_lines=[
            f"Dependencia: {evaluacion.dependencia.nombre}",
            f"Periodo: {evaluacion.periodo.nombre}",
            f"Estado: {detail['summary']['state_label']}",
            f"Cuestionario: {detail['questionnaire_name']} | Ejes: {detail['total_axes']} | Evidencias activas: {detail['total_evidence']}",
        ],
        preliminary_message="Resultado preliminar. Esta evaluacion sigue en seguimiento operativo y todavia no forma parte del universo oficial cerrado.",
    )


def build_assignment_pdf(asignacion) -> BytesIO:
    detail = build_assignment_report_detail(asignacion)
    dependency_name = asignacion.dependencia_visible.nombre if asignacion.dependencia_visible else "Sin dependencia"
    return _build_detail_pdf(
        detail,
        title="Reporte ejecutivo de asignación",
        subtitle_lines=[
            f"Campaña: {asignacion.campana.nombre}",
            f"Objetivo: {asignacion.objetivo_nombre}",
            f"Dependencia: {dependency_name}",
            f"Estado: {detail['summary']['state_label']}",
            f"Cuestionario: {detail['questionnaire_name']} | Ejes: {detail['total_axes']} | Evidencias activas: {detail['total_evidence']}",
        ],
        preliminary_message="Resultado preliminar. Esta asignación sigue abierta y el reporte puede cambiar mientras se completa o valida el cuestionario.",
        cover_highlights=[
            dependency_name,
            f"Cuestionario: {detail['questionnaire_name']}",
        ],
        cover_detail_lines=[],
    )


def build_assignment_excel(asignacion) -> BytesIO:
    detail = build_assignment_report_detail(asignacion)
    return _build_detail_excel(
        detail,
        title="Reporte ejecutivo de asignacion",
        subtitle=f"{asignacion.campana.nombre} | {asignacion.objetivo_nombre}",
    )


def build_assignment_word(asignacion) -> BytesIO:
    detail = build_assignment_report_detail(asignacion)
    dependency_name = asignacion.dependencia_visible.nombre if asignacion.dependencia_visible else "Sin dependencia"
    return _build_detail_word(
        detail,
        title="Reporte ejecutivo de asignacion",
        subtitle_lines=[
            f"Campana: {asignacion.campana.nombre}",
            f"Objetivo: {asignacion.objetivo_nombre}",
            f"Dependencia: {dependency_name}",
            f"Estado: {detail['summary']['state_label']}",
            f"Cuestionario: {detail['questionnaire_name']} | Ejes: {detail['total_axes']} | Evidencias activas: {detail['total_evidence']}",
        ],
        preliminary_message="Resultado preliminar. Esta asignacion sigue abierta y el reporte puede cambiar mientras se completa o valida el cuestionario.",
    )


def build_period_excel(periodo, include_states=None) -> BytesIO:
    summary = summarize_period(periodo, include_states=include_states)
    workbook = Workbook()

    operational_sheet = workbook.active
    operational_sheet.title = "Avance Operativo"
    _write_excel_title(operational_sheet, "Avance operativo del periodo", periodo.nombre, end_column=7)
    operational_sheet.append([])
    operational_sheet.append(["Ranking", "Dependencia", "Estado", "Preliminar", "Indice", "Nivel", "Avance"])
    header_row = operational_sheet.max_row
    for row in summary["operational_ranking"]:
        operational_sheet.append(
            [
                row["operational_rank"],
                row["dependencia"],
                row["estado_label"],
                "Si" if row["is_preliminary"] else "No",
                row["indice"],
                row["nivel"],
                row["avance"],
            ]
        )
    _style_excel_metrics_table(operational_sheet, header_rows={header_row})
    _set_excel_widths(
        operational_sheet,
        {"A": 10, "B": 28, "C": 18, "D": 12, "E": 12, "F": 14, "G": 12},
    )

    official_sheet = workbook.create_sheet("Resultado Oficial")
    _write_excel_title(official_sheet, "Resultado oficial del periodo", periodo.nombre, end_column=6)
    official_sheet.append([])
    official_sheet.append(["Ranking", "Dependencia", "Estado", "Indice", "Nivel", "Avance"])
    official_header_row = official_sheet.max_row
    for row in summary["official_ranking"]:
        official_sheet.append(
            [
                row["official_rank"],
                row["dependencia"],
                row["estado_label"],
                row["indice"],
                row["nivel"],
                row["avance"],
            ]
        )
    _style_excel_metrics_table(official_sheet, header_rows={official_header_row})
    _set_excel_widths(official_sheet, {"A": 10, "B": 28, "C": 18, "D": 12, "E": 14, "F": 12})

    axis_sheet = workbook.create_sheet("Ejes")
    _write_excel_title(axis_sheet, "Detalle por eje", periodo.nombre, end_column=7)
    axis_sheet.append([])
    axis_sheet.append(["Dependencia", "Estado", "Eje", "Promedio", "Ponderacion", "Brecha", "Prioridad"])
    axis_header_row = axis_sheet.max_row
    for row in summary["operational_ranking"]:
        for axis in row["axes"]:
            axis_sheet.append(
                [
                    row["dependencia"],
                    row["estado_label"],
                    axis["nombre"],
                    axis["promedio"],
                    axis["ponderacion"],
                    axis["brecha"],
                    axis["prioridad"],
                ]
            )
    _style_excel_metrics_table(axis_sheet, header_rows={axis_header_row})
    _set_excel_widths(
        axis_sheet,
        {"A": 28, "B": 18, "C": 32, "D": 12, "E": 14, "F": 12, "G": 18},
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_period_csv(periodo, include_states=None) -> BytesIO:
    summary = summarize_period(periodo, include_states=include_states)
    text = StringIO()
    writer = csv.writer(text)
    writer.writerow(["ranking_operativo", "dependencia", "estado", "preliminar", "indice", "nivel", "avance"])
    for row in summary["operational_ranking"]:
        writer.writerow(
            [
                row["operational_rank"],
                row["dependencia"],
                row["estado_label"],
                "si" if row["is_preliminary"] else "no",
                row["indice"],
                row["nivel"],
                row["avance"],
            ]
        )
    payload = BytesIO(text.getvalue().encode("utf-8-sig"))
    payload.seek(0)
    return payload


def _build_detail_pdf(
    detail: dict,
    *,
    title: str,
    subtitle_lines: list[str],
    preliminary_message: str | None = None,
    cover_highlights: list[str] | None = None,
    cover_detail_lines: list[str] | None = None,
) -> BytesIO:
    styles = _pdf_styles()
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.65 * inch,
        rightMargin=0.65 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.55 * inch,
    )

    story = _build_pdf_cover_story(
        detail,
        title=title,
        subtitle_lines=subtitle_lines,
        styles=styles,
        preliminary_message=preliminary_message,
        cover_highlights=cover_highlights or [],
        cover_detail_lines=cover_detail_lines or [],
    )
    story.extend([Paragraph(_paragraph_escape(title), styles["title"])])
    for line in subtitle_lines:
        story.append(Paragraph(_paragraph_escape(line), styles["subtitle"]))
    story.append(Spacer(1, 0.14 * inch))

    if preliminary_message and detail["summary"].get("is_preliminary"):
        story.append(Paragraph(_paragraph_escape(preliminary_message), styles["warning"]))
        story.append(Spacer(1, 0.12 * inch))

    story.extend(
        [
            Paragraph("Índice de madurez", styles["section"]),
            _build_pdf_summary_table(detail["summary"]),
            Spacer(1, 0.12 * inch),
            _build_pdf_overview_panel(detail, styles),
            Spacer(1, 0.18 * inch),
            Paragraph("Mapa de brechas", styles["section"]),
            _build_pdf_heatmap_table(detail, styles),
            Spacer(1, 0.18 * inch),
            Paragraph("Síntesis por eje", styles["section"]),
            _build_pdf_axis_summary_table(detail, styles),
        ]
    )

    if detail["axes"]:
        story.extend(
            [
                PageBreak(),
                Paragraph("Cuestionario respondido", styles["section"]),
                Paragraph(
                    "Cada bloque resume la respuesta elegida, la opción contestada y el contexto de captura por eje.",
                    styles["body_muted"],
                ),
                Spacer(1, 0.12 * inch),
            ]
        )
        for index, axis in enumerate(detail["axes"]):
            if index > 0:
                story.append(PageBreak())
            story.extend(_build_pdf_axis_story(axis, styles))

    document.build(story, onFirstPage=_draw_pdf_chrome, onLaterPages=_draw_pdf_chrome)
    buffer.seek(0)
    return buffer


def _build_detail_excel(detail: dict, *, title: str, subtitle: str) -> BytesIO:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    _write_excel_title(summary_sheet, title, subtitle, end_column=4)
    summary_sheet.append([])
    summary_sheet.append(["Indicador", "Valor"])
    metrics_header_row = summary_sheet.max_row
    summary_sheet.append(["Indice global", detail["summary"]["index_score"]])
    summary_sheet.append(["Nivel", detail["summary"]["level_label"]])
    summary_sheet.append(["Estado", detail["summary"]["state_label"]])
    summary_sheet.append(["Avance", f"{detail['summary']['completion']}%"])
    summary_sheet.append(["Ejes criticos", len(detail["summary"]["critical_axes"])])
    summary_sheet.append(["Cuestionario", detail["questionnaire_name"]])
    summary_sheet.append(["Ejes totales", detail["total_axes"]])
    summary_sheet.append(["Evidencias activas", detail["total_evidence"]])
    _style_excel_metrics_table(summary_sheet, header_rows={metrics_header_row})

    guide_title_row = summary_sheet.max_row + 2
    summary_sheet.merge_cells(start_row=guide_title_row, start_column=1, end_row=guide_title_row, end_column=2)
    guide_title_cell = summary_sheet.cell(row=guide_title_row, column=1)
    guide_title_cell.value = "Guia de interpretacion"
    guide_title_cell.fill = PatternFill("solid", fgColor=EXCEL_THEME["sage"])
    guide_title_cell.font = Font(bold=True, color=EXCEL_THEME["white"])
    guide_title_cell.alignment = Alignment(horizontal="left", vertical="center")

    summary_sheet.append(["Rango", "Nivel"])
    guide_header_row = summary_sheet.max_row
    guide_rows: list[tuple[int, str]] = []
    for guide in detail["score_guide"]:
        summary_sheet.append([guide["range"], guide["label"]])
        guide_rows.append((summary_sheet.max_row, guide["slug"]))
    _style_excel_metrics_table(summary_sheet, header_rows={metrics_header_row, guide_header_row})
    for row_index, slug in guide_rows:
        for column_index in (1, 2):
            cell = summary_sheet.cell(row=row_index, column=column_index)
            cell.fill = PatternFill("solid", fgColor=_theme_color_for_slug(slug))
            cell.font = Font(bold=True, color=EXCEL_THEME["white"])
            cell.alignment = CENTER_ALIGNMENT
    _set_excel_widths(summary_sheet, {"A": 24, "B": 44, "C": 18, "D": 18})

    heatmap_sheet = workbook.create_sheet("Brechas")
    _write_excel_title(heatmap_sheet, "Mapa de brechas", subtitle, end_column=7)
    heatmap_sheet.append([])
    heatmap_sheet.append(["Eje", "Bajo", "Medio", "Alto", "Optimo", "Promedio", "Brecha"])
    heatmap_header_row = heatmap_sheet.max_row
    painted_heatmap_rows: list[tuple[int, str]] = []
    for row in detail["heatmap_rows"]:
        heatmap_sheet.append(
            [
                f"{row['axis_key']} {row['axis_name']}",
                "",
                "",
                "",
                "",
                row["average"],
                row["gap"],
            ]
        )
        painted_heatmap_rows.append((heatmap_sheet.max_row, row["selected_slug"]))
    _style_excel_metrics_table(heatmap_sheet, header_rows={heatmap_header_row})
    for row_index, slug in painted_heatmap_rows:
        _paint_excel_heatmap_row(heatmap_sheet, row_index, slug)
    _set_excel_widths(
        heatmap_sheet,
        {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12},
    )

    axis_sheet = workbook.create_sheet("Ejes")
    _write_excel_title(axis_sheet, "Sintesis por eje", subtitle, end_column=10)
    axis_sheet.append([])
    axis_sheet.append(
        [
            "Clave",
            "Eje",
            "Promedio",
            "Brecha",
            "Prioridad",
            "Avance",
            "Comentario general",
            "Responsable",
            "Actualizado",
            "Evidencias",
        ]
    )
    axis_header_row = axis_sheet.max_row
    for axis in detail["axes"]:
        axis_sheet.append(
            [
                axis["clave"],
                axis["nombre"],
                axis["summary"]["promedio"],
                axis["summary"]["brecha"],
                axis["summary"]["prioridad"],
                f"{axis['summary']['progreso']}%",
                axis["axis_comment"],
                axis["axis_comment_author"],
                axis["axis_comment_updated_at"],
                len(axis["evidence"]),
            ]
        )
    _style_excel_metrics_table(axis_sheet, header_rows={axis_header_row})
    _set_excel_widths(
        axis_sheet,
        {"A": 10, "B": 24, "C": 12, "D": 12, "E": 18, "F": 12, "G": 44, "H": 24, "I": 20, "J": 12},
    )

    questionnaire_sheet = workbook.create_sheet("Cuestionario")
    _write_excel_title(questionnaire_sheet, "Cuestionario respondido", subtitle, end_column=10)
    questionnaire_sheet.append([])
    questionnaire_sheet.append(
        [
            "Clave",
            "Eje",
            "Reactivo",
            "Pregunta",
            "Nivel seleccionado",
            "Opcion contestada",
            "Area responsable",
            "Comentario",
            "Capturado por",
            "Actualizado",
        ]
    )
    questionnaire_header_row = questionnaire_sheet.max_row
    for axis in detail["axes"]:
        for question in axis["questions"]:
            questionnaire_sheet.append(
                [
                    axis["clave"],
                    axis["nombre"],
                    question["codigo"],
                    question["question"],
                    question["selected_level"],
                    question["selected_option"],
                    question["area_name"],
                    question["comment"],
                    question["captured_by"],
                    question["updated_at_label"],
                ]
            )
    _style_excel_metrics_table(questionnaire_sheet, header_rows={questionnaire_header_row})
    _set_excel_widths(
        questionnaire_sheet,
        {"A": 10, "B": 24, "C": 14, "D": 54, "E": 18, "F": 32, "G": 24, "H": 40, "I": 24, "J": 18},
    )

    evidence_sheet = workbook.create_sheet("Evidencias")
    _write_excel_title(evidence_sheet, "Evidencias por eje", subtitle, end_column=6)
    evidence_sheet.append([])
    evidence_sheet.append(["Clave", "Eje", "Evidencia", "Version", "Tipo", "Fecha"])
    evidence_header_row = evidence_sheet.max_row
    has_evidence = False
    for axis in detail["axes"]:
        for evidence in axis["evidence"]:
            has_evidence = True
            evidence_sheet.append(
                [
                    axis["clave"],
                    axis["nombre"],
                    evidence["name"],
                    evidence["version"],
                    evidence["mime_type"],
                    evidence["created_at_label"],
                ]
            )
    if not has_evidence:
        evidence_sheet.append(["", "", "Sin evidencias activas registradas", "", "", ""])
    _style_excel_metrics_table(evidence_sheet, header_rows={evidence_header_row})
    _set_excel_widths(evidence_sheet, {"A": 10, "B": 24, "C": 36, "D": 10, "E": 24, "F": 18})

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_detail_word(detail: dict, *, title: str, subtitle_lines: list[str], preliminary_message: str | None = None) -> BytesIO:
    body: list[str] = [
        _docx_paragraph(title, bold=True, size=34, spacing_after=120),
    ]
    for line in subtitle_lines:
        body.append(_docx_paragraph(line, size=20, color=WORD_THEME["muted"], spacing_after=40))

    if preliminary_message and detail["summary"].get("is_preliminary"):
        body.append(
            _docx_paragraph(
                preliminary_message,
                bold=True,
                size=20,
                color=WORD_THEME["medium"],
                spacing_before=60,
                spacing_after=100,
            )
        )

    body.append(_docx_paragraph("Indice de madurez", bold=True, size=26, spacing_before=60))
    body.append(
        _docx_table(
            [
                _docx_header_row(["Indice global", "Nivel", "Avance", "Ejes criticos"]),
                [
                    str(detail["summary"]["index_score"]),
                    detail["summary"]["level_label"],
                    f"{detail['summary']['completion']}%",
                    str(len(detail["summary"]["critical_axes"])),
                ],
            ],
            col_widths=[2200, 2200, 1800, 2200],
        )
    )
    body.append(_docx_paragraph("", size=4, spacing_after=80))

    body.append(_docx_paragraph("Guia de interpretacion", bold=True, size=24))
    guide_rows = [_docx_header_row(["Rango", "Nivel"])]
    for guide in detail["score_guide"]:
        guide_rows.append(
            [
                {"text": guide["range"], "fill": _theme_color_for_slug(guide["slug"]), "bold": True, "color": WORD_THEME["white"]},
                {"text": guide["label"], "fill": _theme_color_for_slug(guide["slug"]), "bold": True, "color": WORD_THEME["white"]},
            ]
        )
    body.append(_docx_table(guide_rows, col_widths=[3400, 2200]))
    body.append(_docx_paragraph("", size=4, spacing_after=80))

    body.append(_docx_paragraph("Mapa de brechas", bold=True, size=24))
    heatmap_rows = [_docx_header_row(["Eje", "Bajo", "Medio", "Alto", "Optimo", "Promedio", "Brecha"])]
    for row in detail["heatmap_rows"]:
        cells: list[dict | str] = [f"{row['axis_key']} {row['axis_name']}"]
        for slug in ("low", "medium", "high", "optimal"):
            if row["selected_slug"] == slug:
                cells.append({"text": "OK", "fill": _theme_color_for_slug(slug), "bold": True, "color": WORD_THEME["white"]})
            else:
                cells.append("")
        cells.extend([str(row["average"]), str(row["gap"])])
        heatmap_rows.append(cells)
    body.append(_docx_table(heatmap_rows, col_widths=[3000, 1050, 1050, 1050, 1200, 1300, 1300]))
    body.append(_docx_paragraph("", size=4, spacing_after=80))

    body.append(_docx_paragraph("Sintesis por eje", bold=True, size=24))
    axis_rows = [
        _docx_header_row(
            ["Clave", "Eje", "Promedio", "Brecha", "Prioridad", "Avance", "Comentario", "Evidencias"]
        )
    ]
    for axis in detail["axes"]:
        axis_rows.append(
            [
                axis["clave"],
                axis["nombre"],
                str(axis["summary"]["promedio"]),
                str(axis["summary"]["brecha"]),
                axis["summary"]["prioridad"],
                f"{axis['summary']['progreso']}%",
                axis["axis_comment"],
                str(len(axis["evidence"])),
            ]
        )
    body.append(_docx_table(axis_rows, col_widths=[900, 2200, 1100, 1100, 1700, 1200, 3400, 1200]))

    if detail["axes"]:
        body.append(_docx_page_break())
        body.append(_docx_paragraph("Cuestionario respondido", bold=True, size=26))
        for axis_index, axis in enumerate(detail["axes"]):
            if axis_index > 0:
                body.append(_docx_page_break())
            body.append(_docx_paragraph(f"{axis['clave']} {axis['nombre']}", bold=True, size=24, color=WORD_THEME["sage"]))
            body.append(
                _docx_paragraph(
                    f"Promedio {axis['summary']['promedio']} | Brecha {axis['summary']['brecha']} | "
                    f"Avance {axis['summary']['progreso']}% | Prioridad {axis['summary']['prioridad']}",
                    size=18,
                    color=WORD_THEME["muted"],
                    spacing_after=60,
                )
            )
            body.append(_docx_paragraph(f"Comentario del eje: {axis['axis_comment']}", size=18, spacing_after=40))
            body.append(
                _docx_paragraph(
                    f"Responsable: {axis['axis_comment_author']} | Actualizado: {axis['axis_comment_updated_at']}",
                    size=18,
                    color=WORD_THEME["muted"],
                    spacing_after=60,
                )
            )
            if axis["evidence"]:
                evidence_text = "; ".join(f"{item['name']} (v{item['version']})" for item in axis["evidence"])
            else:
                evidence_text = "Sin evidencias activas registradas"
            body.append(_docx_paragraph(f"Evidencias activas: {evidence_text}", size=18, spacing_after=60))

            for question in axis["questions"]:
                body.append(_docx_paragraph(f"{question['codigo']} - {question['question']}", bold=True, size=20, spacing_after=40))
                option_row = []
                for option in question["options"]:
                    option_row.append(
                        {
                            "text": f"Nivel {option['value']}\n{option['label']}",
                            "fill": WORD_THEME["sage_soft"] if option["selected"] else None,
                            "bold": option["selected"],
                        }
                    )
                body.append(_docx_table([option_row], col_widths=[2000, 2000, 2000, 2000]))
                body.append(
                    _docx_table(
                        [
                            _docx_label_value_row("Nivel seleccionado", question["selected_level"]),
                            _docx_label_value_row("Opcion contestada", question["selected_option"]),
                            _docx_label_value_row("Area responsable", question["area_name"]),
                            _docx_label_value_row("Comentario", question["comment"]),
                            _docx_label_value_row(
                                "Capturado por",
                                f"{question['captured_by']} | Ultima actualizacion: {question['updated_at_label']}",
                            ),
                        ],
                        col_widths=[2400, 7200],
                    )
                )
                body.append(_docx_paragraph("", size=4, spacing_after=60))

    return _package_docx("".join(body), title=title)


def _pdf_styles() -> dict[str, ParagraphStyle]:
    styles = getSampleStyleSheet()
    return {
        "cover_kicker": ParagraphStyle(
            "CoverKicker",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=13,
            alignment=TA_CENTER,
            textColor=PDF_THEME["sage"],
            spaceAfter=8,
        ),
        "cover_title": ParagraphStyle(
            "CoverTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=36,
            leading=40,
            alignment=TA_CENTER,
            textColor=PDF_THEME["navy"],
            spaceAfter=14,
        ),
        "cover_subject": ParagraphStyle(
            "CoverSubject",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=28,
            leading=32,
            alignment=TA_CENTER,
            textColor=PDF_THEME["navy"],
            spaceAfter=10,
        ),
        "cover_questionnaire": ParagraphStyle(
            "CoverQuestionnaire",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            alignment=TA_CENTER,
            textColor=PDF_THEME["sage"],
            spaceAfter=12,
        ),
        "cover_meta": ParagraphStyle(
            "CoverMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=11.2,
            leading=14.2,
            alignment=TA_CENTER,
            textColor=PDF_THEME["muted"],
            spaceAfter=4,
        ),
        "cover_warning": ParagraphStyle(
            "CoverWarning",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10.4,
            leading=13.2,
            alignment=TA_CENTER,
            textColor=PDF_THEME["medium"],
            backColor=colors.HexColor("#F8E0C5"),
            borderPadding=10,
        ),
        "title": ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=PDF_THEME["navy"],
            spaceAfter=4,
        ),
        "subtitle": ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            textColor=PDF_THEME["muted"],
            spaceAfter=2,
        ),
        "section": ParagraphStyle(
            "SectionTitle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=15,
            textColor=PDF_THEME["navy"],
            spaceAfter=6,
        ),
        "axis": ParagraphStyle(
            "AxisTitle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=14,
            textColor=PDF_THEME["sage"],
            spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "BodyCopy",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.2,
            leading=12.2,
            textColor=PDF_THEME["navy"],
        ),
        "body_muted": ParagraphStyle(
            "BodyMuted",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=PDF_THEME["muted"],
        ),
        "small": ParagraphStyle(
            "BodySmall",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=PDF_THEME["navy"],
        ),
        "small_center": ParagraphStyle(
            "BodySmallCenter",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            textColor=PDF_THEME["navy"],
        ),
        "warning": ParagraphStyle(
            "Warning",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9.5,
            leading=12,
            textColor=PDF_THEME["medium"],
            backColor=colors.HexColor("#F8E0C5"),
            borderPadding=8,
        ),
    }


def _build_pdf_cover_story(
    detail: dict,
    *,
    title: str,
    subtitle_lines: list[str],
    styles: dict[str, ParagraphStyle],
    preliminary_message: str | None,
    cover_highlights: list[str],
    cover_detail_lines: list[str],
) -> list:
    story: list = [
        Spacer(1, 2.45 * inch),
        Paragraph("Diagnóstico Integral Municipal", styles["cover_kicker"]),
        Paragraph("Reporte Ejecutivo", styles["cover_title"]),
    ]
    if cover_highlights:
        story.append(Paragraph(_paragraph_escape(cover_highlights[0]), styles["cover_subject"]))
    if len(cover_highlights) > 1:
        story.append(Paragraph(_paragraph_escape(cover_highlights[1]), styles["cover_questionnaire"]))
    for line in cover_detail_lines:
        story.append(Paragraph(_paragraph_escape(line), styles["cover_meta"]))

    if preliminary_message and detail["summary"].get("is_preliminary"):
        story.extend(
            [
                Spacer(1, 0.32 * inch),
                Paragraph(_paragraph_escape(preliminary_message), styles["cover_warning"]),
            ]
        )

    story.extend([Spacer(1, 3.1 * inch), PageBreak()])
    return story


def _build_pdf_cover_metadata_table(detail: dict, styles: dict[str, ParagraphStyle]) -> Table:
    label_style = ParagraphStyle(
        "CoverLabel",
        parent=styles["small"],
        fontName="Helvetica-Bold",
        fontSize=10.2,
        leading=12.2,
        textColor=PDF_THEME["navy"],
    )
    value_style = ParagraphStyle(
        "CoverValue",
        parent=styles["small"],
        fontSize=11,
        leading=13,
        textColor=PDF_THEME["navy"],
    )
    rows = [
        [
            Paragraph("Cuestionario", label_style),
            Paragraph(_paragraph_escape(detail["questionnaire_name"]), value_style),
            Paragraph("Estado", label_style),
            Paragraph(_paragraph_escape(detail["summary"]["state_label"]), value_style),
        ],
        [
            Paragraph("Ejes", label_style),
            Paragraph(str(detail["total_axes"]), value_style),
            Paragraph("Evidencias", label_style),
            Paragraph(str(detail["total_evidence"]), value_style),
        ],
        [
            Paragraph("Nivel", label_style),
            Paragraph(_paragraph_escape(detail["summary"]["level_label"]), value_style),
            Paragraph("Avance", label_style),
            Paragraph(f"{detail['summary']['completion']}%", value_style),
        ],
    ]
    table = Table(rows, colWidths=[1.15 * inch, 2.25 * inch, 1.15 * inch, 1.95 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BACKGROUND", (0, 0), (0, -1), PDF_THEME["navy_soft"]),
                ("BACKGROUND", (2, 0), (2, -1), PDF_THEME["navy_soft"]),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return table


def _build_pdf_summary_table(summary: dict) -> Table:
    table = Table(
        [
            ["Índice global", "Nivel", "Avance", "Ejes críticos"],
            [
                str(summary["index_score"]),
                summary["level_label"],
                f"{summary['completion']}%",
                str(len(summary["critical_axes"])),
            ],
        ],
        colWidths=[1.45 * inch, 1.6 * inch, 1.25 * inch, 1.45 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), PDF_THEME["navy_soft"]),
                ("TEXTCOLOR", (0, 1), (-1, 1), PDF_THEME["navy"]),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.6, PDF_THEME["line"]),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _build_pdf_overview_panel(detail: dict, styles: dict[str, ParagraphStyle]) -> Table:
    legend_rows = [[Paragraph("<b>Rango</b>", styles["small_center"]), Paragraph("<b>Nivel</b>", styles["small_center"])]]
    for guide in detail["score_guide"]:
        legend_rows.append(
            [
                Paragraph(guide["range"], styles["small_center"]),
                Paragraph(guide["label"], styles["small_center"]),
            ]
        )
    legend_table = Table(legend_rows, colWidths=[1.15 * inch, 1.1 * inch])
    legend_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F7D8D8")),
                ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#F8E0C5")),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#FAEFC8")),
                ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#DCEDDA")),
                ("TEXTCOLOR", (0, 1), (-1, -1), PDF_THEME["navy"]),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    layout = Table(
        [[legend_table, _build_pdf_radar_chart(detail["axes"])]],
        colWidths=[2.5 * inch, 3.8 * inch],
    )
    layout.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOX", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return layout


def _build_pdf_heatmap_table(detail: dict, styles: dict[str, ParagraphStyle]) -> Table:
    header_style = ParagraphStyle(
        "HeatmapHeader",
        parent=styles["small_center"],
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    axis_style = ParagraphStyle("HeatmapAxis", parent=styles["small"], fontSize=8.2, leading=10)
    metric_style = ParagraphStyle("HeatmapMetric", parent=styles["small_center"], fontSize=8.2, leading=10)
    rows = [[
        Paragraph("<b>Eje</b>", header_style),
        Paragraph("<b>Bajo</b>", header_style),
        Paragraph("<b>Medio</b>", header_style),
        Paragraph("<b>Alto</b>", header_style),
        Paragraph("<b>Óptimo</b>", header_style),
        Paragraph("<b>Promedio</b>", header_style),
        Paragraph("<b>Brecha</b>", header_style),
    ]]
    maturity_column = {"low": 1, "medium": 2, "high": 3, "optimal": 4}
    for axis in detail["axes"]:
        row = [
            Paragraph(f"{axis['clave']} {axis['nombre']}", axis_style),
            "",
            "",
            "",
            "",
            Paragraph(str(axis["summary"]["promedio"]), metric_style),
            Paragraph(str(axis["summary"]["brecha"]), metric_style),
        ]
        row[maturity_column[axis["summary"]["maturity_slug"]]] = Paragraph("<b>OK</b>", metric_style)
        rows.append(row)
    table = Table(rows, repeatRows=1, colWidths=[2.45 * inch, 0.65 * inch, 0.72 * inch, 0.65 * inch, 0.78 * inch, 0.8 * inch, 0.75 * inch])
    styles_table = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
        ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 1), (4, -1), "CENTER"),
        ("ALIGN", (5, 1), (-1, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFBFC")]),
    ]
    for row_index, axis in enumerate(detail["axes"], start=1):
        column_index = maturity_column[axis["summary"]["maturity_slug"]]
        styles_table.extend(
            [
                ("BACKGROUND", (column_index, row_index), (column_index, row_index), PDF_THEME[axis["summary"]["maturity_slug"]]),
                ("TEXTCOLOR", (column_index, row_index), (column_index, row_index), colors.white),
                ("FONTNAME", (column_index, row_index), (column_index, row_index), "Helvetica-Bold"),
            ]
        )
    table.setStyle(TableStyle(styles_table))
    return table


def _build_pdf_axis_summary_table(detail: dict, styles: dict[str, ParagraphStyle]) -> Table:
    header_style = ParagraphStyle(
        "AxisSummaryHeader",
        parent=styles["small_center"],
        fontName="Helvetica-Bold",
        textColor=PDF_THEME["navy"],
        fontSize=8.1,
        leading=9.5,
    )
    cell_style = ParagraphStyle(
        "AxisSummaryCell",
        parent=styles["small"],
        fontSize=8.1,
        leading=9.8,
    )
    center_style = ParagraphStyle(
        "AxisSummaryCenter",
        parent=styles["small_center"],
        fontSize=8.1,
        leading=9.8,
    )
    rows = [[
        Paragraph("<b>Clave</b>", header_style),
        Paragraph("<b>Eje</b>", header_style),
        Paragraph("<b>Promedio</b>", header_style),
        Paragraph("<b>Brecha</b>", header_style),
        Paragraph("<b>Prioridad</b>", header_style),
        Paragraph("<b>Avance</b>", header_style),
        Paragraph("<b>Evidencias</b>", header_style),
    ]]
    for axis in detail["axes"]:
        rows.append(
            [
                Paragraph(axis["clave"], center_style),
                Paragraph(axis["nombre"], cell_style),
                Paragraph(str(axis["summary"]["promedio"]), center_style),
                Paragraph(str(axis["summary"]["brecha"]), center_style),
                Paragraph(axis["summary"]["prioridad"], cell_style),
                Paragraph(f"{axis['summary']['progreso']}%", center_style),
                Paragraph(str(len(axis["evidence"])), center_style),
            ]
        )
    table = Table(
        rows,
        repeatRows=1,
        colWidths=[0.6 * inch, 2.55 * inch, 0.75 * inch, 0.72 * inch, 1.15 * inch, 0.72 * inch, 0.61 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["sage_soft"]),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFBFC")]),
            ]
        )
    )
    return table


def _build_pdf_axis_story(axis: dict, styles: dict[str, ParagraphStyle]) -> list:
    summary = axis["summary"]
    story = [
        Paragraph(f"{axis['clave']} {axis['nombre']}", styles["axis"]),
        _build_pdf_axis_metrics_table(axis),
        Spacer(1, 0.08 * inch),
        Paragraph(f"<b>Comentario del eje:</b> {_paragraph_escape(axis['axis_comment'])}", styles["body"]),
        Paragraph(
            f"<b>Responsable del eje:</b> {_paragraph_escape(axis['axis_comment_area'])}",
            styles["body_muted"],
        ),
        Paragraph(
            f"<b>Prioridad:</b> {_paragraph_escape(summary['prioridad'])} | <b>Madurez:</b> {_paragraph_escape(summary['madurez'])} | <b>Respondidos:</b> {summary['respondidos']}/{summary['total']}",
            styles["body_muted"],
        ),
        Spacer(1, 0.1 * inch),
    ]

    if axis["evidence"]:
        evidence_text = "; ".join(f"{item['name']} (v{item['version']})" for item in axis["evidence"])
    else:
        evidence_text = "Sin evidencias activas registradas"
    story.append(Paragraph(f"<b>Evidencias activas:</b> {_paragraph_escape(evidence_text)}", styles["body"]))
    story.append(Spacer(1, 0.1 * inch))

    for question in axis["questions"]:
        story.append(_build_pdf_question_block(question, styles))
        story.append(Spacer(1, 0.12 * inch))
    return story


def _build_pdf_axis_metrics_table(axis: dict) -> Table:
    summary = axis["summary"]
    table = Table(
        [
            ["Promedio", "Brecha", "Avance", "Evidencias"],
            [
                str(summary["promedio"]),
                str(summary["brecha"]),
                f"{summary['progreso']}%",
                str(len(axis["evidence"])),
            ],
        ],
        colWidths=[1.2 * inch, 1.2 * inch, 1.2 * inch, 1.35 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy_soft"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _build_pdf_question_block(question: dict, styles: dict[str, ParagraphStyle]) -> KeepTogether:
    option_cells = [
        Paragraph(
            f"<b>Nivel {option['value']}</b><br/>{_paragraph_escape(option['label'])}",
            styles["small_center"],
        )
        for option in question["options"]
    ]
    option_table = Table([option_cells], colWidths=[1.47 * inch, 1.47 * inch, 1.47 * inch, 1.47 * inch])
    option_styles = [
        ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]
    for column_index, option in enumerate(question["options"]):
        if option["selected"]:
            option_styles.extend(
                [
                    ("BACKGROUND", (column_index, 0), (column_index, 0), PDF_THEME["sage_soft"]),
                    ("FONTNAME", (column_index, 0), (column_index, 0), "Helvetica-Bold"),
                    ("LINEBELOW", (column_index, 0), (column_index, 0), 1.4, PDF_THEME["sage"]),
                ]
            )
    option_table.setStyle(TableStyle(option_styles))

    meta_table = Table(
        [
            ["Nivel seleccionado", question["selected_level"]],
            ["Opción contestada", question["selected_option"]],
            ["Área responsable", question["area_name"]],
            ["Comentario", question["comment"]],
            ["Capturado por", question["captured_by"]],
        ],
        colWidths=[1.4 * inch, 4.45 * inch],
    )
    meta_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), PDF_THEME["navy_soft"]),
                ("TEXTCOLOR", (0, 0), (0, -1), PDF_THEME["navy"]),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    block = [
        Paragraph(f"<b>{_paragraph_escape(question['codigo'])}</b> - {_paragraph_escape(question['question'])}", styles["body"]),
        Spacer(1, 0.06 * inch),
        option_table,
        Spacer(1, 0.06 * inch),
        meta_table,
    ]
    return KeepTogether(block)


def _build_pdf_radar_chart(axes: list[dict]) -> Drawing:
    width = 265
    height = 230
    drawing = Drawing(width, height)
    center_x = width / 2
    center_y = height / 2 - 8
    radius = 74
    total_axes = max(len(axes), 1)
    labels = [axis["clave"] or axis["nombre"] for axis in axes]
    values = [axis["summary"]["promedio"] for axis in axes]
    angles = [math.pi / 2 - (2 * math.pi * index / total_axes) for index in range(total_axes)]

    for level in range(1, 4):
        scale = level / 3
        points: list[float] = []
        for angle in angles:
            x = center_x + math.cos(angle) * radius * scale
            y = center_y + math.sin(angle) * radius * scale
            points.extend([x, y])
        drawing.add(Polygon(points, fillColor=None, strokeColor=colors.HexColor("#C6D1D8"), strokeWidth=1))

    for angle in angles:
        x = center_x + math.cos(angle) * radius
        y = center_y + math.sin(angle) * radius
        drawing.add(Line(center_x, center_y, x, y, strokeColor=colors.HexColor("#C6D1D8"), strokeWidth=1))

    for level in range(1, 4):
        label_y = center_y + radius * (level / 3)
        drawing.add(
            String(
                center_x + 6,
                label_y - 3,
                str(level),
                fontName="Helvetica",
                fontSize=7,
                fillColor=PDF_THEME["muted"],
            )
        )

    if values:
        data_points: list[float] = []
        for angle, value in zip(angles, values, strict=False):
            scale = min(max(value, 0), 3) / 3
            x = center_x + math.cos(angle) * radius * scale
            y = center_y + math.sin(angle) * radius * scale
            data_points.extend([x, y])
        drawing.add(
            Polygon(
                data_points,
                fillColor=colors.Color(0.447, 0.545, 0.314, alpha=0.18),
                strokeColor=PDF_THEME["medium"],
                strokeWidth=2,
            )
        )
        for x, y in zip(data_points[0::2], data_points[1::2], strict=False):
            drawing.add(Circle(x, y, 3, fillColor=PDF_THEME["medium"], strokeColor=colors.white, strokeWidth=0.8))

    for angle, label in zip(angles, labels, strict=False):
        label_radius = radius + 18
        x = center_x + math.cos(angle) * label_radius
        y = center_y + math.sin(angle) * label_radius
        anchor = "middle"
        if math.cos(angle) > 0.25:
            anchor = "start"
        elif math.cos(angle) < -0.25:
            anchor = "end"
        drawing.add(
            String(
                x,
                y,
                label,
                fontName="Helvetica-Bold",
                fontSize=8,
                fillColor=PDF_THEME["navy"],
                textAnchor=anchor,
            )
        )

    drawing.add(
        String(
            center_x,
            8,
            "Radar institucional por eje",
            fontName="Helvetica",
            fontSize=7.5,
            fillColor=PDF_THEME["muted"],
            textAnchor="middle",
        )
    )
    return drawing


def _draw_pdf_chrome(canvas, doc) -> None:
    canvas.saveState()
    width, height = doc.pagesize
    canvas.setFillColor(PDF_THEME["sage"])
    canvas.rect(doc.leftMargin, height - 0.42 * inch, width - doc.leftMargin - doc.rightMargin, 0.1 * inch, fill=1, stroke=0)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(PDF_THEME["muted"])
    canvas.drawString(doc.leftMargin, 0.32 * inch, "Diagnóstico Integral Municipal")
    canvas.drawCentredString(width / 2, 0.32 * inch, PDF_FOOTER_LEGEND)
    canvas.drawRightString(width - doc.rightMargin, 0.32 * inch, f"Página {canvas.getPageNumber()}")
    canvas.restoreState()


def _write_excel_title(sheet, title: str, subtitle: str, *, end_column: int) -> None:
    end_letter = get_column_letter(end_column)
    sheet.merge_cells(f"A1:{end_letter}1")
    sheet.merge_cells(f"A2:{end_letter}2")
    title_cell = sheet["A1"]
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color=EXCEL_THEME["navy"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    subtitle_cell = sheet["A2"]
    subtitle_cell.value = subtitle
    subtitle_cell.font = Font(size=10, color=EXCEL_THEME["muted"])
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_excel_metrics_table(sheet, *, header_rows: set[int]) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row in (1, 2):
                continue
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT
            if cell.row in header_rows:
                cell.fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
                cell.font = Font(bold=True, color=EXCEL_THEME["white"])
                cell.alignment = CENTER_ALIGNMENT
            elif cell.value is not None:
                cell.font = Font(color=EXCEL_THEME["navy"])


def _paint_excel_heatmap_row(sheet, row_index: int, maturity_slug: str) -> None:
    color_map = {
        "low": EXCEL_THEME["low"],
        "medium": EXCEL_THEME["medium"],
        "high": EXCEL_THEME["high"],
        "optimal": EXCEL_THEME["optimal"],
    }
    column_map = {"low": 3, "medium": 4, "high": 5, "optimal": 6}
    column_index = column_map[maturity_slug]
    cell = sheet.cell(row=row_index, column=column_index)
    cell.fill = PatternFill("solid", fgColor=color_map[maturity_slug])
    cell.font = Font(bold=True, color=EXCEL_THEME["white"])
    cell.alignment = CENTER_ALIGNMENT


def _set_excel_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _docx_paragraph(
    text: str,
    *,
    bold: bool = False,
    size: int = 22,
    color: str = WORD_THEME["navy"],
    spacing_before: int = 0,
    spacing_after: int = 80,
    align: str | None = None,
) -> str:
    alignment = f'<w:jc w:val="{align}"/>' if align else ""
    return (
        "<w:p>"
        f'<w:pPr>{alignment}<w:spacing w:before="{spacing_before}" w:after="{spacing_after}"/></w:pPr>'
        f"{_docx_run(text, bold=bold, size=size, color=color)}"
        "</w:p>"
    )


def _docx_run(text: str, *, bold: bool = False, size: int = 22, color: str = WORD_THEME["navy"]) -> str:
    run_properties = ["<w:rPr>"]
    if bold:
        run_properties.append("<w:b/>")
    if size:
        run_properties.append(f'<w:sz w:val="{size}"/>')
    if color:
        run_properties.append(f'<w:color w:val="{color}"/>')
    run_properties.append("</w:rPr>")
    rpr = "".join(run_properties)
    parts = str(text or "").split("\n")
    runs = []
    for index, part in enumerate(parts):
        runs.append(f'<w:r>{rpr}<w:t xml:space="preserve">{xml_escape(part)}</w:t></w:r>')
        if index < len(parts) - 1:
            runs.append("<w:r><w:br/></w:r>")
    return "".join(runs) if runs else f'<w:r>{rpr}<w:t xml:space="preserve"></w:t></w:r>'


def _docx_table(rows: list[list], *, col_widths: list[int] | None = None) -> str:
    table_parts = [
        "<w:tbl>",
        "<w:tblPr>",
        '<w:tblW w:w="0" w:type="auto"/>',
        "<w:tblBorders>",
        '<w:top w:val="single" w:sz="6" w:color="D9E0E5"/>',
        '<w:left w:val="single" w:sz="6" w:color="D9E0E5"/>',
        '<w:bottom w:val="single" w:sz="6" w:color="D9E0E5"/>',
        '<w:right w:val="single" w:sz="6" w:color="D9E0E5"/>',
        '<w:insideH w:val="single" w:sz="6" w:color="D9E0E5"/>',
        '<w:insideV w:val="single" w:sz="6" w:color="D9E0E5"/>',
        "</w:tblBorders>",
        "</w:tblPr>",
    ]
    if col_widths:
        table_parts.append("<w:tblGrid>")
        for width in col_widths:
            table_parts.append(f'<w:gridCol w:w="{width}"/>')
        table_parts.append("</w:tblGrid>")

    for row in rows:
        table_parts.append("<w:tr>")
        for index, cell in enumerate(row):
            if isinstance(cell, dict):
                text = str(cell.get("text", ""))
                fill = cell.get("fill")
                bold = bool(cell.get("bold"))
                color = cell.get("color", WORD_THEME["navy"])
            else:
                text = str(cell)
                fill = None
                bold = False
                color = WORD_THEME["navy"]
            width = col_widths[index] if col_widths and index < len(col_widths) else None
            tc_properties = ["<w:tcPr>"]
            if width:
                tc_properties.append(f'<w:tcW w:w="{width}" w:type="dxa"/>')
            if fill:
                tc_properties.append(f'<w:shd w:val="clear" w:color="auto" w:fill="{fill}"/>')
            tc_properties.append('<w:vAlign w:val="center"/>')
            tc_properties.append("</w:tcPr>")
            table_parts.append("<w:tc>")
            table_parts.append("".join(tc_properties))
            table_parts.append(_docx_paragraph(text, bold=bold, size=20, color=color, spacing_after=0))
            table_parts.append("</w:tc>")
        table_parts.append("</w:tr>")
    table_parts.append("</w:tbl>")
    return "".join(table_parts)


def _docx_page_break() -> str:
    return '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'


def _docx_header_row(values: list[str]) -> list[dict]:
    return [
        {"text": value, "fill": WORD_THEME["navy"], "bold": True, "color": WORD_THEME["white"]}
        for value in values
    ]


def _docx_label_value_row(label: str, value: str) -> list[dict | str]:
    return [
        {"text": label, "fill": WORD_THEME["sage_soft"], "bold": True},
        value,
    ]


def _package_docx(body_xml: str, *, title: str) -> BytesIO:
    created_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        f"{body_xml}"
        "<w:sectPr>"
        '<w:pgSz w:w="12240" w:h="15840"/>'
        '<w:pgMar w:top="1080" w:right="900" w:bottom="900" w:left="900" w:header="720" w:footer="720" w:gutter="0"/>'
        "</w:sectPr>"
        "</w:body>"
        "</w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        "</Types>"
    )
    relationships = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        "</Relationships>"
    )
    core_properties = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        f"<dc:title>{xml_escape(title)}</dc:title>"
        "<dc:creator>Diagnostico Integral Municipal</dc:creator>"
        "<cp:lastModifiedBy>Diagnostico Integral Municipal</cp:lastModifiedBy>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{created_at}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{created_at}</dcterms:modified>'
        "</cp:coreProperties>"
    )
    app_properties = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        "<Application>Diagnostico Integral Municipal</Application>"
        "</Properties>"
    )

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", relationships)
        archive.writestr("docProps/core.xml", core_properties)
        archive.writestr("docProps/app.xml", app_properties)
        archive.writestr("word/document.xml", document_xml)
    buffer.seek(0)
    return buffer


def _theme_color_for_slug(slug: str) -> str:
    return {
        "low": EXCEL_THEME["low"],
        "medium": EXCEL_THEME["medium"],
        "high": EXCEL_THEME["high"],
        "optimal": EXCEL_THEME["optimal"],
    }.get(slug, EXCEL_THEME["navy"])


def _paragraph_escape(value: str) -> str:
    return xml_escape(str(value or "")).replace("\n", "<br/>")

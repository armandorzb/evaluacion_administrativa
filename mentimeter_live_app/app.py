from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict, deque
from datetime import datetime
from functools import wraps
from io import BytesIO, StringIO
from math import ceil
import os
from pathlib import Path
from random import randint
import re
from secrets import compare_digest, token_urlsafe
from time import monotonic
from typing import Any

import qrcode
from flask import (
    Flask,
    abort,
    current_app,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    session as flask_session,
    url_for,
)
from flask_socketio import SocketIO, emit, join_room
from sqlalchemy import inspect
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

try:
    from .models import db, Option, Participant, Question, Response, Session, SessionRun, utcnow
except ImportError:  # Allows `python app.py` from this folder.
    from models import db, Option, Participant, Question, Response, Session, SessionRun, utcnow


socketio = SocketIO(
    async_mode=os.getenv("MENTI_SOCKETIO_ASYNC_MODE") or None,
    cors_allowed_origins="*",
    manage_session=False,
)
# Contract shared by backend validation and the two vanilla JS frontends.
QUESTION_TYPES = {"content_slide", "multiple_choice", "word_cloud", "scale", "open_text", "ranking", "quiz"}
SINGLE_RESPONSE_TYPES = {"multiple_choice", "scale", "ranking", "quiz"}
PARTICIPANT_COOKIE = "menti_participant_token"
RESULT_LAYOUTS = {"auto", "chart", "list", "grid", "cloud", "cards", "leaderboard"}
LAYOUT_BLOCK_IDS = ("question", "activity", "results")
DEFAULT_LAYOUT_BLOCKS = {
    "question": {"id": "question", "x": 7, "y": 12, "w": 86, "h": 25, "z": 1},
    "activity": {"id": "activity", "x": 7, "y": 42, "w": 42, "h": 43, "z": 2},
    "results": {"id": "results", "x": 53, "y": 42, "w": 40, "h": 43, "z": 3},
}
TEXT_BOX_HEX_RE = re.compile(r"^#[0-9a-fA-F]{6}$")
TEXT_BOX_MAX_ITEMS = 24
TEXT_BOX_MAX_TEXT = 1200
TEXT_BOX_TOTAL_TEXT = 3000
socket_participants: dict[str, tuple[int, str]] = {}
public_rate_buckets: dict[str, deque[float]] = {}


def default_asset_version() -> str:
    app_root = Path(__file__).resolve().parent
    candidates = [
        Path(__file__),
        app_root / "static" / "css" / "app.css",
        app_root / "static" / "js" / "admin.js",
        app_root / "static" / "js" / "audience.js",
        app_root / "templates" / "admin.html",
        app_root / "templates" / "admin_login.html",
        app_root / "templates" / "audience.html",
        app_root / "templates" / "join.html",
    ]
    mtimes = [path.stat().st_mtime for path in candidates if path.exists()]
    return str(int(max(mtimes, default=0)))


def create_app(test_config: dict[str, Any] | None = None) -> Flask:
    app = Flask(__name__, instance_relative_config=True)
    database_url = os.getenv("DATABASE_URL")
    local_database = Path(os.getenv("MENTI_DB_PATH", str(Path(app.instance_path) / "mentimeter.sqlite3")))
    app.config.update(
        SECRET_KEY=os.getenv("SECRET_KEY", "dev-mentimeter-local"),
        SQLALCHEMY_DATABASE_URI=database_url or f"sqlite:///{local_database}",
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
        MAX_CONTENT_LENGTH=int(os.getenv("MENTI_MAX_CONTENT_LENGTH", str(1024 * 1024))),
        MENTI_SEED_DEMO=os.getenv("MENTI_SEED_DEMO", "true").lower() in {"1", "true", "yes", "si"},
        MENTI_SOCKETIO_CORS=os.getenv("MENTI_SOCKETIO_CORS", "*"),
        MENTI_ADMIN_PIN=os.getenv("MENTI_ADMIN_PIN"),
        MENTI_ADMIN_USERNAME=os.getenv("MENTI_ADMIN_USERNAME"),
        MENTI_ADMIN_PASSWORD=os.getenv("MENTI_ADMIN_PASSWORD"),
        MENTI_RESPONSE_RATE_LIMIT=int(os.getenv("MENTI_RESPONSE_RATE_LIMIT", "120")),
        MENTI_RESPONSE_RATE_WINDOW=int(os.getenv("MENTI_RESPONSE_RATE_WINDOW", "60")),
        ASSET_VERSION=os.getenv("MENTI_ASSET_VERSION") or default_asset_version(),
    )
    if test_config:
        app.config.update(test_config)

    Path(app.instance_path).mkdir(parents=True, exist_ok=True)
    if not database_url:
        local_database.parent.mkdir(parents=True, exist_ok=True)
    db.init_app(app)
    if os.getenv("MENTI_PROXY_FIX", "false").lower() in {"1", "true", "yes", "si"}:
        app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

    socketio.init_app(app, cors_allowed_origins=app.config["MENTI_SOCKETIO_CORS"])

    @app.context_processor
    def inject_asset_version() -> dict[str, str]:
        return {"asset_version": str(app.config.get("ASSET_VERSION") or "1")}

    register_routes(app)
    register_socket_events()

    with app.app_context():
        db.create_all()
        ensure_schema_compatibility()
        if app.config.get("MENTI_SEED_DEMO", True):
            ensure_demo_session()

    return app


def ensure_schema_compatibility() -> None:
    inspector = inspect(db.engine)
    if "responses" not in inspector.get_table_names():
        return
    response_columns = {column["name"] for column in inspector.get_columns("responses")}
    if "run_id" not in response_columns:
        with db.engine.begin() as connection:
            connection.exec_driver_sql("ALTER TABLE responses ADD COLUMN run_id INTEGER")
            connection.exec_driver_sql("CREATE INDEX IF NOT EXISTS ix_responses_run_id ON responses (run_id)")


def is_admin_authenticated() -> bool:
    return admin_auth_mode() == "open" or flask_session.get("menti_admin_ok") is True


def admin_auth_mode() -> str:
    username = current_app.config.get("MENTI_ADMIN_USERNAME")
    password = current_app.config.get("MENTI_ADMIN_PASSWORD")
    if username and password:
        return "password"
    if current_app.config.get("MENTI_ADMIN_PIN"):
        return "pin"
    return "open"


def validate_admin_login(form) -> bool:
    mode = admin_auth_mode()
    if mode == "password":
        expected_user = str(current_app.config["MENTI_ADMIN_USERNAME"])
        expected_password = str(current_app.config["MENTI_ADMIN_PASSWORD"])
        submitted_user = str(form.get("username", ""))
        submitted_password = str(form.get("password", ""))
        return compare_digest(submitted_user, expected_user) and compare_digest(submitted_password, expected_password)
    if mode == "pin":
        return compare_digest(str(form.get("pin", "")), str(current_app.config["MENTI_ADMIN_PIN"]))
    return True


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if is_admin_authenticated():
            return view(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "No autorizado."}), 401
        return redirect(url_for("admin_login", next=request.full_path))

    return wrapped


def safe_next_url(value: str | None) -> str:
    if value and value.startswith("/") and not value.startswith("//"):
        return value
    return url_for("admin")


def rate_limit_identity(scope: str, session: Session | None = None, participant: Participant | None = None) -> str:
    forwarded = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    remote_addr = forwarded or request.remote_addr or "unknown"
    session_key = session.code if session else "-"
    participant_key = participant.token if participant else "-"
    return f"{scope}:{session_key}:{participant_key}:{remote_addr}"


def consume_rate_limit(key: str) -> tuple[bool, int]:
    limit = max(1, int(current_app.config.get("MENTI_RESPONSE_RATE_LIMIT", 30)))
    window = max(1, int(current_app.config.get("MENTI_RESPONSE_RATE_WINDOW", 60)))
    now = monotonic()
    bucket = public_rate_buckets.setdefault(key, deque())
    while bucket and now - bucket[0] >= window:
        bucket.popleft()
    if len(bucket) >= limit:
        retry_after = max(1, ceil(window - (now - bucket[0])))
        return False, retry_after
    bucket.append(now)
    return True, 0


def rate_limited_response(retry_after: int):
    response = jsonify({"ok": False, "error": "Demasiadas respuestas. Intenta nuevamente en unos segundos."})
    response.status_code = 429
    response.headers["Retry-After"] = str(retry_after)
    return response


def socket_payload_too_large(payload: Any) -> bool:
    max_size = int(current_app.config.get("MAX_CONTENT_LENGTH") or 0)
    if not max_size:
        return False
    try:
        size = len(json.dumps(payload or {}, ensure_ascii=False).encode("utf-8"))
    except TypeError:
        size = max_size + 1
    return size > max_size


def register_routes(app: Flask) -> None:
    @app.before_request
    def reject_oversized_payload():
        max_size = int(current_app.config.get("MAX_CONTENT_LENGTH") or 0)
        if max_size and request.content_length and request.content_length > max_size:
            return jsonify({"ok": False, "error": "Payload demasiado grande."}), 413

    @app.get("/")
    def index():
        return redirect(url_for("join"))

    @app.get("/healthz")
    def healthz():
        return jsonify({"ok": True, "service": "mentimeter_live_app"})

    @app.route("/admin-login", methods=["GET", "POST"])
    def admin_login():
        mode = admin_auth_mode()
        if mode == "open":
            return redirect(url_for("admin"))
        if request.method == "POST":
            if validate_admin_login(request.form):
                flask_session["menti_admin_ok"] = True
                flask_session["menti_admin_mode"] = mode
                return redirect(safe_next_url(request.args.get("next")))
            return render_template("admin_login.html", auth_mode=mode, error="Credenciales incorrectas."), 401
        return render_template("admin_login.html", auth_mode=mode)

    @app.post("/admin-logout")
    def admin_logout():
        flask_session.pop("menti_admin_ok", None)
        return redirect(url_for("admin_login"))

    @app.get("/admin")
    @admin_required
    def admin():
        sessions = Session.query.order_by(Session.updated_at.desc()).all()
        selected = find_session(request.args.get("code")) if request.args.get("code") else (sessions[0] if sessions else None)
        return render_template(
            "admin.html",
            sessions=sessions,
            selected=selected,
            question_types=sorted(QUESTION_TYPES),
            admin_auth_required=admin_auth_mode() != "open",
        )

    @app.get("/present/<code>")
    @admin_required
    def present(code: str):
        session = find_session(code)
        if session is None:
            abort(404)
        return render_template(
            "admin.html",
            sessions=[session],
            selected=session,
            question_types=sorted(QUESTION_TYPES),
            present_only=True,
            admin_auth_required=admin_auth_mode() != "open",
        )

    @app.route("/join", methods=["GET", "POST"])
    def join():
        code = request.values.get("code", "").strip()
        if request.method == "POST" or code:
            session = find_session(code)
            if session is None:
                return render_template("join.html", error="No encontramos una sesión con ese código.", code=code), 404
            return redirect(url_for("audience", code=session.code))
        return render_template("join.html")

    @app.get("/s/<code>")
    def audience(code: str):
        session = find_session(code)
        if session is None:
            abort(404)
        participant = get_or_create_participant(session, request.cookies.get(PARTICIPANT_COOKIE))
        db.session.commit()
        response = make_response(render_template("audience.html", session=session, participant=participant))
        response.set_cookie(PARTICIPANT_COOKIE, participant.token, httponly=True, samesite="Lax")
        return response

    @app.get("/qr/<code>.png")
    def qr_png(code: str):
        session = find_session(code)
        if session is None:
            abort(404)
        qr_url = url_for("audience", code=session.code, _external=True)
        img = qrcode.make(qr_url)
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        return send_file(buffer, mimetype="image/png", download_name=f"join-{session.code}.png")

    @app.get("/api/sessions")
    @admin_required
    def api_sessions():
        return jsonify({"ok": True, "sessions": [serialize_session(item) for item in Session.query.order_by(Session.updated_at.desc())]})

    @app.get("/api/question-templates")
    @admin_required
    def api_question_templates():
        return jsonify({"ok": True, "templates": question_templates()})

    @app.post("/api/sessions")
    @admin_required
    def api_create_session():
        payload = request.get_json(silent=True) or {}
        try:
            session = Session(title=clean_text(payload.get("title"), 180, required=True), code=generate_session_code())
            db.session.add(session)
            db.session.commit()
            return jsonify({"ok": True, "session": serialize_session(session)}), 201
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)

    @app.get("/api/sessions/<code>")
    def api_session(code: str):
        session = require_session(code)
        refresh_session_timers(session)
        db.session.commit()
        return jsonify({"ok": True, "session": serialize_session(session)})

    @app.patch("/api/sessions/<code>")
    @admin_required
    def api_update_session(code: str):
        session = require_session(code)
        payload = request.get_json(silent=True) or {}
        try:
            if "title" in payload:
                session.title = clean_text(payload.get("title"), 180, required=True)
            if "theme" in payload:
                config = dict(session.config_json or {})
                config["theme"] = normalize_choice(payload.get("theme"), {"civic", "ocean", "contrast"}, config.get("theme", "civic"))
                session.config_json = config
            db.session.commit()
            broadcast_session(session)
            return jsonify({"ok": True, "session": serialize_session(session)})
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)

    @app.post("/api/sessions/<code>/questions")
    @admin_required
    def api_create_question(code: str):
        session = require_session(code)
        try:
            question = add_question(session, request.get_json(silent=True) or {})
            db.session.commit()
            broadcast_session(session)
            return jsonify({"ok": True, "question": serialize_question(question), "session": serialize_session(session)}), 201
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)

    @app.patch("/api/sessions/<code>/questions/<int:question_id>")
    @admin_required
    def api_update_question(code: str, question_id: int):
        session = require_session(code)
        question = require_question(session, question_id)
        try:
            update_question(question, request.get_json(silent=True) or {})
            db.session.commit()
            broadcast_session(session)
            return jsonify({"ok": True, "question": serialize_question(question), "session": serialize_session(session)})
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)

    @app.post("/api/sessions/<code>/questions/<int:question_id>/duplicate")
    @admin_required
    def api_duplicate_question(code: str, question_id: int):
        session = require_session(code)
        question = require_question(session, question_id)
        duplicate = duplicate_question(session, question)
        db.session.commit()
        broadcast_session(session)
        return jsonify({"ok": True, "question": serialize_question(duplicate), "session": serialize_session(session)}), 201

    @app.delete("/api/sessions/<code>/questions/<int:question_id>")
    @admin_required
    def api_delete_question(code: str, question_id: int):
        session = require_session(code)
        question = require_question(session, question_id)
        db.session.delete(question)
        db.session.flush()
        normalize_question_positions(session)
        clamp_active_index(session)
        db.session.commit()
        broadcast_session(session)
        return jsonify({"ok": True, "session": serialize_session(session)})

    @app.post("/api/sessions/<code>/questions/reorder")
    @admin_required
    def api_reorder_questions(code: str):
        session = require_session(code)
        payload = request.get_json(silent=True) or {}
        try:
            reorder_questions(session, payload.get("question_ids") or [])
            db.session.commit()
            broadcast_session(session)
            return jsonify({"ok": True, "session": serialize_session(session)})
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)

    @app.post("/api/sessions/<code>/questions/<int:question_id>/responses/<int:response_id>/moderate")
    @admin_required
    def api_moderate_response(code: str, question_id: int, response_id: int):
        session = require_session(code)
        question = require_question(session, question_id)
        try:
            response_record = moderate_response(question, response_id, request.get_json(silent=True) or {})
            db.session.commit()
            emit_results(session, question)
            broadcast_session(session)
            return jsonify({"ok": True, "response_id": response_record.id, "results": aggregate_question(question)})
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)

    @app.get("/api/sessions/<code>/insights")
    @admin_required
    def api_session_insights(code: str):
        session = require_session(code)
        refresh_session_timers(session)
        db.session.commit()
        return jsonify({"ok": True, "insights": build_insights(session)})

    @app.get("/api/sessions/<code>/export.csv")
    @admin_required
    def export_session_csv(code: str):
        session = require_session(code)
        return send_file(
            build_csv_export(session),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"mentimeter-{session.code}.csv",
        )

    @app.get("/api/sessions/<code>/export.xlsx")
    @admin_required
    def export_session_xlsx(code: str):
        session = require_session(code)
        return send_file(
            build_xlsx_export(session),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"mentimeter-{session.code}.xlsx",
        )

    @app.get("/api/sessions/<code>/runs/<int:run_id>/export.xlsx")
    @admin_required
    def export_session_run_xlsx(code: str, run_id: int):
        session = require_session(code)
        run = require_session_run(session, run_id)
        return send_file(
            build_xlsx_export(session, run),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"mentimeter-{session.code}-ejecucion-{run.run_number}.xlsx",
        )

    @app.get("/api/sessions/<code>/export.pdf")
    @admin_required
    def export_session_pdf(code: str):
        session = require_session(code)
        return send_file(
            build_pdf_export(session),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"mentimeter-{session.code}.pdf",
        )

    @app.get("/api/sessions/<code>/runs/<int:run_id>/export.pdf")
    @admin_required
    def export_session_run_pdf(code: str, run_id: int):
        session = require_session(code)
        run = require_session_run(session, run_id)
        return send_file(
            build_pdf_export(session, run),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"mentimeter-{session.code}-ejecucion-{run.run_number}.pdf",
        )

    @app.post("/api/sessions/<code>/control")
    @admin_required
    def api_control_session(code: str):
        session = require_session(code)
        try:
            apply_control(session, request.get_json(silent=True) or {})
            refresh_session_timers(session)
            db.session.commit()
            broadcast_session(session)
            return jsonify({"ok": True, "session": serialize_session(session)})
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)

    @app.post("/api/sessions/<code>/questions/<int:question_id>/responses")
    def api_submit_response(code: str, question_id: int):
        session = require_session(code)
        question = require_question(session, question_id)
        participant = get_or_create_participant(session, request.cookies.get(PARTICIPANT_COOKIE))
        allowed, retry_after = consume_rate_limit(rate_limit_identity("response", session))
        if not allowed:
            response = rate_limited_response(retry_after)
            return response
        try:
            if refresh_session_timers(session):
                db.session.commit()
                broadcast_session(session)
                raise ValueError("Tiempo agotado para esta pregunta.")
            response_record = record_response(session, question, participant, request.get_json(silent=True) or {})
            db.session.commit()
            emit_results(session, question)
            response = jsonify(
                {
                    "ok": True,
                    "participant_token": participant.token,
                    "response_id": response_record.id,
                    "results": aggregate_question(question),
                    "score": participant.score,
                }
            )
            response.set_cookie(PARTICIPANT_COOKIE, participant.token, httponly=True, samesite="Lax")
            return response
        except ValueError as exc:
            db.session.rollback()
            return json_error(exc)


def register_socket_events() -> None:
    @socketio.on("join_session")
    def socket_join_session(data):
        session = find_session((data or {}).get("code"))
        if session is None:
            return {"ok": False, "error": "Sesión no encontrada."}
        refresh_session_timers(session)
        participant = get_or_create_participant(session, (data or {}).get("participant_token"))
        participant.connected = True
        join_room(room_name(session.code))
        socket_participants[request.sid] = (session.id, participant.token)
        db.session.commit()
        emit("participant_count", participant_count(session), to=room_name(session.code))
        return {"ok": True, "session": serialize_session(session), "participant_token": participant.token}

    @socketio.on("submit_response")
    def socket_submit_response(data):
        if socket_payload_too_large(data):
            return {"ok": False, "error": "Payload demasiado grande."}
        session = find_session((data or {}).get("code"))
        if session is None:
            return {"ok": False, "error": "Sesión no encontrada."}
        question = db.session.get(Question, int((data or {}).get("question_id") or 0))
        if question is None or question.session_id != session.id:
            return {"ok": False, "error": "Pregunta no encontrada."}
        participant = get_or_create_participant(session, (data or {}).get("participant_token"))
        allowed, retry_after = consume_rate_limit(rate_limit_identity("socket-response", session))
        if not allowed:
            return {
                "ok": False,
                "error": "Demasiadas respuestas. Intenta nuevamente en unos segundos.",
                "retry_after": retry_after,
            }
        try:
            if refresh_session_timers(session):
                db.session.commit()
                emit("session_state", serialize_session(session), to=room_name(session.code))
                raise ValueError("Tiempo agotado para esta pregunta.")
            response_record = record_response(session, question, participant, (data or {}).get("payload") or {})
            db.session.commit()
            emit_results(session, question)
            emit("participant_count", participant_count(session), to=room_name(session.code))
            return {
                "ok": True,
                "participant_token": participant.token,
                "response_id": response_record.id,
                "results": aggregate_question(question),
                "score": participant.score,
            }
        except ValueError as exc:
            db.session.rollback()
            return {"ok": False, "error": str(exc)}

    @socketio.on("presenter_control")
    def socket_presenter_control(data):
        if not is_admin_authenticated():
            return {"ok": False, "error": "No autorizado."}
        session = find_session((data or {}).get("code"))
        if session is None:
            return {"ok": False, "error": "Sesión no encontrada."}
        try:
            apply_control(session, data or {})
            refresh_session_timers(session)
            db.session.commit()
            state = serialize_session(session)
            emit("session_state", state, to=room_name(session.code))
            return {"ok": True, "session": state}
        except ValueError as exc:
            db.session.rollback()
            return {"ok": False, "error": str(exc)}

    @socketio.on("moderate_response")
    def socket_moderate_response(data):
        if not is_admin_authenticated():
            return {"ok": False, "error": "No autorizado."}
        session = find_session((data or {}).get("code"))
        if session is None:
            return {"ok": False, "error": "Sesión no encontrada."}
        question = db.session.get(Question, int((data or {}).get("question_id") or 0))
        if question is None or question.session_id != session.id:
            return {"ok": False, "error": "Pregunta no encontrada."}
        try:
            response_record = moderate_response(question, int((data or {}).get("response_id") or 0), data or {})
            db.session.commit()
            emit_results(session, question)
            emit("session_state", serialize_session(session), to=room_name(session.code))
            return {"ok": True, "response_id": response_record.id, "results": aggregate_question(question)}
        except ValueError as exc:
            db.session.rollback()
            return {"ok": False, "error": str(exc)}

    @socketio.on("disconnect")
    def socket_disconnect():
        tracked = socket_participants.pop(request.sid, None)
        if not tracked:
            return
        session_id, token = tracked
        participant = Participant.query.filter_by(session_id=session_id, token=token).first()
        if participant:
            participant.connected = False
            db.session.commit()
            emit("participant_count", participant_count(participant.session), to=room_name(participant.session.code))


def add_question(session: Session, payload: dict[str, Any]) -> Question:
    normalized = normalize_question_payload(payload)
    question = Question(
        session=session,
        type=normalized["type"],
        title=normalized["title"],
        prompt=normalized["prompt"],
        position=next_question_position(session),
        config_json=normalized["config"],
        is_open=True,
    )
    db.session.add(question)
    db.session.flush()
    replace_options(question, normalized["options"], normalized.get("correct_option_labels") or [])
    return question


def update_question(question: Question, payload: dict[str, Any]) -> Question:
    merged = {
        "type": payload.get("type", question.type),
        "title": payload.get("title", question.title),
        "prompt": payload.get("prompt", question.prompt),
        "config": {**(question.config_json or {}), **(payload.get("config") or {})},
        "options": payload.get("options", [option.label for option in question.options]),
        "correct_option_labels": payload.get(
            "correct_option_labels",
            [option.label for option in question.options if option.is_correct],
        ),
    }
    normalized = normalize_question_payload(merged)
    question.type = normalized["type"]
    question.title = normalized["title"]
    question.prompt = normalized["prompt"]
    question.config_json = normalized["config"]
    replace_options(question, normalized["options"], normalized.get("correct_option_labels") or [])
    return question


def duplicate_question(session: Session, source: Question) -> Question:
    payload = {
        "type": source.type,
        "title": f"{source.title} (copia)"[:180],
        "prompt": source.prompt,
        "config": dict(source.config_json or {}),
        "options": [option.label for option in source.options],
        "correct_option_labels": [option.label for option in source.options if option.is_correct],
    }
    payload["config"].pop("timer_started_at", None)
    duplicate = add_question(session, payload)
    ordered = sorted(session.questions, key=lambda item: item.position)
    ordered_ids = [question.id for question in ordered if question.id != duplicate.id]
    source_index = ordered_ids.index(source.id)
    ordered_ids.insert(source_index + 1, duplicate.id)
    reorder_questions(session, ordered_ids)
    return duplicate


def default_text_boxes(title: str, body: str) -> list[dict[str, Any]]:
    boxes: list[dict[str, Any]] = [
        {
            "id": "title",
            "text": title,
            "x": 8,
            "y": 14,
            "w": 62,
            "h": 22,
            "font_size": 60,
            "font_weight": 800,
            "color": "#17212f",
            "background": "transparent",
            "align": "left",
            "auto_fit": True,
            "z": 1,
        }
    ]
    if body:
        boxes.append(
            {
                "id": "body",
                "text": body,
                "x": 8,
                "y": 42,
                "w": 60,
                "h": 28,
                "font_size": 28,
                "font_weight": 400,
                "color": "#334155",
                "background": "transparent",
                "align": "left",
                "auto_fit": True,
                "z": 2,
            }
        )
    return boxes


def sanitize_text_boxes(raw_boxes: Any, *, title: str, body: str) -> list[dict[str, Any]]:
    if not isinstance(raw_boxes, list) or not raw_boxes:
        raw_boxes = default_text_boxes(title, body)

    boxes: list[dict[str, Any]] = []
    total_text = 0
    for index, item in enumerate(raw_boxes[:TEXT_BOX_MAX_ITEMS], start=1):
        if not isinstance(item, dict):
            continue
        text = clean_text(item.get("text"), TEXT_BOX_MAX_TEXT)
        total_text += len(text)
        if total_text > TEXT_BOX_TOTAL_TEXT:
            raise ValueError("El texto total de los cuadros no puede superar 3000 caracteres.")
        box_id = sanitize_text_box_id(item.get("id"), index)
        boxes.append(
            {
                "id": box_id,
                "text": text,
                "x": clamp_float(item.get("x"), 0, 100, 8),
                "y": clamp_float(item.get("y"), 0, 100, 14),
                "w": clamp_float(item.get("w"), 5, 100, 50),
                "h": clamp_float(item.get("h"), 5, 100, 16),
                "font_size": clamp_int(item.get("font_size", 32), 12, 120),
                "font_weight": normalize_font_weight(item.get("font_weight")),
                "color": sanitize_hex_color(item.get("color"), "#17212f"),
                "background": sanitize_box_background(item.get("background")),
                "align": normalize_choice(item.get("align"), {"left", "center", "right"}, "left"),
                "auto_fit": as_bool(item.get("auto_fit"), True),
                "z": clamp_int(item.get("z", index), 0, 100),
            }
        )
    if not boxes:
        boxes = default_text_boxes(title, body)
    return boxes


def body_from_text_boxes(boxes: list[dict[str, Any]], fallback: str) -> str:
    for box in boxes:
        if box.get("id") == "body" and box.get("text"):
            return clean_text(box.get("text"), 1800)
    for box in boxes:
        if box.get("id") != "title" and box.get("text"):
            return clean_text(box.get("text"), 1800)
    return fallback


def sanitize_layout_blocks(raw_blocks: Any) -> dict[str, dict[str, Any]]:
    source: dict[str, Any] = {}
    if isinstance(raw_blocks, dict):
        source = raw_blocks
    elif isinstance(raw_blocks, list):
        for item in raw_blocks:
            if isinstance(item, dict) and str(item.get("id") or "") in LAYOUT_BLOCK_IDS:
                source[str(item.get("id"))] = item

    blocks: dict[str, dict[str, Any]] = {}
    for block_id in LAYOUT_BLOCK_IDS:
        defaults = DEFAULT_LAYOUT_BLOCKS[block_id]
        item = source.get(block_id)
        if not isinstance(item, dict):
            item = {}
        width = clamp_float(item.get("w"), 12, 100, defaults["w"])
        height = clamp_float(item.get("h"), 10, 100, defaults["h"])
        x = clamp_float(item.get("x"), 0, max(0, 100 - width), defaults["x"])
        y = clamp_float(item.get("y"), 0, max(0, 100 - height), defaults["y"])
        blocks[block_id] = {
            "id": block_id,
            "x": x,
            "y": y,
            "w": width,
            "h": height,
            "z": clamp_int(item.get("z", defaults["z"]), 0, 100),
        }
    return blocks


def sanitize_text_styles(raw_styles: Any) -> dict[str, dict[str, Any]]:
    if not isinstance(raw_styles, dict):
        return {}
    styles: dict[str, dict[str, Any]] = {}
    for raw_key, raw_item in list(raw_styles.items())[:80]:
        key = "".join(ch for ch in str(raw_key or "") if ch.isalnum() or ch in "-_:").strip()[:80]
        if not key or not isinstance(raw_item, dict):
            continue
        try:
            color = sanitize_hex_color(raw_item.get("color"), "#17212f")
        except ValueError:
            color = "#17212f"
        try:
            background = sanitize_box_background(raw_item.get("background"))
        except ValueError:
            background = "transparent"
        item: dict[str, Any] = {
            "font_size": clamp_int(raw_item.get("font_size", 32), 8, 120),
            "font_weight": normalize_font_weight(raw_item.get("font_weight")),
            "color": color,
            "background": background,
            "align": normalize_choice(raw_item.get("align"), {"left", "center", "right"}, "left"),
            "auto_fit": as_bool(raw_item.get("auto_fit"), True),
        }
        styles[key] = item
    return styles


def default_result_layout(question_type: str) -> str:
    if question_type == "word_cloud":
        return "cloud"
    if question_type == "open_text":
        return "cards"
    if question_type == "quiz":
        return "leaderboard"
    return "chart"


def normalize_result_contract(question_type: str, config: dict[str, Any]) -> dict[str, Any]:
    return {
        "result_placement": "slide",
        "show_results": as_bool(config.get("show_results"), True),
        "result_layout": normalize_choice(
            config.get("result_layout"),
            RESULT_LAYOUTS,
            default_result_layout(question_type),
        ),
    }


def serialized_question_config(question: Question) -> dict[str, Any]:
    config = dict(question.config_json or {})
    if question.type == "content_slide":
        return config
    return {
        **config,
        **normalize_result_contract(question.type, config),
        "layout_blocks": sanitize_layout_blocks(config.get("layout_blocks")),
    }


def normalize_question_payload(payload: dict[str, Any]) -> dict[str, Any]:
    question_type = str(payload.get("type") or "").strip()
    if question_type not in QUESTION_TYPES:
        raise ValueError("Tipo de pregunta no soportado.")
    title = clean_text(payload.get("title"), 180, required=True)
    prompt = clean_text(payload.get("prompt"), 1200, required=question_type != "content_slide")
    config = dict(payload.get("config") or {})
    options = parse_lines(payload.get("options"))
    correct_labels = set(parse_lines(payload.get("correct_option_labels") or config.get("correct_options")))
    result_contract = normalize_result_contract(question_type, config) if question_type != "content_slide" else {}
    layout_blocks = sanitize_layout_blocks(config.get("layout_blocks")) if question_type != "content_slide" else {}
    text_styles = sanitize_text_styles(config.get("text_styles")) if question_type != "content_slide" else {}

    if question_type == "content_slide":
        body = clean_text(config.get("body"), 1800, required=False)
        text_boxes = sanitize_text_boxes(config.get("text_boxes"), title=title, body=body)
        config = {
            "layout": normalize_choice(config.get("layout"), {"title", "text", "instructions", "qr"}, "title"),
            "body": body_from_text_boxes(text_boxes, body),
            "media_url": clean_text(config.get("media_url"), 800, required=False),
            "show_qr": as_bool(config.get("show_qr"), False),
            "text_boxes": text_boxes,
        }
        options = []
        correct_labels = set()
    elif question_type in {"multiple_choice", "ranking", "quiz"} and len(options) < 2:
        raise ValueError("Agrega al menos dos opciones.")
    if question_type == "quiz" and not correct_labels:
        raise ValueError("Marca al menos una respuesta correcta para el quiz.")
    if question_type == "content_slide":
        pass
    elif question_type == "scale":
        minimum = clamp_int(config.get("min", 1), 1, 10)
        maximum = clamp_int(config.get("max", 5), 2, 10)
        if maximum <= minimum:
            raise ValueError("La escala requiere un máximo mayor al mínimo.")
        config = {**result_contract, "layout_blocks": layout_blocks, "text_styles": text_styles, "min": minimum, "max": maximum}
    elif question_type == "quiz":
        config = {
            **result_contract,
            "layout_blocks": layout_blocks,
            "text_styles": text_styles,
            "timer_seconds": clamp_int(config.get("timer_seconds", 30), 5, 600),
            "points": clamp_int(config.get("points", 100), 1, 1000),
        }
    elif question_type in {"word_cloud", "open_text"}:
        config = {
            **result_contract,
            "layout_blocks": layout_blocks,
            "text_styles": text_styles,
            "moderation": normalize_choice(config.get("moderation"), {"none", "manual"}, "none"),
        }
    else:
        config = {
            **result_contract,
            "layout_blocks": layout_blocks,
            "text_styles": text_styles,
            **({"max_entries": config.get("max_entries")} if config.get("max_entries") is not None else {}),
        }

    return {
        "type": question_type,
        "title": title,
        "prompt": prompt,
        "config": config,
        "options": options,
        "correct_option_labels": correct_labels,
    }


def replace_options(question: Question, labels: list[str], correct_labels: set[str]) -> None:
    for option in list(question.options):
        db.session.delete(option)
    db.session.flush()
    for index, label in enumerate(labels, start=1):
        db.session.add(
            Option(
                question=question,
                label=label,
                position=index,
                is_correct=label in correct_labels,
            )
        )
    db.session.flush()


RUN_DEFAULT = object()
RUN_LEGACY = object()
RUN_EMPTY = object()


def sorted_runs(session: Session) -> list[SessionRun]:
    return sorted(session.runs, key=lambda item: item.run_number)


def active_session_run(session: Session) -> SessionRun | None:
    return next((run for run in sorted_runs(session) if run.status == "active"), None)


def latest_session_run(session: Session) -> SessionRun | None:
    runs = sorted_runs(session)
    return runs[-1] if runs else None


def next_run_number(session: Session) -> int:
    return max([run.run_number for run in session.runs] or [0]) + 1


def get_or_create_active_run(session: Session) -> tuple[SessionRun, bool]:
    run = active_session_run(session)
    if run is not None:
        return run, False
    run = SessionRun(
        session=session,
        run_number=next_run_number(session),
        status="active",
        started_at=utcnow(),
        summary_json={},
    )
    db.session.add(run)
    db.session.flush()
    return run, True


def close_active_run(session: Session) -> SessionRun | None:
    run = active_session_run(session)
    if run is None:
        return None
    run.status = "closed"
    run.ended_at = run.ended_at or utcnow()
    run.summary_json = build_run_snapshot(session, run)
    return run


def default_live_run(session: Session):
    run = active_session_run(session)
    if run is not None:
        return run
    if session.status == "closed":
        return latest_session_run(session) or (RUN_LEGACY if session.responses else RUN_EMPTY)
    return RUN_EMPTY if session.runs else RUN_LEGACY


def default_report_run(session: Session):
    return active_session_run(session) or latest_session_run(session) or (RUN_LEGACY if session.responses else RUN_EMPTY)


def require_session_run(session: Session, run_id: int) -> SessionRun:
    run = db.session.get(SessionRun, int(run_id))
    if run is None or run.session_id != session.id:
        abort(404)
    return run


def response_key_for_run(question_type: str, run) -> str:
    base_key = "default" if question_type in SINGLE_RESPONSE_TYPES else token_urlsafe(8)
    if run in {RUN_LEGACY, RUN_EMPTY, None}:
        return base_key
    return f"run:{run.id}:{base_key}"


def display_response_key(response: Response) -> str:
    key = str(response.response_key or "")
    if key.startswith("run:"):
        parts = key.split(":", 2)
        return parts[2] if len(parts) == 3 else key
    return key


def scoped_question_responses(question: Question, run=RUN_DEFAULT, *, active_only: bool = True) -> list[Response]:
    if run is RUN_DEFAULT:
        run = default_live_run(question.session)
    responses = list(question.responses)
    if run is RUN_EMPTY:
        responses = []
    elif run is RUN_LEGACY:
        responses = [response for response in responses if response.run_id is None]
    elif isinstance(run, SessionRun):
        responses = [response for response in responses if response.run_id == run.id]
    if active_only:
        responses = [response for response in responses if response.is_active]
    return responses


def scoped_session_responses(session: Session, run=RUN_DEFAULT, *, active_only: bool = False) -> list[Response]:
    if run is RUN_DEFAULT:
        run = default_report_run(session)
    responses = list(session.responses)
    if run is RUN_EMPTY:
        responses = []
    elif run is RUN_LEGACY:
        responses = [response for response in responses if response.run_id is None]
    elif isinstance(run, SessionRun):
        responses = [response for response in responses if response.run_id == run.id]
    if active_only:
        responses = [response for response in responses if response.is_active]
    return sorted(responses, key=lambda item: (item.question.position if item.question else 0, item.created_at or utcnow()))


def record_response(session: Session, question: Question, participant: Participant, payload: dict[str, Any]) -> Response:
    # Single-response questions reuse "default" so a participant edits/replaces the vote instead of duplicating it.
    if session.status != "active":
        raise ValueError("La sesión no está activa.")
    if question.type == "content_slide":
        raise ValueError("Esta diapositiva no acepta respuestas.")
    if session.active_question and session.active_question.id != question.id:
        raise ValueError("Esta no es la pregunta activa.")
    if not question.is_open:
        raise ValueError("La votación de esta pregunta está cerrada.")

    normalized = normalize_response_payload(question, payload)
    run, _created = get_or_create_active_run(session)
    response_key = response_key_for_run(question.type, run)
    existing = Response.query.filter_by(
        question_id=question.id,
        participant_id=participant.id,
        response_key=response_key,
        run_id=run.id,
    ).first()
    if existing and question.type == "quiz":
        participant.score = max(0, participant.score - existing.score_awarded)

    score_awarded = quiz_score(question, normalized) if question.type == "quiz" else 0
    if existing:
        existing.payload_json = normalized
        existing.score_awarded = score_awarded
        existing.is_active = True
        response_record = existing
    else:
        response_record = Response(
            session=session,
            run=run,
            question=question,
            participant=participant,
            response_key=response_key,
            payload_json=normalized,
            score_awarded=score_awarded,
        )
        db.session.add(response_record)

    if question.type == "quiz":
        participant.score += score_awarded
    participant.connected = True
    return response_record


def normalize_response_payload(question: Question, payload: dict[str, Any]) -> dict[str, Any]:
    if question.type == "content_slide":
        raise ValueError("Esta diapositiva no acepta respuestas.")
    if question.type in {"multiple_choice", "quiz"}:
        option_id = int(payload.get("option_id") or payload.get("choice") or 0)
        if option_id not in {option.id for option in question.options}:
            raise ValueError("Selecciona una opción válida.")
        return {"option_id": option_id}
    if question.type == "word_cloud":
        text = clean_text(payload.get("text"), 80, required=True)
        return {"text": text, "status": initial_response_status(question)}
    if question.type == "open_text":
        text = clean_text(payload.get("text"), 500, required=True)
        return {"text": text, "status": initial_response_status(question)}
    if question.type == "scale":
        minimum = int((question.config_json or {}).get("min", 1))
        maximum = int((question.config_json or {}).get("max", 5))
        value = clamp_int(payload.get("value"), minimum, maximum)
        return {"value": value}
    if question.type == "ranking":
        option_ids = [int(value) for value in payload.get("ranking") or []]
        expected = {option.id for option in question.options}
        if set(option_ids) != expected:
            raise ValueError("Ordena todas las opciones del ranking.")
        return {"ranking": option_ids}
    raise ValueError("Tipo de pregunta no soportado.")


def moderate_response(question: Question, response_id: int, payload: dict[str, Any]) -> Response:
    response_record = db.session.get(Response, response_id)
    if response_record is None or response_record.question_id != question.id:
        raise ValueError("Respuesta no encontrada.")
    action = normalize_choice(payload.get("action"), {"approve", "reject"}, "approve")
    next_status = "approved" if action == "approve" else "rejected"
    data = dict(response_record.payload_json or {})
    data["status"] = next_status
    response_record.payload_json = data
    response_record.is_active = next_status != "rejected"
    return response_record


def apply_control(session: Session, payload: dict[str, Any]) -> None:
    action = str(payload.get("action") or "").strip()
    questions = sorted(session.questions, key=lambda item: item.position)
    if action == "start":
        _run, created = get_or_create_active_run(session)
        if created:
            for participant in session.participants:
                participant.score = 0
        session.status = "active"
        session.active_question_index = min(session.active_question_index, max(len(questions) - 1, 0))
        if session.active_question:
            session.active_question.is_open = True
            start_timer_if_needed(session.active_question)
    elif action == "close":
        session.status = "closed"
        for question in questions:
            question.is_open = False
        close_active_run(session)
    elif action == "reset":
        close_active_run(session)
        for participant in session.participants:
            participant.score = 0
        for question in questions:
            question.is_open = True
            clear_question_timer(question)
        session.status = "draft"
        session.active_question_index = 0
        session.config_json = {**(session.config_json or {}), "theme": (session.config_json or {}).get("theme", "civic")}
    elif action in {"next", "next_slide"}:
        session.active_question_index = min(session.active_question_index + 1, max(len(questions) - 1, 0))
        if session.status == "active" and session.active_question:
            session.active_question.is_open = True
            start_timer_if_needed(session.active_question)
    elif action in {"previous", "previous_slide"}:
        session.active_question_index = max(session.active_question_index - 1, 0)
        if session.status == "active" and session.active_question:
            session.active_question.is_open = True
            start_timer_if_needed(session.active_question)
    elif action in {"go", "go_to_slide"}:
        index = int(payload.get("index") or 0)
        session.active_question_index = min(max(index, 0), max(len(questions) - 1, 0))
        if session.status == "active" and session.active_question:
            session.active_question.is_open = True
            start_timer_if_needed(session.active_question)
    elif action == "close_question":
        if session.active_question:
            session.active_question.is_open = False
    elif action == "open_question":
        if session.active_question:
            session.active_question.is_open = True
            start_timer_if_needed(session.active_question, restart=True)
    elif action == "set_theme":
        theme = normalize_choice(payload.get("theme"), {"civic", "ocean", "contrast"}, "civic")
        config = dict(session.config_json or {})
        config["theme"] = theme
        session.config_json = config
    else:
        raise ValueError("Acción no soportada.")


def aggregate_question(question: Question, run=RUN_DEFAULT) -> dict[str, Any]:
    # Results are aggregated on the server so every presenter/audience client receives the same live state.
    responses = scoped_question_responses(question, run)
    if question.type == "content_slide":
        return {"type": "content_slide", "total": 0}
    if question.type in {"multiple_choice", "quiz"}:
        counts = Counter(int(response.payload_json.get("option_id") or 0) for response in responses)
        return {
            "type": question.type,
            "total": sum(counts.values()),
            "options": [
                {
                    "id": option.id,
                    "label": option.label,
                    "count": counts.get(option.id, 0),
                    "is_correct": option.is_correct,
                }
                for option in question.options
            ],
            "leaderboard": leaderboard(question.session, run) if question.type == "quiz" else [],
        }
    if question.type == "word_cloud":
        counts = Counter(
            normalize_word(response.payload_json.get("text", ""))
            for response in responses
            if response.payload_json.get("status", "approved") == "approved"
        )
        counts.pop("", None)
        return {
            "type": "word_cloud",
            "total": sum(counts.values()),
            "words": [{"text": text, "count": count} for text, count in counts.most_common(80)],
        }
    if question.type == "open_text":
        visible = [response for response in responses if response.payload_json.get("status", "approved") == "approved"]
        return {
            "type": "open_text",
            "total": len(visible),
            "cards": [
                {"id": response.id, "text": response.payload_json.get("text", "")}
                for response in sorted(visible, key=lambda item: item.created_at, reverse=True)
            ],
        }
    if question.type == "scale":
        values = [int(response.payload_json.get("value") or 0) for response in responses]
        counts = Counter(values)
        minimum = int((question.config_json or {}).get("min", 1))
        maximum = int((question.config_json or {}).get("max", 5))
        return {
            "type": "scale",
            "total": len(values),
            "average": round(sum(values) / len(values), 2) if values else 0,
            "values": [{"value": value, "count": counts.get(value, 0)} for value in range(minimum, maximum + 1)],
        }
    if question.type == "ranking":
        scores: defaultdict[int, int] = defaultdict(int)
        option_count = len(question.options)
        for response in responses:
            for index, option_id in enumerate(response.payload_json.get("ranking") or []):
                scores[int(option_id)] += option_count - index
        return {
            "type": "ranking",
            "total": len(responses),
            "options": [
                {"id": option.id, "label": option.label, "score": scores.get(option.id, 0)}
                for option in question.options
            ],
        }
    return {"type": question.type, "total": len(responses)}


def serialize_session(session: Session) -> dict[str, Any]:
    refresh_session_timers(session)
    active = session.active_question
    active_run = active_session_run(session)
    report_run = default_report_run(session)
    return {
        "id": session.id,
        "title": session.title,
        "code": session.code,
        "status": session.status,
        "theme": (session.config_json or {}).get("theme", "civic"),
        "active_run_id": active_run.id if active_run else None,
        "report_run_id": report_run.id if isinstance(report_run, SessionRun) else None,
        "active_question_index": session.active_question_index,
        "active_question_id": active.id if active else None,
        "participant_count": len(session.participants),
        "connected_count": len([participant for participant in session.participants if participant.connected]),
        "join_url": url_for("audience", code=session.code, _external=True),
        "qr_url": url_for("qr_png", code=session.code),
        "runs": [serialize_run(run) for run in sorted_runs(session)],
        "questions": [serialize_question(question) for question in sorted(session.questions, key=lambda item: item.position)],
    }


def serialize_run(run: SessionRun) -> dict[str, Any]:
    responses = scoped_session_responses(run.session, run, active_only=False)
    active_responses = [response for response in responses if response.is_active]
    return {
        "id": run.id,
        "run_number": run.run_number,
        "status": run.status,
        "started_at": run.started_at.isoformat() if run.started_at else None,
        "ended_at": run.ended_at.isoformat() if run.ended_at else None,
        "participant_count": len({response.participant_id for response in active_responses}),
        "response_count": len(active_responses),
        "xlsx_url": url_for("export_session_run_xlsx", code=run.session.code, run_id=run.id),
        "pdf_url": url_for("export_session_run_pdf", code=run.session.code, run_id=run.id),
    }


def serialize_question(question: Question) -> dict[str, Any]:
    return {
        "id": question.id,
        "type": question.type,
        "title": question.title,
        "prompt": question.prompt,
        "position": question.position,
        "is_open": question.is_open,
        "config": serialized_question_config(question),
        "timer": timer_state(question),
        "options": [
            {"id": option.id, "label": option.label, "position": option.position, "is_correct": option.is_correct}
            for option in sorted(question.options, key=lambda item: item.position)
        ],
        "pending_responses": pending_responses(question),
        "results": aggregate_question(question),
    }


def ensure_demo_session() -> Session:
    # Idempotent seed: keeps local startup friendly without overwriting user-created sessions.
    session = Session.query.filter_by(code="123456").first()
    if session:
        return session
    session = Session(title="Demo participativa", code="123456", status="draft")
    db.session.add(session)
    db.session.flush()
    add_question(
        session,
        {
            "type": "content_slide",
            "title": "Demo participativa",
            "prompt": "",
            "config": {
                "layout": "title",
                "body": "Presentación interactiva para talleres, consultas y capacitaciones.",
            },
        },
    )
    add_question(
        session,
        {
            "type": "multiple_choice",
            "title": "Prioridad del taller",
            "prompt": "Que tema deberiamos atender primero?",
            "options": ["Atencion ciudadana", "Procesos internos", "Datos y reportes", "Coordinacion"],
        },
    )
    add_question(
        session,
        {
            "type": "word_cloud",
            "title": "Una palabra",
            "prompt": "Describe el reto principal en una palabra.",
        },
    )
    add_question(
        session,
        {
            "type": "quiz",
            "title": "Quiz rápido",
            "prompt": "Que herramienta permite actualizaciones en tiempo real?",
            "options": ["CSV", "WebSockets", "PDF", "Correo"],
            "correct_option_labels": ["WebSockets"],
            "config": {"timer_seconds": 30, "points": 100},
        },
    )
    db.session.commit()
    return session


def question_templates() -> list[dict[str, Any]]:
    return [
        {
            "name": "Portada",
            "payload": {
                "type": "content_slide",
                "title": "Título de la presentación",
                "prompt": "",
                "config": {"layout": "title", "body": "Subtítulo o contexto del taller."},
            },
        },
        {
            "name": "Instrucciones con QR",
            "payload": {
                "type": "content_slide",
                "title": "Participa con tu celular",
                "prompt": "",
                "config": {"layout": "qr", "body": "Escanea el QR o entra con el código de la sesión.", "show_qr": True},
            },
        },
        {
            "name": "Pulso de prioridad",
            "payload": {
                "type": "multiple_choice",
                "title": "Pulso de prioridad",
                "prompt": "Que tema merece atencion inmediata?",
                "options": ["Atencion ciudadana", "Eficiencia interna", "Transparencia", "Coordinacion"],
            },
        },
        {
            "name": "Lluvia con moderación",
            "payload": {
                "type": "word_cloud",
                "title": "Lluvia de ideas",
                "prompt": "Escribe una palabra clave para el diagnóstico.",
                "config": {"moderation": "manual"},
            },
        },
        {
            "name": "Escala 1 a 10",
            "payload": {
                "type": "scale",
                "title": "Nivel de acuerdo",
                "prompt": "Califica del 1 al 10.",
                "config": {"min": 1, "max": 10},
            },
        },
        {
            "name": "Quiz rápido",
            "payload": {
                "type": "quiz",
                "title": "Quiz rápido",
                "prompt": "Selecciona la respuesta correcta.",
                "options": ["Opción A", "Opción B", "Opción C"],
                "correct_option_labels": ["Opción B"],
                "config": {"timer_seconds": 30, "points": 100},
            },
        },
    ]


def refresh_session_timers(session: Session) -> bool:
    changed = False
    if session.status != "active":
        return changed
    for question in session.questions:
        if question.type == "quiz" and question.is_open and timer_remaining(question) == 0:
            question.is_open = False
            changed = True
    return changed


def start_timer_if_needed(question: Question, *, restart: bool = False) -> None:
    if question.type != "quiz":
        return
    config = dict(question.config_json or {})
    if restart or not config.get("timer_started_at"):
        config["timer_started_at"] = utcnow().isoformat()
        question.config_json = config


def clear_question_timer(question: Question) -> None:
    if question.type != "quiz":
        return
    config = dict(question.config_json or {})
    config.pop("timer_started_at", None)
    question.config_json = config


def timer_remaining(question: Question) -> int | None:
    if question.type != "quiz":
        return None
    config = question.config_json or {}
    duration = int(config.get("timer_seconds") or 30)
    started_at = parse_datetime(config.get("timer_started_at"))
    if not started_at:
        return duration
    elapsed = int((utcnow() - started_at).total_seconds())
    return max(0, duration - elapsed)


def timer_state(question: Question) -> dict[str, Any] | None:
    if question.type != "quiz":
        return None
    config = question.config_json or {}
    return {
        "duration": int(config.get("timer_seconds") or 30),
        "started_at": config.get("timer_started_at"),
        "remaining": timer_remaining(question),
    }


def pending_responses(question: Question) -> list[dict[str, Any]]:
    if question.type not in {"word_cloud", "open_text"}:
        return []
    return [
        {"id": response.id, "text": response.payload_json.get("text", ""), "status": response.payload_json.get("status", "approved")}
        for response in scoped_question_responses(question)
        if response.payload_json.get("status") == "pending"
    ]


def initial_response_status(question: Question) -> str:
    return "pending" if (question.config_json or {}).get("moderation") == "manual" else "approved"


def build_insights(session: Session) -> dict[str, Any]:
    refresh_session_timers(session)
    return build_session_report(session, default_report_run(session), include_responses=False)


def selected_report_run(session: Session, run=RUN_DEFAULT):
    return default_report_run(session) if run is RUN_DEFAULT else run


def run_label(run) -> str:
    if isinstance(run, SessionRun):
        return f"Ejecucion {run.run_number}"
    if run is RUN_LEGACY:
        return "Historico sin ejecucion"
    return "Sin ejecucion"


def run_status(run) -> str:
    if isinstance(run, SessionRun):
        return run.status
    if run is RUN_LEGACY:
        return "legacy"
    return "empty"


def run_started_at(run) -> str:
    return run.started_at.isoformat() if isinstance(run, SessionRun) and run.started_at else ""


def run_ended_at(run) -> str:
    return run.ended_at.isoformat() if isinstance(run, SessionRun) and run.ended_at else ""


def participant_aliases(responses: list[Response]) -> dict[int, str]:
    ordered_ids: list[int] = []
    for response in sorted(responses, key=lambda item: item.created_at or utcnow()):
        if response.participant_id not in ordered_ids:
            ordered_ids.append(response.participant_id)
    return {participant_id: f"Participante {index}" for index, participant_id in enumerate(ordered_ids, start=1)}


def label_for_question_type(question_type: str) -> str:
    labels = {
        "content_slide": "Contenido",
        "multiple_choice": "Opcion multiple",
        "word_cloud": "Nube",
        "scale": "Escala",
        "open_text": "Texto abierto",
        "ranking": "Ranking",
        "quiz": "Quiz",
    }
    return labels.get(question_type, question_type)


def response_answer_text(response: Response) -> str:
    question = response.question
    payload = response.payload_json or {}
    if question is None:
        return json.dumps(payload, ensure_ascii=False)
    if question.type in {"multiple_choice", "quiz"}:
        option_id = int(payload.get("option_id") or 0)
        option = next((item for item in question.options if item.id == option_id), None)
        return option.label if option else str(option_id or "")
    if question.type in {"word_cloud", "open_text"}:
        return str(payload.get("text") or "")
    if question.type == "scale":
        return str(payload.get("value") or "")
    if question.type == "ranking":
        by_id = {option.id: option.label for option in question.options}
        labels = [by_id.get(int(option_id), str(option_id)) for option_id in payload.get("ranking") or []]
        return " > ".join(labels)
    return json.dumps(payload, ensure_ascii=False)


def interpreted_result_rows(question: Question, results: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if question.type in {"multiple_choice", "quiz"}:
        total = int(results.get("total") or 0)
        for option in results.get("options") or []:
            count = int(option.get("count") or 0)
            rows.append(
                {
                    "metric": option.get("label") or "",
                    "value": count,
                    "count": count,
                    "percent": round((count / total) * 100, 2) if total else 0,
                    "detail": "Correcta" if option.get("is_correct") else "",
                }
            )
    elif question.type == "word_cloud":
        for word in results.get("words") or []:
            count = int(word.get("count") or 0)
            rows.append({"metric": word.get("text") or "", "value": count, "count": count, "percent": "", "detail": "menciones"})
    elif question.type == "open_text":
        for card in results.get("cards") or []:
            rows.append({"metric": f"Respuesta {card.get('id')}", "value": card.get("text") or "", "count": 1, "percent": "", "detail": ""})
    elif question.type == "scale":
        rows.append({"metric": "Promedio", "value": results.get("average") or 0, "count": results.get("total") or 0, "percent": "", "detail": ""})
        for value in results.get("values") or []:
            count = int(value.get("count") or 0)
            rows.append({"metric": str(value.get("value")), "value": count, "count": count, "percent": "", "detail": "respuestas"})
    elif question.type == "ranking":
        ordered = sorted(results.get("options") or [], key=lambda item: item.get("score") or 0, reverse=True)
        for index, option in enumerate(ordered, start=1):
            rows.append({"metric": option.get("label") or "", "value": int(option.get("score") or 0), "count": index, "percent": "", "detail": f"Lugar {index}"})
    else:
        rows.append({"metric": label_for_question_type(question.type), "value": results.get("total") or 0, "count": "", "percent": "", "detail": ""})
    return rows


def build_session_report(session: Session, run=RUN_DEFAULT, *, include_responses: bool = True) -> dict[str, Any]:
    refresh_session_timers(session)
    selected_run = selected_report_run(session, run)
    questions = sorted(session.questions, key=lambda item: item.position)
    responses = scoped_session_responses(session, selected_run, active_only=False)
    active_responses = [response for response in responses if response.is_active]
    aliases = participant_aliases(responses)
    slides: list[dict[str, Any]] = []
    for question in questions:
        question_responses = scoped_question_responses(question, selected_run, active_only=False)
        active_question_responses = [response for response in question_responses if response.is_active]
        results = aggregate_question(question, selected_run)
        slides.append(
            {
                "id": question.id,
                "position": question.position,
                "title": question.title,
                "prompt": question.prompt,
                "type": question.type,
                "type_label": label_for_question_type(question.type),
                "is_open": question.is_open,
                "response_count": len(active_question_responses),
                "raw_response_count": len(question_responses),
                "results": results,
                "result_rows": interpreted_result_rows(question, results),
            }
        )
    raw_responses = []
    if include_responses:
        for response in responses:
            raw_responses.append(
                {
                    "id": response.id,
                    "slide_position": response.question.position if response.question else "",
                    "slide_title": response.question.title if response.question else "",
                    "type": response.question.type if response.question else "",
                    "participant": aliases.get(response.participant_id, "Participante"),
                    "response": response_answer_text(response),
                    "status": (response.payload_json or {}).get("status", "activa" if response.is_active else "inactiva"),
                    "score_awarded": response.score_awarded,
                    "created_at": response.created_at.isoformat() if response.created_at else "",
                    "payload": response.payload_json or {},
                    "response_key": display_response_key(response),
                    "active": response.is_active,
                }
            )
    metadata = {
        "code": session.code,
        "title": session.title,
        "status": session.status,
        "theme": (session.config_json or {}).get("theme", "civic"),
        "run_id": selected_run.id if isinstance(selected_run, SessionRun) else None,
        "run_number": selected_run.run_number if isinstance(selected_run, SessionRun) else None,
        "run_label": run_label(selected_run),
        "run_status": run_status(selected_run),
        "started_at": run_started_at(selected_run),
        "ended_at": run_ended_at(selected_run),
        "question_count": len(questions),
        "participant_count": len({response.participant_id for response in active_responses}),
        "response_count": len(active_responses),
        "raw_response_count": len(responses),
        "active_question": session.active_question.title if session.active_question else None,
    }
    return {
        **metadata,
        "metadata": metadata,
        "questions": slides,
        "slides": slides,
        "responses": raw_responses,
        "leaderboard": leaderboard(session, selected_run),
    }


def build_run_snapshot(session: Session, run: SessionRun) -> dict[str, Any]:
    report = build_session_report(session, run, include_responses=False)
    return {
        "title": report["title"],
        "code": report["code"],
        "run_number": report["run_number"],
        "participant_count": report["participant_count"],
        "response_count": report["response_count"],
        "question_count": report["question_count"],
        "closed_at": run.ended_at.isoformat() if run.ended_at else "",
        "slides": [
            {
                "id": slide["id"],
                "position": slide["position"],
                "title": slide["title"],
                "type": slide["type"],
                "response_count": slide["response_count"],
                "results": slide["result_rows"][:20],
            }
            for slide in report["slides"]
        ],
    }


def export_rows(session: Session, run=RUN_DEFAULT) -> list[list[Any]]:
    report = build_session_report(session, run)
    rows: list[list[Any]] = [["session_code", "run", "slide", "type", "participant", "response", "status", "score_awarded", "created_at"]]
    for response in report["responses"]:
        rows.append(
            [
                report["code"],
                report["run_label"],
                response["slide_title"],
                response["type"],
                response["participant"],
                response["response"],
                response["status"],
                response["score_awarded"],
                response["created_at"],
            ]
        )
    return rows


def build_csv_export(session: Session, run=RUN_DEFAULT) -> BytesIO:
    text_buffer = StringIO()
    writer = csv.writer(text_buffer)
    writer.writerows(export_rows(session, run))
    buffer = BytesIO(text_buffer.getvalue().encode("utf-8-sig"))
    buffer.seek(0)
    return buffer


def build_xlsx_export(session: Session, run=RUN_DEFAULT) -> BytesIO:
    report = build_session_report(session, run)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    append_table(
        summary,
        ["Campo", "Valor"],
        [
            ["Presentacion", report["title"]],
            ["Codigo", report["code"]],
            ["Ejecucion", report["run_label"]],
            ["Estado de ejecucion", report["run_status"]],
            ["Inicio", report["started_at"]],
            ["Cierre", report["ended_at"]],
            ["Diapositivas", report["question_count"]],
            ["Participantes", report["participant_count"]],
            ["Respuestas activas", report["response_count"]],
            ["Respuestas totales", report["raw_response_count"]],
        ],
    )
    append_table(
        workbook.create_sheet("Diapositivas"),
        ["#", "Titulo", "Tipo", "Pregunta", "Estado", "Respuestas"],
        [
            [slide["position"], slide["title"], slide["type_label"], slide["prompt"], "Abierta" if slide["is_open"] else "Cerrada", slide["response_count"]]
            for slide in report["slides"]
        ],
    )
    append_table(
        workbook.create_sheet("Resultados"),
        ["#", "Diapositiva", "Tipo", "Metrica", "Valor", "Conteo", "Porcentaje", "Detalle"],
        [
            [slide["position"], slide["title"], slide["type_label"], row["metric"], row["value"], row["count"], row["percent"], row["detail"]]
            for slide in report["slides"]
            for row in slide["result_rows"]
        ],
    )
    append_table(
        workbook.create_sheet("Respuestas"),
        ["#", "Diapositiva", "Tipo", "Participante", "Respuesta", "Estado", "Puntos", "Fecha", "Payload"],
        [
            [
                response["slide_position"],
                response["slide_title"],
                label_for_question_type(response["type"]),
                response["participant"],
                response["response"],
                response["status"],
                response["score_awarded"],
                response["created_at"],
                json.dumps(response["payload"], ensure_ascii=False),
            ]
            for response in report["responses"]
        ],
    )
    add_type_specific_sheets(workbook, report)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def add_type_specific_sheets(workbook: Workbook, report: dict[str, Any]) -> None:
    sheet_by_type = {
        "multiple_choice": "Opcion multiple",
        "word_cloud": "Nube",
        "open_text": "Abiertas",
        "scale": "Escalas",
        "ranking": "Ranking",
        "quiz": "Quiz",
    }
    for question_type, sheet_name in sheet_by_type.items():
        rows = [
            [slide["position"], slide["title"], row["metric"], row["value"], row["count"], row["percent"], row["detail"]]
            for slide in report["slides"]
            if slide["type"] == question_type
            for row in slide["result_rows"]
        ]
        if rows:
            append_table(workbook.create_sheet(sheet_name), ["#", "Diapositiva", "Metrica", "Valor", "Conteo", "Porcentaje", "Detalle"], rows)
    if report["leaderboard"]:
        append_table(
            workbook.create_sheet("Leaderboard"),
            ["Lugar", "Participante", "Puntos"],
            [[index, item["label"], item["score"]] for index, item in enumerate(report["leaderboard"], start=1)],
        )


def append_table(sheet, headers: list[Any], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="17212F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 54)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_legacy_pdf_export(session: Session) -> BytesIO:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 54
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(54, y, f"Resultados {session.title}")
    y -= 24
    pdf.setFont("Helvetica", 10)
    pdf.drawString(54, y, f"Código: {session.code} | Participantes: {len(session.participants)} | Respuestas: {len(session.responses)}")
    y -= 30
    for question in sorted(session.questions, key=lambda item: item.position):
        if y < 120:
            pdf.showPage()
            y = height - 54
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(54, y, f"{question.position}. {question.title} ({question.type})")
        y -= 16
        pdf.setFont("Helvetica", 9)
        results = aggregate_question(question)
        for line in json.dumps(results, ensure_ascii=False)[:1000].split(","):
            if y < 54:
                pdf.showPage()
                y = height - 54
            pdf.drawString(64, y, line[:105])
            y -= 12
        y -= 8
    pdf.save()
    buffer.seek(0)
    return buffer


def build_pdf_export(session: Session, run=RUN_DEFAULT) -> BytesIO:
    report = build_session_report(session, run)
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 54
    y = height - margin
    y = pdf_title(pdf, y, "Resultados de sesion", report["title"])
    for line in [
        f"Codigo: {report['code']} | {report['run_label']} | Estado: {report['run_status']}",
        f"Participantes: {report['participant_count']} | Respuestas activas: {report['response_count']} | Diapositivas: {report['question_count']}",
        f"Inicio: {report['started_at'] or 'N/D'} | Cierre: {report['ended_at'] or 'N/D'}",
    ]:
        y = pdf_line(pdf, y, line, margin, width, height, font_size=10)
    y -= 14
    y = pdf_section(pdf, y, "Dinamicas", width, height)
    for slide in report["slides"]:
        y = pdf_line(pdf, y, f"{slide['position']}. {slide['title']} ({slide['type_label']}) - {slide['response_count']} respuestas", margin, width, height, bold=True)
    y -= 10
    for slide in report["slides"]:
        y = pdf_section(pdf, y, f"{slide['position']}. {slide['title']}", width, height)
        y = pdf_line(pdf, y, f"Tipo: {slide['type_label']} | Pregunta: {slide['prompt'] or 'N/D'}", margin, width, height, font_size=9)
        for row in slide["result_rows"][:18]:
            detail = f" ({row['detail']})" if row["detail"] else ""
            percent = f" - {row['percent']}%" if row["percent"] not in ("", None) else ""
            y = pdf_line(pdf, y, f"- {row['metric']}: {row['value']}{percent}{detail}", margin + 10, width, height, font_size=9)
        if not slide["result_rows"]:
            y = pdf_line(pdf, y, "- Sin resultados capturados.", margin + 10, width, height, font_size=9)
        y -= 8
    if report["leaderboard"]:
        y = pdf_section(pdf, y, "Leaderboard", width, height)
        for index, item in enumerate(report["leaderboard"], start=1):
            y = pdf_line(pdf, y, f"{index}. {item['label']} - {item['score']} puntos", margin + 10, width, height, font_size=9)
    y = pdf_section(pdf, y, "Anexo de respuestas anonimas", width, height)
    for response in report["responses"][:120]:
        y = pdf_line(
            pdf,
            y,
            f"{response['participant']} | {response['slide_title']} | {response['response']} | {response['created_at']}",
            margin + 10,
            width,
            height,
            font_size=8,
        )
    pdf.save()
    buffer.seek(0)
    return buffer


def pdf_title(pdf, y: float, title: str, subtitle: str) -> float:
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(54, y, title)
    y -= 24
    pdf.setFont("Helvetica-Bold", 13)
    for line in wrap_pdf_text(subtitle, 78):
        pdf.drawString(54, y, line)
        y -= 16
    return y - 8


def pdf_section(pdf, y: float, title: str, width: float, height: float) -> float:
    if y < 96:
        pdf.showPage()
        y = height - 54
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(54, y, title[:95])
    return y - 18


def pdf_line(pdf, y: float, text: str, x: float, width: float, height: float, *, font_size: int = 9, bold: bool = False) -> float:
    pdf.setFont("Helvetica-Bold" if bold else "Helvetica", font_size)
    max_chars = max(44, int((width - x - 54) / (font_size * 0.46)))
    for line in wrap_pdf_text(text, max_chars):
        if y < 54:
            pdf.showPage()
            y = height - 54
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", font_size)
        pdf.drawString(x, y, line)
        y -= font_size + 4
    return y


def wrap_pdf_text(text: Any, max_chars: int) -> list[str]:
    clean = " ".join(str(text or "").replace("\n", " ").split())
    if not clean:
        return [""]
    lines: list[str] = []
    current = ""
    for word in clean.split(" "):
        if len(current) + len(word) + 1 <= max_chars:
            current = f"{current} {word}".strip()
        else:
            if current:
                lines.append(current)
            current = word[:max_chars]
    if current:
        lines.append(current)
    return lines


def find_session(code: str | None) -> Session | None:
    normalized = "".join(ch for ch in str(code or "") if ch.isdigit())[:6]
    if len(normalized) != 6:
        return None
    return Session.query.filter_by(code=normalized).first()


def require_session(code: str) -> Session:
    session = find_session(code)
    if session is None:
        abort(404)
    return session


def require_question(session: Session, question_id: int) -> Question:
    question = db.session.get(Question, question_id)
    if question is None or question.session_id != session.id:
        abort(404)
    return question


def get_or_create_participant(session: Session, token: str | None) -> Participant:
    participant = None
    if token:
        participant = Participant.query.filter_by(session_id=session.id, token=token).first()
    if participant is None:
        participant = Participant(session=session, token=token or token_urlsafe(32), connected=True)
        db.session.add(participant)
        db.session.flush()
    participant.connected = True
    return participant


def generate_session_code() -> str:
    for _ in range(100):
        code = f"{randint(0, 999999):06d}"
        if Session.query.filter_by(code=code).first() is None:
            return code
    raise ValueError("No se pudo generar un código único.")


def next_question_position(session: Session) -> int:
    return max([question.position for question in session.questions] or [0]) + 1


def normalize_question_positions(session: Session) -> None:
    for index, question in enumerate(sorted(session.questions, key=lambda item: item.position), start=1):
        question.position = index


def reorder_questions(session: Session, question_ids: list[Any]) -> None:
    ordered_ids = [int(value) for value in question_ids]
    questions = sorted(session.questions, key=lambda item: item.position)
    existing_ids = [question.id for question in questions]
    if sorted(ordered_ids) != sorted(existing_ids):
        raise ValueError("El orden no coincide con las preguntas de la sesión.")
    by_id = {question.id: question for question in questions}
    for index, question_id in enumerate(ordered_ids, start=1):
        by_id[question_id].position = -index
    db.session.flush()
    for index, question_id in enumerate(ordered_ids, start=1):
        by_id[question_id].position = index
    clamp_active_index(session)


def clamp_active_index(session: Session) -> None:
    session.active_question_index = min(max(session.active_question_index, 0), max(len(session.questions) - 1, 0))


def room_name(code: str) -> str:
    return f"session:{code}"


def participant_count(session: Session) -> dict[str, int | str]:
    return {
        "code": session.code,
        "participant_count": len(session.participants),
        "connected_count": len([participant for participant in session.participants if participant.connected]),
    }


def broadcast_session(session: Session) -> None:
    socketio.emit("session_state", serialize_session(session), to=room_name(session.code))


def emit_results(session: Session, question: Question) -> None:
    socketio.emit(
        "results_updated",
        {"code": session.code, "question_id": question.id, "results": aggregate_question(question)},
        to=room_name(session.code),
    )


def quiz_score(question: Question, payload: dict[str, Any]) -> int:
    correct_ids = {option.id for option in question.options if option.is_correct}
    if int(payload.get("option_id") or 0) in correct_ids:
        return int((question.config_json or {}).get("points") or 100)
    return 0


def leaderboard(session: Session, run=RUN_DEFAULT) -> list[dict[str, Any]]:
    scores: defaultdict[int, int] = defaultdict(int)
    for response in scoped_session_responses(session, run, active_only=True):
        if response.question and response.question.type == "quiz":
            scores[response.participant_id] += int(response.score_awarded or 0)
    ordered = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    return [
        {"participant_id": participant_id, "label": f"Participante {index}", "score": score}
        for index, (participant_id, score) in enumerate(ordered[:10], start=1)
        if score > 0
    ]


def clean_text(value: Any, max_length: int, *, required: bool = False) -> str:
    text = str(value or "").strip()
    if required and len(text) < 1:
        raise ValueError("Captura el texto requerido.")
    if len(text) > max_length:
        raise ValueError(f"El texto no puede superar {max_length} caracteres.")
    return text


def parse_lines(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = str(value).replace("\r", "\n").split("\n")
    items = []
    for item in raw_items:
        text = clean_text(item, 180)
        if text and text not in items:
            items.append(text)
    return items[:12]


def clamp_int(value: Any, minimum: int, maximum: int) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        number = minimum
    return min(max(number, minimum), maximum)


def clamp_float(value: Any, minimum: float, maximum: float, default: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = default
    return round(min(max(number, minimum), maximum), 2)


def as_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "on", "yes", "si"}


def normalize_choice(value: Any, allowed: set[str], default: str) -> str:
    text = str(value or "").strip()
    return text if text in allowed else default


def sanitize_text_box_id(value: Any, index: int) -> str:
    text = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(value or "")).strip("-")[:40]
    return text or f"box-{index}"


def sanitize_hex_color(value: Any, default: str) -> str:
    if value in (None, ""):
        return default
    color = str(value).strip()
    if not TEXT_BOX_HEX_RE.match(color):
        raise ValueError("Color invalido en cuadro de texto.")
    return color.lower()


def sanitize_box_background(value: Any) -> str:
    if value in (None, "", "transparent"):
        return "transparent"
    return sanitize_hex_color(value, "transparent")


def normalize_font_weight(value: Any) -> int:
    if isinstance(value, bool):
        return 800 if value else 400
    text = str(value or "").strip().lower()
    if text in {"bold", "b", "true"}:
        return 800
    try:
        return 800 if int(float(text)) >= 600 else 400
    except (TypeError, ValueError):
        return 400


def parse_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def normalize_word(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def json_error(exc: Exception, status_code: int = 400):
    return jsonify({"ok": False, "error": str(exc)}), status_code


app = create_app()


if __name__ == "__main__":
    socketio.run(app, host="127.0.0.1", port=5000, debug=True)

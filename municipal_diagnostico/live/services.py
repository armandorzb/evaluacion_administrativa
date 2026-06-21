from __future__ import annotations

import base64
import json
import re
import secrets
from collections import Counter, defaultdict
from io import BytesIO
from typing import Any

import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from municipal_diagnostico.extensions import db
from municipal_diagnostico.live.schemas import (
    ACTIVITY_BRAINSTORM,
    ACTIVITY_CONTENT_SLIDE,
    ACTIVITY_MATRIX_2X2,
    ACTIVITY_MULTIPLE_CHOICE,
    ACTIVITY_POINTS_100,
    ACTIVITY_QA,
    ACTIVITY_QUIZ_CHOICE,
    ACTIVITY_QUIZ_TEXT,
    ACTIVITY_RANKING,
    ACTIVITY_SCALE,
    ActivityPayload,
    PresenterControlPayload,
    SessionPayload,
    TemplatePayload,
    normalize_response_payload,
)
from municipal_diagnostico.models import (
    LiveActivity,
    LiveParticipant,
    LiveReactivoTemplate,
    LiveResponse,
    LiveSession,
)
from municipal_diagnostico.timeutils import utcnow


LIVE_PARTICIPANT_COOKIE = "live_participant_token"
LIVE_CODE_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
STOPWORDS = {
    "con",
    "del",
    "las",
    "los",
    "para",
    "por",
    "que",
    "una",
    "uno",
    "como",
    "mas",
    "mas",
    "sin",
    "sus",
}
REPLACEABLE_RESPONSE_TYPES = {
    ACTIVITY_MULTIPLE_CHOICE,
    ACTIVITY_SCALE,
    ACTIVITY_RANKING,
    ACTIVITY_POINTS_100,
    ACTIVITY_MATRIX_2X2,
    ACTIVITY_QUIZ_CHOICE,
    ACTIVITY_QUIZ_TEXT,
}
QUIZ_TYPES = {ACTIVITY_QUIZ_CHOICE, ACTIVITY_QUIZ_TEXT}


class LiveValidationError(ValueError):
    pass


def room_name(session_id: int) -> str:
    return f"session:{session_id}"


def generate_session_code() -> str:
    while True:
        code = "".join(secrets.choice(LIVE_CODE_ALPHABET) for _ in range(6))
        if LiveSession.query.filter_by(code=code).first() is None:
            return code


def generate_participant_token() -> str:
    return secrets.token_urlsafe(32)


def create_template(raw_payload: dict[str, Any], user) -> LiveReactivoTemplate:
    payload = TemplatePayload.model_validate(raw_payload)
    template = LiveReactivoTemplate(
        tipo=payload.tipo,
        titulo=payload.titulo,
        prompt=payload.prompt,
        config_json=payload.config,
        creado_por=user if getattr(user, "is_authenticated", False) else None,
        activo=True,
    )
    db.session.add(template)
    return template


def update_template(template: LiveReactivoTemplate, raw_payload: dict[str, Any]) -> LiveReactivoTemplate:
    payload = TemplatePayload.model_validate(raw_payload)
    template.tipo = payload.tipo
    template.titulo = payload.titulo
    template.prompt = payload.prompt
    template.config_json = payload.config
    return template


def create_session(raw_payload: dict[str, Any], user) -> LiveSession:
    payload = SessionPayload.model_validate(raw_payload)
    session = LiveSession(
        titulo=payload.titulo,
        descripcion=payload.descripcion,
        mode=payload.mode,
        estado="draft",
        code=generate_session_code(),
        presentador=user if getattr(user, "is_authenticated", False) else None,
        config_json=payload.config,
    )
    db.session.add(session)
    db.session.flush()

    for template_id in payload.template_ids:
        template = db.session.get(LiveReactivoTemplate, int(template_id))
        if template is None or not template.activo:
            raise LiveValidationError("Uno de los reactivos seleccionados no existe o está inactivo.")
        add_activity_from_template(session, template)
    return session


def add_activity(session: LiveSession, raw_payload: dict[str, Any]) -> LiveActivity:
    payload = ActivityPayload.model_validate(raw_payload)
    if payload.template_id:
        template = db.session.get(LiveReactivoTemplate, int(payload.template_id))
        if template is None or not template.activo:
            raise LiveValidationError("El reactivo seleccionado no existe o está inactivo.")
        return add_activity_from_template(session, template)

    activity = LiveActivity(
        session=session,
        orden=next_activity_order(session),
        tipo=payload.tipo,
        titulo=payload.titulo,
        prompt=payload.prompt,
        config_json=payload.config,
        payload_json={},
        estado=initial_activity_state(session, payload.tipo),
    )
    db.session.add(activity)
    return activity


def add_activity_from_template(session: LiveSession, template: LiveReactivoTemplate) -> LiveActivity:
    activity = LiveActivity(
        session=session,
        template=template,
        orden=next_activity_order(session),
        tipo=template.tipo,
        titulo=template.titulo,
        prompt=template.prompt,
        config_json=dict(template.config_json or {}),
        payload_json={"template_id": template.id},
        estado=initial_activity_state(session, template.tipo),
    )
    db.session.add(activity)
    db.session.flush()
    return activity


def update_activity(activity: LiveActivity, raw_payload: dict[str, Any]) -> LiveActivity:
    payload = ActivityPayload.model_validate(
        {
            "tipo": raw_payload.get("tipo") or activity.tipo,
            "titulo": raw_payload.get("titulo") or activity.titulo,
            "prompt": raw_payload.get("prompt") or activity.prompt,
            "config": raw_payload.get("config") if raw_payload.get("config") is not None else activity.config_json,
        }
    )
    activity.tipo = payload.tipo
    activity.titulo = payload.titulo
    activity.prompt = payload.prompt
    activity.config_json = payload.config
    return activity


def duplicate_activity(session: LiveSession, activity_id: int) -> LiveActivity:
    source = get_session_activity(session, activity_id)
    duplicate = LiveActivity(
        session=session,
        template=source.template,
        orden=next_activity_order(session),
        tipo=source.tipo,
        titulo=f"{source.titulo} (copia)"[:180],
        prompt=source.prompt,
        config_json=dict(source.config_json or {}),
        payload_json=dict(source.payload_json or {}),
        estado=initial_activity_state(session, source.tipo),
    )
    duplicate.payload_json.pop("timer_started_at", None)
    duplicate.payload_json.pop("visited", None)
    db.session.add(duplicate)
    db.session.flush()

    ordered_ids = [activity.id for activity in sorted(session.activities, key=lambda item: item.orden)]
    ordered_ids.remove(duplicate.id)
    source_index = ordered_ids.index(source.id)
    ordered_ids.insert(source_index + 1, duplicate.id)
    reorder_activities(session, ordered_ids)
    return duplicate


def delete_activity(session: LiveSession, activity_id: int) -> None:
    activity = get_session_activity(session, activity_id)
    ordered = [item for item in sorted(session.activities, key=lambda item: item.orden) if item.id != activity.id]
    was_active = session.active_activity_id == activity.id
    next_active_id = None
    if was_active and ordered:
        later = [item for item in ordered if item.orden > activity.orden]
        next_active_id = (later[0] if later else ordered[-1]).id

    db.session.delete(activity)
    db.session.flush()
    db.session.expire(session, ["activities"])
    remaining = sorted(session.activities, key=lambda item: item.orden)
    if remaining:
        for index, item in enumerate(remaining, start=1):
            item.orden = -1000 - index
        db.session.flush()
        for index, item in enumerate(remaining, start=1):
            item.orden = index
    if was_active:
        session.active_activity_id = None
        if next_active_id:
            open_activity(session, next_active_id)


def reorder_activities(session: LiveSession, ordered_ids: list[int]) -> None:
    activities_by_id = {activity.id: activity for activity in session.activities}
    normalized_ids = [int(activity_id) for activity_id in ordered_ids]
    if len(normalized_ids) != len(set(normalized_ids)):
        raise LiveValidationError("El orden contiene diapositivas repetidas.")
    if set(normalized_ids) != set(activities_by_id):
        raise LiveValidationError("El orden debe incluir todas las diapositivas de la presentación.")

    for index, activity in enumerate(session.activities, start=1):
        activity.orden = -1000 - index
    db.session.flush()
    for index, activity_id in enumerate(normalized_ids, start=1):
        activities_by_id[activity_id].orden = index


def go_to_slide(session: LiveSession, activity_id: int) -> LiveActivity:
    activity = open_activity(session, activity_id)
    set_activity_payload(activity, {"visited": True})
    return activity


def next_slide(session: LiveSession) -> LiveActivity:
    ordered = sorted(session.activities, key=lambda item: item.orden)
    if not ordered:
        raise LiveValidationError("Agrega al menos una diapositiva para presentar.")
    current_index = current_slide_index(session, ordered)
    next_index = min(current_index + 1, len(ordered) - 1)
    return go_to_slide(session, ordered[next_index].id)


def previous_slide(session: LiveSession) -> LiveActivity:
    ordered = sorted(session.activities, key=lambda item: item.orden)
    if not ordered:
        raise LiveValidationError("Agrega al menos una diapositiva para presentar.")
    current_index = current_slide_index(session, ordered)
    previous_index = max(current_index - 1, 0)
    return go_to_slide(session, ordered[previous_index].id)


def current_slide_index(session: LiveSession, ordered: list[LiveActivity]) -> int:
    if session.active_activity_id:
        for index, activity in enumerate(ordered):
            if activity.id == session.active_activity_id:
                return index
    for index, activity in enumerate(ordered):
        if activity.estado == "open":
            return index
    return 0


def initial_activity_state(session: LiveSession, activity_type: str | None) -> str:
    if session.estado == "active" and session.mode == "self_paced" and activity_type not in QUIZ_TYPES:
        return "open"
    return "draft"


def next_activity_order(session: LiveSession) -> int:
    if not session.activities:
        return 1
    return max(activity.orden for activity in session.activities) + 1


def find_session_by_code(code: str | None) -> LiveSession | None:
    normalized = (code or "").strip().upper()
    if not normalized:
        return None
    return LiveSession.query.filter_by(code=normalized).first()


def get_or_create_participant(session: LiveSession, token: str | None = None) -> LiveParticipant:
    token = (token or "").strip() or generate_participant_token()
    participant = LiveParticipant.query.filter_by(session_id=session.id, token=token).first()
    if participant is None:
        participant = LiveParticipant(session=session, token=token, ultima_actividad_at=utcnow(), connected=False)
        db.session.add(participant)
        db.session.flush()
    else:
        participant.ultima_actividad_at = utcnow()
    return participant


def mark_participant_seen(session: LiveSession, token: str | None, *, connected: bool = True) -> LiveParticipant:
    participant = get_or_create_participant(session, token)
    participant.connected = connected
    participant.ultima_actividad_at = utcnow()
    return participant


def submit_response(
    session: LiveSession,
    activity: LiveActivity,
    participant: LiveParticipant,
    raw_payload: dict[str, Any],
) -> LiveResponse:
    if activity.session_id != session.id:
        raise LiveValidationError("La actividad no pertenece a esta sesión.")
    if session.estado != "active":
        raise LiveValidationError("La sesión no está abierta.")
    if activity.estado != "open":
        raise LiveValidationError("La actividad no está abierta.")
    if activity.tipo in QUIZ_TYPES and session.mode != "guided":
        raise LiveValidationError("Los quizzes competitivos requieren modo guiado.")
    if session.mode == "guided" and session.active_activity_id != activity.id:
        raise LiveValidationError("Esta actividad no está activa en este momento.")

    payload = normalize_response_payload(activity.tipo, activity.config_json or {}, raw_payload)
    participant.ultima_actividad_at = utcnow()

    if activity.tipo in REPLACEABLE_RESPONSE_TYPES:
        response_key = "quiz" if activity.tipo in QUIZ_TYPES else activity.tipo
        return upsert_response(session, activity, participant, response_key, payload)

    if activity.tipo == ACTIVITY_BRAINSTORM:
        max_ideas = int((activity.config_json or {}).get("max_ideas_per_participant") or 5)
        active_ideas = LiveResponse.query.filter_by(
            activity_id=activity.id,
            participant_id=participant.id,
            is_active=True,
        ).count()
        if active_ideas >= max_ideas:
            raise LiveValidationError(f"Puedes enviar hasta {max_ideas} ideas en esta actividad.")
        payload["status"] = "pending" if (activity.config_json or {}).get("moderation") == "manual" else "approved"
        response = LiveResponse(
            session=session,
            activity=activity,
            participant=participant,
            response_key=secrets.token_urlsafe(8),
            payload_json=payload,
            is_active=True,
        )
        db.session.add(response)
        return response

    if activity.tipo == ACTIVITY_QA:
        payload.update(
            {
                "status": "pending" if (activity.config_json or {}).get("moderation") == "manual" else "approved",
                "upvotes": 0,
                "upvoter_tokens": [],
                "shown": False,
                "answered": False,
            }
        )
        response = LiveResponse(
            session=session,
            activity=activity,
            participant=participant,
            response_key=secrets.token_urlsafe(8),
            payload_json=payload,
            is_active=True,
        )
        db.session.add(response)
        return response

    raise LiveValidationError("Tipo de actividad no soportado.")


def upsert_response(
    session: LiveSession,
    activity: LiveActivity,
    participant: LiveParticipant,
    response_key: str,
    payload: dict[str, Any],
) -> LiveResponse:
    response = LiveResponse.query.filter_by(
        activity_id=activity.id,
        participant_id=participant.id,
        response_key=response_key,
    ).first()
    if response is None:
        response = LiveResponse(
            session=session,
            activity=activity,
            participant=participant,
            response_key=response_key,
            payload_json=payload,
            is_active=True,
        )
        db.session.add(response)
    else:
        response.payload_json = payload
        response.is_active = True
    return response


def apply_presenter_control(session: LiveSession, raw_payload: dict[str, Any]) -> LiveSession:
    payload = PresenterControlPayload.model_validate(raw_payload)
    if payload.session_id != session.id:
        raise LiveValidationError("La sesión no coincide con el control solicitado.")

    if payload.action == "open_session":
        open_session(session)
    elif payload.action == "close_session":
        close_session(session)
    elif payload.action == "next_slide":
        next_slide(session)
    elif payload.action == "previous_slide":
        previous_slide(session)
    elif payload.action == "go_to_slide":
        if not payload.activity_id:
            raise LiveValidationError("Selecciona una diapositiva para presentar.")
        go_to_slide(session, payload.activity_id)
    elif payload.action == "open_activity":
        if not payload.activity_id:
            raise LiveValidationError("Selecciona una actividad para abrir.")
        go_to_slide(session, payload.activity_id)
    elif payload.action == "close_activity":
        if not payload.activity_id:
            raise LiveValidationError("Selecciona una actividad para cerrar.")
        close_activity(session, payload.activity_id)
    elif payload.action == "set_mode":
        if not payload.mode:
            raise LiveValidationError("Selecciona un modo de sesión.")
        session.mode = payload.mode
        if session.estado == "active":
            configure_open_activities_for_mode(session)
    elif payload.action in {"reveal_results", "hide_results"}:
        activity = get_session_activity(session, payload.activity_id)
        set_activity_payload(activity, {"show_results": payload.action == "reveal_results"})
    elif payload.action == "set_timer":
        activity = get_session_activity(session, payload.activity_id)
        seconds = payload.timer_seconds
        if seconds is None:
            seconds = int((activity.config_json or {}).get("timer_seconds") or 0)
        set_activity_payload(activity, {"timer_seconds": seconds, "timer_started_at": utcnow().isoformat()})
    elif payload.action == "show_question":
        activity = get_session_activity(session, payload.activity_id)
        if not payload.response_id:
            raise LiveValidationError("Selecciona una pregunta para mostrar.")
        moderate_response(session, activity, payload.response_id, "show")
    return session


def open_session(session: LiveSession) -> None:
    session.estado = "active"
    session.closed_at = None
    if session.opened_at is None:
        session.opened_at = utcnow()
    configure_open_activities_for_mode(session)


def close_session(session: LiveSession) -> None:
    session.estado = "closed"
    session.closed_at = utcnow()
    session.active_activity_id = None
    for activity in session.activities:
        if activity.estado != "closed":
            activity.estado = "closed"
            activity.closed_at = utcnow()


def configure_open_activities_for_mode(session: LiveSession) -> None:
    if not session.activities:
        session.active_activity_id = None
        return
    if session.mode == "self_paced":
        session.active_activity_id = None
        for activity in session.activities:
            if activity.tipo in QUIZ_TYPES:
                activity.estado = "draft"
                continue
            if activity.estado != "closed":
                activity.estado = "open"
                activity.opened_at = activity.opened_at or utcnow()
        return

    selected = None
    if session.active_activity_id:
        selected = next((activity for activity in session.activities if activity.id == session.active_activity_id), None)
    if selected is None or selected.estado == "closed":
        selected = next((activity for activity in session.activities if activity.estado != "closed"), session.activities[0])
    for activity in session.activities:
        if activity.id == selected.id:
            activity.estado = "open"
            activity.opened_at = activity.opened_at or utcnow()
            activity.closed_at = None
            session.active_activity_id = activity.id
        elif activity.estado == "open":
            activity.estado = "draft"


def open_activity(session: LiveSession, activity_id: int) -> LiveActivity:
    activity = db.session.get(LiveActivity, int(activity_id))
    if activity is None or activity.session_id != session.id:
        raise LiveValidationError("La actividad solicitada no existe.")
    if activity.tipo in QUIZ_TYPES and session.mode != "guided":
        raise LiveValidationError("Los quizzes competitivos requieren modo guiado.")
    if session.estado != "active":
        session.estado = "active"
        session.opened_at = session.opened_at or utcnow()
        session.closed_at = None
    if session.mode == "guided":
        for current in session.activities:
            if current.id != activity.id and current.estado == "open":
                current.estado = "draft"
    session.active_activity_id = activity.id
    activity.estado = "open"
    activity.opened_at = activity.opened_at or utcnow()
    activity.closed_at = None
    return activity


def close_activity(session: LiveSession, activity_id: int) -> LiveActivity:
    activity = db.session.get(LiveActivity, int(activity_id))
    if activity is None or activity.session_id != session.id:
        raise LiveValidationError("La actividad solicitada no existe.")
    activity.estado = "closed"
    activity.closed_at = utcnow()
    if session.active_activity_id == activity.id:
        session.active_activity_id = None
    return activity


def get_session_activity(session: LiveSession, activity_id: int | None) -> LiveActivity:
    if not activity_id:
        raise LiveValidationError("Selecciona una actividad.")
    activity = db.session.get(LiveActivity, int(activity_id))
    if activity is None or activity.session_id != session.id:
        raise LiveValidationError("La actividad solicitada no existe.")
    return activity


def set_activity_payload(activity: LiveActivity, values: dict[str, Any]) -> None:
    payload = dict(activity.payload_json or {})
    payload.update(values)
    activity.payload_json = payload


def moderate_response(session: LiveSession, activity: LiveActivity, response_id: int, action: str) -> LiveResponse:
    if activity.session_id != session.id:
        raise LiveValidationError("La actividad no pertenece a esta sesión.")
    response = db.session.get(LiveResponse, int(response_id))
    if response is None or response.activity_id != activity.id:
        raise LiveValidationError("La respuesta solicitada no existe.")
    if activity.tipo not in {ACTIVITY_QA, ACTIVITY_BRAINSTORM}:
        raise LiveValidationError("Esta actividad no admite moderación.")

    payload = dict(response.payload_json or {})
    if action == "approve":
        payload["status"] = "approved"
    elif action == "reject":
        payload["status"] = "rejected"
    elif action in {"answer", "answered"}:
        payload["status"] = "answered"
        payload["answered"] = True
    elif action == "show":
        payload["status"] = "approved" if payload.get("status") == "pending" else payload.get("status", "approved")
        payload["shown"] = True
        set_activity_payload(activity, {"shown_response_id": response.id})
    else:
        raise LiveValidationError("Acción de moderación no soportada.")
    response.payload_json = payload
    return response


def upvote_response(session: LiveSession, activity: LiveActivity, response_id: int, participant: LiveParticipant) -> LiveResponse:
    if activity.session_id != session.id:
        raise LiveValidationError("La actividad no pertenece a esta sesión.")
    if activity.tipo != ACTIVITY_QA:
        raise LiveValidationError("Solo Q&A admite votos a favor.")
    if not (activity.config_json or {}).get("allow_upvotes", True):
        raise LiveValidationError("Los votos están desactivados para esta actividad.")
    response = db.session.get(LiveResponse, int(response_id))
    if response is None or response.activity_id != activity.id:
        raise LiveValidationError("La pregunta solicitada no existe.")

    payload = dict(response.payload_json or {})
    upvoters = list(payload.get("upvoter_tokens") or [])
    if participant.token not in upvoters:
        upvoters.append(participant.token)
    payload["upvoter_tokens"] = upvoters
    payload["upvotes"] = len(upvoters)
    response.payload_json = payload
    return response


def serialize_session(session: LiveSession) -> dict[str, Any]:
    return {
        "id": session.id,
        "titulo": session.titulo,
        "descripcion": session.descripcion,
        "code": session.code,
        "mode": session.mode,
        "estado": session.estado,
        "active_activity_id": session.active_activity_id,
        "participant_count": len(session.participants),
        "connected_count": len([participant for participant in session.participants if participant.connected]),
        "activities": [serialize_activity(activity) for activity in session.activities],
    }


def serialize_activity(activity: LiveActivity, *, include_results: bool = True) -> dict[str, Any]:
    payload = {
        "id": activity.id,
        "session_id": activity.session_id,
        "template_id": activity.template_id,
        "orden": activity.orden,
        "tipo": activity.tipo,
        "titulo": activity.titulo,
        "prompt": activity.prompt,
        "estado": activity.estado,
        "config": activity.config_json or {},
        "payload": activity_runtime_payload(activity),
    }
    if include_results:
        payload["results"] = aggregate_results(activity)
    return payload


def activity_runtime_payload(activity: LiveActivity) -> dict[str, Any]:
    config = activity.config_json or {}
    payload = dict(activity.payload_json or {})
    payload.setdefault("show_results", bool(config.get("show_results", True)))
    if "timer_seconds" not in payload and config.get("timer_seconds"):
        payload["timer_seconds"] = int(config.get("timer_seconds") or 0)
    return payload


def aggregate_results(activity: LiveActivity) -> dict[str, Any]:
    responses = [response for response in activity.responses if response.is_active]
    if activity.tipo == ACTIVITY_CONTENT_SLIDE:
        return {"type": ACTIVITY_CONTENT_SLIDE, "total": 0}
    if activity.tipo == ACTIVITY_MULTIPLE_CHOICE:
        options = list((activity.config_json or {}).get("options", []))
        counts = Counter(str(response.payload_json.get("choice")) for response in responses)
        return {
            "type": ACTIVITY_MULTIPLE_CHOICE,
            "total": sum(counts.values()),
            "options": [{"label": option, "count": counts.get(option, 0)} for option in options],
        }
    if activity.tipo == ACTIVITY_BRAINSTORM:
        ideas = [
            {
                "id": response.id,
                "idea": response.payload_json.get("idea", ""),
                "status": response.payload_json.get("status", "approved"),
                "created_at": response.created_at.isoformat() if response.created_at else None,
            }
            for response in responses
            if response.payload_json.get("idea") and response.payload_json.get("status", "approved") != "rejected"
        ]
        visible_ideas = [idea for idea in ideas if idea["status"] in {"approved", "answered"}]
        return {
            "type": ACTIVITY_BRAINSTORM,
            "total": len(ideas),
            "approved_total": len(visible_ideas),
            "ideas": ideas,
            "words": build_word_frequencies([idea["idea"] for idea in visible_ideas or ideas]),
        }
    if activity.tipo == ACTIVITY_SCALE:
        return aggregate_scale(activity, responses)
    if activity.tipo == ACTIVITY_RANKING:
        return aggregate_ranking(activity, responses)
    if activity.tipo == ACTIVITY_POINTS_100:
        return aggregate_points(activity, responses)
    if activity.tipo == ACTIVITY_MATRIX_2X2:
        return aggregate_matrix(activity, responses)
    if activity.tipo == ACTIVITY_QA:
        return aggregate_qa(activity, responses)
    if activity.tipo == ACTIVITY_QUIZ_CHOICE:
        return aggregate_quiz_choice(activity, responses)
    if activity.tipo == ACTIVITY_QUIZ_TEXT:
        return aggregate_quiz_text(activity, responses)
    return {"type": activity.tipo, "total": len(responses)}


def aggregate_scale(activity: LiveActivity, responses: list[LiveResponse]) -> dict[str, Any]:
    buckets: dict[str, list[float]] = defaultdict(list)
    for response in responses:
        for item, value in (response.payload_json.get("ratings") or {}).items():
            buckets[str(item)].append(float(value))
    items = []
    for item in (activity.config_json or {}).get("items", []):
        values = buckets.get(str(item), [])
        items.append(
            {
                "label": item,
                "count": len(values),
                "average": round(sum(values) / len(values), 2) if values else 0,
                "min": min(values) if values else None,
                "max": max(values) if values else None,
            }
        )
    return {"type": ACTIVITY_SCALE, "total": len(responses), "items": items}


def aggregate_ranking(activity: LiveActivity, responses: list[LiveResponse]) -> dict[str, Any]:
    items = [str(item) for item in (activity.config_json or {}).get("items", [])]
    max_ranked = int((activity.config_json or {}).get("max_ranked") or len(items))
    scores = Counter({item: 0 for item in items})
    first_places = Counter({item: 0 for item in items})
    for response in responses:
        ranking = [str(item) for item in response.payload_json.get("ranking", [])]
        for index, item in enumerate(ranking[:max_ranked]):
            scores[item] += max(max_ranked - index, 1)
            if index == 0:
                first_places[item] += 1
    ordered = [
        {"label": item, "score": scores.get(item, 0), "first_places": first_places.get(item, 0)}
        for item in items
    ]
    ordered.sort(key=lambda item: item["score"], reverse=True)
    return {"type": ACTIVITY_RANKING, "total": len(responses), "items": ordered}


def aggregate_points(activity: LiveActivity, responses: list[LiveResponse]) -> dict[str, Any]:
    items = [str(item) for item in (activity.config_json or {}).get("items", [])]
    totals = Counter({item: 0 for item in items})
    for response in responses:
        for item, value in (response.payload_json.get("points") or {}).items():
            totals[str(item)] += int(value or 0)
    ordered = [
        {
            "label": item,
            "points": totals.get(item, 0),
            "average": round(totals.get(item, 0) / len(responses), 2) if responses else 0,
        }
        for item in items
    ]
    ordered.sort(key=lambda item: item["points"], reverse=True)
    return {"type": ACTIVITY_POINTS_100, "total": len(responses), "items": ordered}


def aggregate_matrix(activity: LiveActivity, responses: list[LiveResponse]) -> dict[str, Any]:
    buckets: dict[str, list[dict[str, float]]] = defaultdict(list)
    points = []
    for response in responses:
        for item, coords in (response.payload_json.get("ratings") or {}).items():
            point = {"label": str(item), "x": float(coords.get("x", 0)), "y": float(coords.get("y", 0))}
            buckets[str(item)].append(point)
            points.append(point)
    items = []
    for item in (activity.config_json or {}).get("items", []):
        values = buckets.get(str(item), [])
        items.append(
            {
                "label": item,
                "count": len(values),
                "x": round(sum(point["x"] for point in values) / len(values), 2) if values else 0,
                "y": round(sum(point["y"] for point in values) / len(values), 2) if values else 0,
            }
        )
    return {
        "type": ACTIVITY_MATRIX_2X2,
        "total": len(responses),
        "items": items,
        "points": points,
        "x_axis": (activity.config_json or {}).get("x_axis", {}),
        "y_axis": (activity.config_json or {}).get("y_axis", {}),
        "min": (activity.config_json or {}).get("min", -5),
        "max": (activity.config_json or {}).get("max", 5),
    }


def aggregate_qa(activity: LiveActivity, responses: list[LiveResponse]) -> dict[str, Any]:
    questions = []
    status_counts = Counter()
    for response in responses:
        payload = response.payload_json or {}
        status = str(payload.get("status") or "approved")
        status_counts[status] += 1
        if status == "rejected":
            continue
        questions.append(
            {
                "id": response.id,
                "question": payload.get("question", ""),
                "status": status,
                "upvotes": int(payload.get("upvotes") or 0),
                "shown": bool(payload.get("shown")),
                "answered": bool(payload.get("answered")),
                "created_at": response.created_at.isoformat() if response.created_at else None,
            }
        )
    questions.sort(key=lambda item: (item["shown"], item["upvotes"], item["created_at"] or ""), reverse=True)
    return {
        "type": ACTIVITY_QA,
        "total": len(responses),
        "questions": questions,
        "pending": status_counts.get("pending", 0),
        "approved": status_counts.get("approved", 0),
        "answered": status_counts.get("answered", 0),
        "rejected": status_counts.get("rejected", 0),
    }


def aggregate_quiz_choice(activity: LiveActivity, responses: list[LiveResponse]) -> dict[str, Any]:
    options = list((activity.config_json or {}).get("options", []))
    counts = Counter(str(response.payload_json.get("choice")) for response in responses)
    correct_count = sum(1 for response in responses if response.payload_json.get("is_correct"))
    return {
        "type": ACTIVITY_QUIZ_CHOICE,
        "total": len(responses),
        "correct": correct_count,
        "options": [{"label": option, "count": counts.get(option, 0)} for option in options],
        "leaderboard": build_leaderboard(responses),
    }


def aggregate_quiz_text(activity: LiveActivity, responses: list[LiveResponse]) -> dict[str, Any]:
    answers = Counter(str(response.payload_json.get("answer") or "").strip() for response in responses)
    correct_count = sum(1 for response in responses if response.payload_json.get("is_correct"))
    return {
        "type": ACTIVITY_QUIZ_TEXT,
        "total": len(responses),
        "correct": correct_count,
        "answers": [{"label": answer, "count": count} for answer, count in answers.most_common(20) if answer],
        "leaderboard": build_leaderboard(responses),
    }


def build_leaderboard(responses: list[LiveResponse]) -> list[dict[str, Any]]:
    by_participant: dict[int, dict[str, Any]] = {}
    for response in responses:
        entry = by_participant.setdefault(
            response.participant_id,
            {"participant_id": response.participant_id, "score": 0, "correct": 0, "answers": 0},
        )
        entry["score"] += int((response.payload_json or {}).get("score_awarded") or 0)
        entry["correct"] += 1 if (response.payload_json or {}).get("is_correct") else 0
        entry["answers"] += 1
    ordered = sorted(by_participant.values(), key=lambda item: (item["score"], item["correct"]), reverse=True)
    for index, item in enumerate(ordered, start=1):
        item["rank"] = index
        item["label"] = f"Participante {index}"
    return ordered[:25]


def build_word_frequencies(ideas: list[str]) -> list[list[Any]]:
    counter: Counter[str] = Counter()
    for idea in ideas:
        for word in re.findall(r"[\w]+", idea.lower()):
            normalized = word.strip("_")
            if len(normalized) >= 3 and normalized not in STOPWORDS:
                counter[normalized] += 1
    return [[word, count] for word, count in counter.most_common(60)]


def build_qr_data_uri(url: str) -> str:
    buffer = build_qr_png(url)
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def build_qr_png(url: str) -> BytesIO:
    image = qrcode.make(url)
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


def build_live_excel(session: LiveSession) -> BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.append(["Sesión", session.titulo])
    summary.append(["Código", session.code])
    summary.append(["Modo", session.mode])
    summary.append(["Estado", session.estado])
    summary.append(["Participantes", len(session.participants)])

    activities = workbook.create_sheet("Actividades")
    activities.append(["Orden", "Título", "Tipo", "Estado", "Respuestas", "Agregado JSON"])
    for activity in session.activities:
        results = aggregate_results(activity)
        activities.append(
            [
                activity.orden,
                activity.titulo,
                activity.tipo,
                activity.estado,
                results.get("total", 0),
                json.dumps(results, ensure_ascii=False),
            ]
        )

    responses = workbook.create_sheet("Respuestas")
    responses.append(["Actividad", "Tipo", "Participante anónimo", "Llave", "Activa", "Payload JSON", "Creada"])
    for response in session.responses:
        responses.append(
            [
                response.activity.titulo if response.activity else "",
                response.activity.tipo if response.activity else "",
                response.participant_id,
                response.response_key,
                "si" if response.is_active else "no",
                json.dumps(response.payload_json or {}, ensure_ascii=False),
                response.created_at.isoformat() if response.created_at else "",
            ]
        )

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5A74")
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_live_pdf(session: LiveSession) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title=f"Live {session.code}")
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Sesión Live: {session.titulo}", styles["Title"]),
        Paragraph(f"Código: {session.code} | Modo: {session.mode} | Estado: {session.estado}", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [["Orden", "Actividad", "Tipo", "Respuestas"]]
    for activity in session.activities:
        data.append([str(activity.orden), activity.titulo, activity.tipo, str(aggregate_results(activity).get("total", 0))])
    table = Table(data, repeatRows=1, colWidths=[45, 240, 120, 80])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5A74")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7DEE2")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 12))
    for activity in session.activities:
        results = aggregate_results(activity)
        story.append(Paragraph(activity.titulo, styles["Heading2"]))
        story.append(Paragraph(activity.prompt, styles["Normal"]))
        story.append(Paragraph(f"Tipo: {activity.tipo} | Respuestas: {results.get('total', 0)}", styles["Normal"]))
        story.append(Spacer(1, 8))
    doc.build(story)
    buffer.seek(0)
    return buffer

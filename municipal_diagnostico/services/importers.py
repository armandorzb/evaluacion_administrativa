from __future__ import annotations

import csv
from io import StringIO

from openpyxl import load_workbook

from municipal_diagnostico.extensions import db
from municipal_diagnostico.models import Area, Dependencia, Usuario
from municipal_diagnostico.services.module_access import WELLBEING_ALLOWED_ROLES, normalize_module_flags, parse_optional_flag


def load_rows(file_storage) -> list[dict]:
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        text = file_storage.stream.read().decode("utf-8-sig")
        return list(csv.DictReader(StringIO(text)))

    workbook = load_workbook(file_storage, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    payload = []
    for row in rows[1:]:
        payload.append(
            {
                headers[index]: (value.strip() if isinstance(value, str) else value)
                for index, value in enumerate(row)
                if headers[index]
            }
        )
    return payload


def import_dependencias(rows: list[dict]) -> dict:
    created = updated = 0
    errors = []
    for row in rows:
        name = str(row.get("nombre", "")).strip()
        if not name:
            errors.append("Fila sin nombre de dependencia.")
            continue
        dependency = Dependencia.query.filter_by(nombre=name).first()
        if dependency:
            updated += 1
        else:
            dependency = Dependencia(nombre=name)
            db.session.add(dependency)
            created += 1
        dependency.tipo = str(row.get("tipo", "Administrativa")).strip() or "Administrativa"
        dependency.descripcion = str(row.get("descripcion", "")).strip() or None
        dependency.activa = True
    db.session.commit()
    return {"created": created, "updated": updated, "errors": errors}


def import_areas(rows: list[dict]) -> dict:
    created = updated = 0
    errors = []
    for row in rows:
        dependency_name = str(row.get("dependencia", "")).strip()
        area_name = str(row.get("nombre", "")).strip()
        if not dependency_name or not area_name:
            errors.append("Fila sin dependencia o nombre de área.")
            continue
        dependency = Dependencia.query.filter_by(nombre=dependency_name).first()
        if not dependency:
            errors.append(f"Dependencia no encontrada: {dependency_name}")
            continue
        area = Area.query.filter_by(dependencia_id=dependency.id, nombre=area_name).first()
        if area:
            updated += 1
        else:
            area = Area(nombre=area_name, dependencia=dependency)
            db.session.add(area)
            created += 1
        area.activa = True
    db.session.commit()
    return {"created": created, "updated": updated, "errors": errors}


def import_usuarios(rows: list[dict]) -> dict:
    created = updated = 0
    errors = []
    for row in rows:
        email = str(row.get("correo", "")).strip().lower()
        role = str(row.get("rol", "")).strip().lower()
        password = str(row.get("password", "")).strip()
        if not email or role not in {"administrador", "revisor", "evaluador", "respondente", "consulta"}:
            errors.append("Fila con correo o rol inválido.")
            continue
        if not password and not Usuario.query.filter_by(correo=email).first():
            errors.append(f"Usuario nuevo sin contraseña: {email}")
            continue

        try:
            requested_diagnostic = parse_optional_flag(row.get("acceso_diagnostico"))
            requested_wellbeing = parse_optional_flag(row.get("acceso_bienestar"))
        except ValueError:
            errors.append(f"Valor de acceso inválido para usuario {email}. Usa si/no, true/false o 1/0.")
            continue
        if requested_wellbeing is True and role not in WELLBEING_ALLOWED_ROLES:
            errors.append(f"Bienestar Policial solo puede asignarse a administrador o consulta: {email}")
            continue

        module_access = normalize_module_flags(role, requested_diagnostic, requested_wellbeing)
        if not module_access["acceso_diagnostico"] and not module_access["acceso_bienestar"]:
            errors.append(f"El usuario {email} debe conservar al menos un módulo activo.")
            continue

        dependency = None
        area = None
        dependency_name = str(row.get("dependencia", "")).strip()
        area_name = str(row.get("area", "")).strip()
        if dependency_name:
            dependency = Dependencia.query.filter_by(nombre=dependency_name).first()
            if not dependency:
                errors.append(f"Dependencia no encontrada para usuario {email}: {dependency_name}")
                continue
        if area_name and dependency:
            area = Area.query.filter_by(dependencia_id=dependency.id, nombre=area_name).first()
            if not area:
                errors.append(f"Área no encontrada para usuario {email}: {area_name}")
                continue

        user = Usuario.query.filter_by(correo=email).first()
        if user:
            updated += 1
        else:
            user = Usuario(correo=email, nombre=str(row.get("nombre", email)).strip())
            db.session.add(user)
            created += 1

        user.nombre = str(row.get("nombre", user.nombre)).strip() or user.nombre
        user.rol = role
        user.dependencia = dependency
        user.area = area
        user.acceso_diagnostico = module_access["acceso_diagnostico"]
        user.acceso_bienestar = module_access["acceso_bienestar"]
        user.activo = True
        if password:
            user.set_password(password)

    db.session.commit()
    return {"created": created, "updated": updated, "errors": errors}

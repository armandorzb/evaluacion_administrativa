import pytest
from pydantic import ValidationError
from io import BytesIO
from openpyxl import load_workbook

from municipal_diagnostico import create_app
from municipal_diagnostico.extensions import db, socketio
from municipal_diagnostico.live.schemas import TemplatePayload, normalize_response_payload
from municipal_diagnostico.live.services import (
    add_activity,
    add_activity_from_template,
    aggregate_results,
    apply_presenter_control,
    create_session,
    create_template,
    duplicate_activity,
    open_session,
    reorder_activities,
)
from municipal_diagnostico.models import LiveActivity, LiveParticipant, LiveResponse, LiveSession, Usuario


class TestConfig:
    SECRET_KEY = "test"
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = "tests/uploads"
    ALLOWED_EXTENSIONS = {"pdf"}
    AUTO_INIT_DATABASE = True
    BOOTSTRAP_ADMIN_EMAIL = None
    BOOTSTRAP_ADMIN_PASSWORD = None
    BOOTSTRAP_ADMIN_NAME = None


def login(client, email: str, password: str = "secret123"):
    return client.post(
        "/auth/login",
        data={"correo": email, "password": password},
        follow_redirects=True,
    )


def build_app():
    app = create_app(TestConfig)
    with app.app_context():
        admin = Usuario(
            nombre="Admin Live",
            correo="admin-live@test.local",
            rol="administrador",
            activo=True,
            acceso_diagnostico=False,
            acceso_bienestar=False,
            acceso_iso9001=False,
            acceso_live=True,
        )
        admin.set_password("secret123")
        consulta = Usuario(
            nombre="Consulta Live",
            correo="consulta-live@test.local",
            rol="consulta",
            activo=True,
            acceso_diagnostico=False,
            acceso_bienestar=False,
            acceso_iso9001=False,
            acceso_live=True,
        )
        consulta.set_password("secret123")
        denied = Usuario(
            nombre="Sin Live",
            correo="sin-live@test.local",
            rol="consulta",
            activo=True,
            acceso_diagnostico=True,
            acceso_bienestar=False,
            acceso_iso9001=False,
            acceso_live=False,
        )
        denied.set_password("secret123")
        db.session.add_all([admin, consulta, denied])
        db.session.commit()
    return app


def test_live_pydantic_validation_normalizes_supported_types():
    brainstorm = TemplatePayload.model_validate(
        {
            "tipo": "brainstorm",
            "titulo": "Ideas",
            "prompt": "Comparte una mejora",
            "config": {"max_ideas_per_participant": "2", "max_length": "80"},
        }
    )
    assert brainstorm.config["max_ideas_per_participant"] == 2
    assert brainstorm.config["max_length"] == 80

    multiple = TemplatePayload.model_validate(
        {
            "tipo": "multiple_choice",
            "titulo": "Prioridad",
            "prompt": "Elige una prioridad",
            "config": {"options": ["Seguridad", "Servicios"]},
        }
    )
    assert multiple.config["options"] == ["Seguridad", "Servicios"]
    assert normalize_response_payload("multiple_choice", multiple.config, {"choice": "Seguridad"}) == {"choice": "Seguridad"}

    scale = TemplatePayload.model_validate(
        {
            "tipo": "scale",
            "titulo": "Madurez",
            "prompt": "Evalua capacidades",
            "config": {"items": ["Procesos", "Datos"], "min": 1, "max": 5},
        }
    )
    assert normalize_response_payload("scale", scale.config, {"ratings": {"Procesos": 4, "Datos": 5}})["ratings"]["Datos"] == 5

    ranking = TemplatePayload.model_validate(
        {
            "tipo": "ranking",
            "titulo": "Priorizar",
            "prompt": "Ordena prioridades",
            "config": {"items": ["A", "B", "C"], "max_ranked": 2},
        }
    )
    assert normalize_response_payload("ranking", ranking.config, {"ranking": ["B", "A"]}) == {"ranking": ["B", "A"]}

    points = TemplatePayload.model_validate(
        {
            "tipo": "points_100",
            "titulo": "Asignar recursos",
            "prompt": "Distribuye puntos",
            "config": {"items": ["A", "B"], "total_points": 100, "step": 10},
        }
    )
    assert normalize_response_payload("points_100", points.config, {"points": {"A": 60, "B": 40}})["points"]["A"] == 60

    matrix = TemplatePayload.model_validate(
        {
            "tipo": "matrix_2x2",
            "titulo": "Stakeholders",
            "prompt": "Ubica actores",
            "config": {"items": ["Cabildo"], "min": -5, "max": 5},
        }
    )
    assert normalize_response_payload("matrix_2x2", matrix.config, {"ratings": {"Cabildo": {"x": 4, "y": 3}}})["ratings"]["Cabildo"]["x"] == 4

    qa = TemplatePayload.model_validate(
        {"tipo": "qa", "titulo": "Preguntas", "prompt": "Haz preguntas", "config": {"moderation": "manual"}}
    )
    assert normalize_response_payload("qa", qa.config, {"question": "Que sigue?"}) == {"question": "Que sigue?"}

    quiz_choice = TemplatePayload.model_validate(
        {
            "tipo": "quiz_choice",
            "titulo": "Quiz",
            "prompt": "Elige",
            "config": {"options": ["Si", "No"], "correct_options": ["Si"]},
        }
    )
    assert normalize_response_payload("quiz_choice", quiz_choice.config, {"choice": "Si"})["is_correct"] is True

    quiz_text = TemplatePayload.model_validate(
        {
            "tipo": "quiz_text",
            "titulo": "Quiz texto",
            "prompt": "Responde",
            "config": {"answers": ["Hermosillo"]},
        }
    )
    assert normalize_response_payload("quiz_text", quiz_text.config, {"answer": "hermosillo"})["is_correct"] is True

    content = TemplatePayload.model_validate(
        {
            "tipo": "content_slide",
            "titulo": "Bienvenida",
            "prompt": "Instrucciones iniciales",
            "config": {"layout": "qr", "body": "Escanea y espera al presentador.", "timer_seconds": "45"},
        }
    )
    assert content.config["layout"] == "qr"
    assert content.config["timer_seconds"] == 45
    with pytest.raises(ValueError):
        normalize_response_payload("content_slide", content.config, {})

    with pytest.raises(ValidationError):
        TemplatePayload.model_validate(
            {
                "tipo": "multiple_choice",
                "titulo": "Mal",
                "prompt": "Sin opciones suficientes",
                "config": {"options": ["Unica"]},
            }
        )
    with pytest.raises(ValueError):
        normalize_response_payload("points_100", points.config, {"points": {"A": 50, "B": 40}})


def test_live_access_guards_allow_admin_and_consulta_only():
    app = build_app()
    client = app.test_client()

    response = login(client, "admin-live@test.local")
    assert response.status_code == 200
    assert "Live en Tiempo Real" in response.get_data(as_text=True)
    assert client.get("/live/").status_code == 200

    client.get("/auth/logout", follow_redirects=True)
    login(client, "consulta-live@test.local")
    assert client.get("/live/").status_code == 200
    assert client.post("/live/api/templates", json={}).status_code == 403

    client.get("/auth/logout", follow_redirects=True)
    login(client, "sin-live@test.local")
    denied = client.get("/live/")
    assert denied.status_code == 403


def test_admin_can_create_template_and_session_snapshot_via_api():
    app = build_app()
    client = app.test_client()
    login(client, "admin-live@test.local")

    template_response = client.post(
        "/live/api/templates",
        json={
            "tipo": "multiple_choice",
            "titulo": "Prioridad FODA",
            "prompt": "Selecciona la prioridad",
            "config": {"options": ["Alta", "Media", "Baja"]},
        },
    )
    assert template_response.status_code == 201
    template_id = template_response.get_json()["template"]["id"]

    session_response = client.post(
        "/live/api/sessions",
        json={"titulo": "Taller directivo", "mode": "guided", "template_ids": [template_id]},
    )
    assert session_response.status_code == 201
    payload = session_response.get_json()["session"]
    assert payload["activities"][0]["titulo"] == "Prioridad FODA"
    assert payload["activities"][0]["config"]["options"] == ["Alta", "Media", "Baja"]

    with app.app_context():
        activity = LiveActivity.query.first()
        assert activity.template_id == template_id
        assert activity.config_json["options"] == ["Alta", "Media", "Baja"]


def test_presentation_slides_support_content_navigation_reorder_and_duplicate():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "multiple_choice",
                "titulo": "Pulso",
                "prompt": "Elige",
                "config": {"options": ["A", "B"]},
            },
            admin,
        )
        db.session.flush()
        session = create_session({"titulo": "Presentacion", "mode": "guided", "template_ids": [template.id]}, admin)
        content = add_activity(
            session,
            {
                "tipo": "content_slide",
                "titulo": "Bienvenida",
                "prompt": "Escanea el codigo",
                "config": {"layout": "qr", "body": "Usa tu telefono.", "timer_seconds": 30},
            },
        )
        closing = add_activity(
            session,
            {
                "tipo": "content_slide",
                "titulo": "Cierre",
                "prompt": "Gracias",
                "config": {"layout": "title", "body": "Siguientes pasos"},
            },
        )
        db.session.flush()
        open_session(session)
        first_id = session.activities[0].id
        content_id = content.id
        closing_id = closing.id

        apply_presenter_control(session, {"session_id": session.id, "action": "next_slide"})
        assert session.active_activity_id == content_id
        apply_presenter_control(session, {"session_id": session.id, "action": "set_timer", "activity_id": content_id, "timer_seconds": 15})
        assert session.active_activity_id == content_id
        assert content.payload_json["timer_seconds"] == 15
        apply_presenter_control(session, {"session_id": session.id, "action": "previous_slide"})
        assert session.active_activity_id == first_id
        apply_presenter_control(session, {"session_id": session.id, "action": "go_to_slide", "activity_id": closing_id})
        assert session.active_activity_id == closing_id

        reorder_activities(session, [closing_id, first_id, content_id])
        assert [activity.id for activity in sorted(session.activities, key=lambda item: item.orden)] == [closing_id, first_id, content_id]
        duplicate = duplicate_activity(session, first_id)
        assert duplicate.responses == []
        assert duplicate.tipo == "multiple_choice"
        db.session.commit()


def test_admin_can_manage_slides_via_flask_api():
    app = build_app()
    client = app.test_client()
    login(client, "admin-live@test.local")

    session_response = client.post("/live/api/sessions", json={"titulo": "Deck API", "mode": "guided"})
    assert session_response.status_code == 201
    session_id = session_response.get_json()["session"]["id"]

    created = client.post(
        f"/live/api/sessions/{session_id}/activities",
        json={
            "tipo": "content_slide",
            "titulo": "Intro",
            "prompt": "Bienvenida",
            "config": {"layout": "text", "body": "Arranque"},
        },
    )
    assert created.status_code == 201
    slide_id = created.get_json()["activity"]["id"]

    updated = client.patch(
        f"/live/api/sessions/{session_id}/activities/{slide_id}",
        json={
            "titulo": "Intro editada",
            "prompt": "Nueva bienvenida",
            "config": {"layout": "instructions", "body": "Paso 1"},
        },
    )
    assert updated.status_code == 200
    assert updated.get_json()["activity"]["config"]["layout"] == "instructions"

    duplicated = client.post(f"/live/api/sessions/{session_id}/activities/{slide_id}/duplicate")
    assert duplicated.status_code == 201
    duplicate_id = duplicated.get_json()["activity"]["id"]

    reordered = client.post(f"/live/api/sessions/{session_id}/activities/reorder", json={"activity_ids": [duplicate_id, slide_id]})
    assert reordered.status_code == 200
    assert [item["id"] for item in reordered.get_json()["session"]["activities"]] == [duplicate_id, slide_id]

    deleted = client.delete(f"/live/api/sessions/{session_id}/activities/{slide_id}")
    assert deleted.status_code == 200
    assert [item["id"] for item in deleted.get_json()["session"]["activities"]] == [duplicate_id]


def test_public_participant_cookie_and_multiple_choice_replacement():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "multiple_choice",
                "titulo": "Votación",
                "prompt": "Elige",
                "config": {"options": ["A", "B"]},
            },
            admin,
        )
        db.session.flush()
        session = create_session({"titulo": "Sesión pública", "template_ids": [template.id]}, admin)
        open_session(session)
        db.session.commit()
        code = session.code
        activity_id = session.activities[0].id

    client = app.test_client()
    page = client.get(f"/live/s/{code}")
    assert page.status_code == 200
    assert "live_participant_token" in page.headers.get("Set-Cookie", "")

    first = client.post(f"/live/api/s/{code}/activities/{activity_id}/responses", json={"choice": "A"})
    second = client.post(f"/live/api/s/{code}/activities/{activity_id}/responses", json={"choice": "B"})
    assert first.status_code == 200
    assert second.status_code == 200

    with app.app_context():
        responses = LiveResponse.query.filter_by(activity_id=activity_id).all()
        assert len(responses) == 1
        assert responses[0].payload_json == {"choice": "B"}


def test_public_qr_route_opens_session_and_png_is_shareable():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "multiple_choice",
                "titulo": "Acceso QR",
                "prompt": "Elige",
                "config": {"options": ["A", "B"]},
            },
            admin,
        )
        db.session.flush()
        session = create_session({"titulo": "Sesion QR", "template_ids": [template.id]}, admin)
        open_session(session)
        db.session.commit()
        code = session.code
        session_id = session.id

    public_client = app.test_client()
    qr_png = public_client.get(f"/live/qr/{code}.png")
    assert qr_png.status_code == 200
    assert qr_png.mimetype == "image/png"
    assert qr_png.data.startswith(b"\x89PNG")

    scanned = public_client.get(f"/live/q/{code}", follow_redirects=True)
    assert scanned.status_code == 200
    assert f"Código {code}" in scanned.get_data(as_text=True) or f"Codigo {code}" in scanned.get_data(as_text=True)
    assert "live_participant_token" in scanned.headers.get("Set-Cookie", "")

    admin_client = app.test_client()
    login(admin_client, "admin-live@test.local")
    detail = admin_client.get(f"/live/sessions/{session_id}")
    html = detail.get_data(as_text=True)
    assert f"/live/qr/{code}.png" in html
    assert f"/live/q/{code}" in html


def test_brainstorm_allows_multiple_ideas_until_limit():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "brainstorm",
                "titulo": "Ideas",
                "prompt": "Comparte ideas",
                "config": {"max_ideas_per_participant": 2, "max_length": 80},
            },
            admin,
        )
        db.session.flush()
        session = create_session({"titulo": "Lluvia", "mode": "self_paced", "template_ids": [template.id]}, admin)
        open_session(session)
        db.session.commit()
        code = session.code
        activity_id = session.activities[0].id

    client = app.test_client()
    client.get(f"/live/s/{code}/activity/{activity_id}")
    assert client.post(f"/live/api/s/{code}/activities/{activity_id}/responses", json={"idea": "Primera idea"}).status_code == 200
    assert client.post(f"/live/api/s/{code}/activities/{activity_id}/responses", json={"idea": "Segunda idea"}).status_code == 200
    rejected = client.post(f"/live/api/s/{code}/activities/{activity_id}/responses", json={"idea": "Tercera idea"})
    assert rejected.status_code == 400

    with app.app_context():
        participant = LiveParticipant.query.first()
        assert participant is not None
        assert LiveResponse.query.filter_by(participant_id=participant.id).count() == 2


def test_strategic_activity_types_accept_and_aggregate_public_responses():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        templates = [
            create_template(
                {
                    "tipo": "scale",
                    "titulo": "Madurez administrativa",
                    "prompt": "Evalua cada capacidad",
                    "config": {"items": ["Procesos", "Datos"], "min": 1, "max": 5},
                },
                admin,
            ),
            create_template(
                {
                    "tipo": "ranking",
                    "titulo": "Prioridad FODA",
                    "prompt": "Ordena prioridades",
                    "config": {"items": ["Fortaleza", "Oportunidad", "Amenaza"], "max_ranked": 2},
                },
                admin,
            ),
            create_template(
                {
                    "tipo": "points_100",
                    "titulo": "Asignacion de acciones",
                    "prompt": "Distribuye recursos",
                    "config": {"items": ["Accion A", "Accion B"], "total_points": 100, "step": 10},
                },
                admin,
            ),
            create_template(
                {
                    "tipo": "matrix_2x2",
                    "titulo": "Partes interesadas",
                    "prompt": "Ubica cada actor",
                    "config": {
                        "items": ["Cabildo", "Ciudadania"],
                        "x_axis": {"min_label": "Bajo poder", "max_label": "Alto poder"},
                        "y_axis": {"min_label": "Bajo interes", "max_label": "Alto interes"},
                        "min": -5,
                        "max": 5,
                    },
                },
                admin,
            ),
        ]
        db.session.flush()
        session = create_session({"titulo": "Taller FODA", "mode": "self_paced", "template_ids": [item.id for item in templates]}, admin)
        open_session(session)
        db.session.commit()
        code = session.code
        activity_ids = {activity.tipo: activity.id for activity in session.activities}

    client = app.test_client()
    assert client.get(f"/live/s/{code}/activity/{activity_ids['scale']}").status_code == 200
    assert client.post(
        f"/live/api/s/{code}/activities/{activity_ids['scale']}/responses",
        json={"ratings": {"Procesos": 4, "Datos": 5}},
    ).status_code == 200
    assert client.post(
        f"/live/api/s/{code}/activities/{activity_ids['ranking']}/responses",
        json={"ranking": ["Oportunidad", "Fortaleza"]},
    ).status_code == 200
    assert client.post(
        f"/live/api/s/{code}/activities/{activity_ids['points_100']}/responses",
        json={"points": {"Accion A": 70, "Accion B": 30}},
    ).status_code == 200
    assert client.post(
        f"/live/api/s/{code}/activities/{activity_ids['matrix_2x2']}/responses",
        json={"ratings": {"Cabildo": {"x": 4, "y": 5}, "Ciudadania": {"x": -1, "y": 5}}},
    ).status_code == 200

    with app.app_context():
        scale = aggregate_results(db.session.get(LiveActivity, activity_ids["scale"]))
        ranking = aggregate_results(db.session.get(LiveActivity, activity_ids["ranking"]))
        points = aggregate_results(db.session.get(LiveActivity, activity_ids["points_100"]))
        matrix = aggregate_results(db.session.get(LiveActivity, activity_ids["matrix_2x2"]))
        assert scale["items"][0]["average"] == 4
        assert ranking["items"][0]["label"] == "Oportunidad"
        assert points["items"][0]["points"] == 70
        assert matrix["items"][0]["x"] == 4


def test_qa_supports_pending_moderation_upvotes_and_answered_status():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "qa",
                "titulo": "Preguntas ejecutivas",
                "prompt": "Pregunta lo necesario",
                "config": {"moderation": "manual", "allow_upvotes": True},
            },
            admin,
        )
        db.session.flush()
        session = create_session({"titulo": "Sesion Q&A", "mode": "self_paced", "template_ids": [template.id]}, admin)
        open_session(session)
        db.session.commit()
        code = session.code
        session_id = session.id
        activity_id = session.activities[0].id

    public_client = app.test_client()
    public_client.get(f"/live/s/{code}/activity/{activity_id}")
    submitted = public_client.post(
        f"/live/api/s/{code}/activities/{activity_id}/responses",
        json={"question": "Como se dara seguimiento?"},
    )
    assert submitted.status_code == 200
    response_id = submitted.get_json()["response_id"]

    admin_client = app.test_client()
    login(admin_client, "admin-live@test.local")
    approved = admin_client.post(
        f"/live/api/sessions/{session_id}/activities/{activity_id}/responses/{response_id}/moderate",
        json={"action": "approve"},
    )
    assert approved.status_code == 200

    upvote = public_client.post(f"/live/api/s/{code}/activities/{activity_id}/responses/{response_id}/upvote")
    assert upvote.status_code == 200

    answered = admin_client.post(
        f"/live/api/sessions/{session_id}/activities/{activity_id}/responses/{response_id}/moderate",
        json={"action": "answer"},
    )
    assert answered.status_code == 200

    with app.app_context():
        results = aggregate_results(db.session.get(LiveActivity, activity_id))
        assert results["answered"] == 1
        assert results["questions"][0]["upvotes"] == 1
        assert results["questions"][0]["status"] == "answered"


def test_guided_quiz_choice_scores_and_builds_leaderboard():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "quiz_choice",
                "titulo": "Quiz normativo",
                "prompt": "Selecciona la respuesta correcta",
                "config": {"options": ["Si", "No"], "correct_options": ["Si"], "timer_seconds": 20, "points": 150},
            },
            admin,
        )
        db.session.flush()
        session = create_session({"titulo": "Quiz guiado", "mode": "guided", "template_ids": [template.id]}, admin)
        open_session(session)
        db.session.commit()
        code = session.code
        activity_id = session.activities[0].id

    client = app.test_client()
    client.get(f"/live/s/{code}")
    response = client.post(f"/live/api/s/{code}/activities/{activity_id}/responses", json={"choice": "Si"})
    assert response.status_code == 200
    results = response.get_json()["results"]
    assert results["correct"] == 1
    assert results["leaderboard"][0]["score"] == 150


def test_live_session_exports_pdf_and_excel():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "multiple_choice",
                "titulo": "Pulso",
                "prompt": "Selecciona",
                "config": {"options": ["A", "B"]},
            },
            admin,
        )
        db.session.flush()
        session = create_session({"titulo": "Exportable", "mode": "guided", "template_ids": [template.id]}, admin)
        open_session(session)
        db.session.commit()
        session_id = session.id

    client = app.test_client()
    login(client, "admin-live@test.local")
    pdf = client.get(f"/live/sessions/{session_id}/export.pdf")
    assert pdf.status_code == 200
    assert pdf.mimetype == "application/pdf"
    assert pdf.data.startswith(b"%PDF")

    xlsx = client.get(f"/live/sessions/{session_id}/export.xlsx")
    assert xlsx.status_code == 200
    assert xlsx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(xlsx.data))
    assert "Actividades" in workbook.sheetnames


def test_socketio_join_submit_and_presenter_control_emit_updates():
    app = build_app()
    with app.app_context():
        admin = Usuario.query.filter_by(correo="admin-live@test.local").first()
        template = create_template(
            {
                "tipo": "multiple_choice",
                "titulo": "Pulso",
                "prompt": "Selecciona",
                "config": {"options": ["Si", "No"]},
            },
            admin,
        )
        db.session.flush()
        session = LiveSession(titulo="Socket", code="ABC123", mode="guided", estado="draft", presentador=admin, config_json={})
        db.session.add(session)
        db.session.flush()
        activity = add_activity_from_template(session, template)
        content = add_activity(
            session,
            {
                "tipo": "content_slide",
                "titulo": "Instrucciones",
                "prompt": "Siguiente paso",
                "config": {"layout": "text", "body": "Continua con la dinamica."},
            },
        )
        db.session.commit()
        session_id = session.id
        activity_id = activity.id
        content_id = content.id

    flask_client = app.test_client()
    login(flask_client, "admin-live@test.local")
    admin_socket = socketio.test_client(app, flask_test_client=flask_client)
    join_ack = admin_socket.emit("live:join_session", {"session_id": session_id}, callback=True)
    assert join_ack["ok"] is True

    control_ack = admin_socket.emit(
        "live:presenter_control",
        {"session_id": session_id, "action": "open_activity", "activity_id": activity_id},
        callback=True,
    )
    assert control_ack["ok"] is True
    assert control_ack["session"]["active_activity_id"] == activity_id

    participant_socket = socketio.test_client(app)
    participant_socket.emit("live:join_session", {"session_id": session_id, "participant_token": "anon-token"}, callback=True)
    submit_ack = participant_socket.emit(
        "live:submit_response",
        {
            "session_id": session_id,
            "activity_id": activity_id,
            "participant_token": "anon-token",
            "payload": {"choice": "Si"},
        },
        callback=True,
    )
    assert submit_ack["ok"] is True
    received = admin_socket.get_received()
    assert any(item["name"] == "live:results_updated" for item in received)

    slide_ack = admin_socket.emit(
        "live:presenter_control",
        {"session_id": session_id, "action": "go_to_slide", "activity_id": content_id},
        callback=True,
    )
    assert slide_ack["ok"] is True
    assert slide_ack["session"]["active_activity_id"] == content_id
    slide_events = admin_socket.get_received()
    assert any(item["name"] == "live:session_state" for item in slide_events)
    assert any(item["name"] == "live:activity_state" for item in slide_events)

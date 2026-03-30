import zipfile
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from municipal_diagnostico import create_app
from municipal_diagnostico.extensions import db
from municipal_diagnostico.models import Area, ComentarioEje, EvidenciaEje, Evaluacion, PeriodoEvaluacion, Respuesta, Usuario, Dependencia
from municipal_diagnostico.seeds import ensure_official_questionnaire


class TestConfig:
    SECRET_KEY = "test"
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = "tests/uploads"
    ALLOWED_EXTENSIONS = {"pdf"}
    AUTO_INIT_DATABASE = True
    BOOTSTRAP_ADMIN_EMAIL = None
    BOOTSTRAP_ADMIN_PASSWORD = None
    BOOTSTRAP_ADMIN_NAME = None


def add_axis_responses(evaluacion, usuario, area, reactivos, valor, comentario):
    for reactivo in reactivos:
        db.session.add(
            Respuesta(
                evaluacion=evaluacion,
                reactivo_version=reactivo,
                area=area,
                usuario_captura=usuario,
                valor=valor,
                comentario=comentario,
            )
        )


def build_app_with_reporting_data():
    app = create_app(TestConfig)
    with app.app_context():
        questionnaire = ensure_official_questionnaire()

        dependencia_preliminar = Dependencia(nombre="Oficialía Mayor", tipo="Administrativa")
        dependencia_oficial = Dependencia(nombre="Tesorería Municipal", tipo="Administrativa")
        area_preliminar = Area(nombre="Recursos Humanos", dependencia=dependencia_preliminar)
        area_oficial = Area(nombre="Ingresos", dependencia=dependencia_oficial)

        administrador = Usuario(nombre="Administrador", correo="admin@test.local", rol="administrador", activo=True)
        administrador.set_password("secret123")
        consulta = Usuario(nombre="Consulta", correo="consulta@test.local", rol="consulta", activo=True)
        consulta.set_password("secret123")
        capturista_preliminar = Usuario(
            nombre="Capturista OM",
            correo="captura.om@test.local",
            rol="evaluador",
            activo=True,
            dependencia=dependencia_preliminar,
            area=area_preliminar,
        )
        capturista_preliminar.set_password("secret123")
        capturista_oficial = Usuario(
            nombre="Capturista Tesorería",
            correo="captura.teso@test.local",
            rol="evaluador",
            activo=True,
            dependencia=dependencia_oficial,
            area=area_oficial,
        )
        capturista_oficial.set_password("secret123")

        periodo_actual = PeriodoEvaluacion(
            nombre="Periodo Institucional 2026",
            estado="abierto",
            fecha_inicio=date(2026, 1, 1),
            fecha_cierre=date(2026, 12, 31),
            cuestionario_version=questionnaire,
        )
        periodo_anterior = PeriodoEvaluacion(
            nombre="Periodo Institucional 2025",
            estado="cerrado",
            fecha_inicio=date(2025, 1, 1),
            fecha_cierre=date(2025, 12, 31),
            cuestionario_version=questionnaire,
        )

        evaluacion_preliminar = Evaluacion(
            periodo=periodo_actual,
            dependencia=dependencia_preliminar,
            estado="en_captura",
        )
        evaluacion_oficial = Evaluacion(
            periodo=periodo_actual,
            dependencia=dependencia_oficial,
            estado="aprobada",
        )
        evaluacion_historica = Evaluacion(
            periodo=periodo_anterior,
            dependencia=dependencia_oficial,
            estado="cerrada",
        )

        db.session.add_all(
            [
                dependencia_preliminar,
                dependencia_oficial,
                area_preliminar,
                area_oficial,
                administrador,
                consulta,
                capturista_preliminar,
                capturista_oficial,
                periodo_actual,
                periodo_anterior,
                evaluacion_preliminar,
                evaluacion_oficial,
                evaluacion_historica,
            ]
        )
        db.session.flush()

        primer_eje = questionnaire.ejes[0]
        segundo_eje = questionnaire.ejes[1]

        add_axis_responses(
            evaluacion_preliminar,
            capturista_preliminar,
            area_preliminar,
            primer_eje.reactivos[:3],
            2,
            "Avance documentado",
        )
        add_axis_responses(
            evaluacion_preliminar,
            capturista_preliminar,
            area_preliminar,
            segundo_eje.reactivos[:2],
            1,
            "Capacidad incipiente",
        )
        add_axis_responses(
            evaluacion_oficial,
            capturista_oficial,
            area_oficial,
            primer_eje.reactivos[:3],
            3,
            "Práctica consolidada",
        )
        add_axis_responses(
            evaluacion_oficial,
            capturista_oficial,
            area_oficial,
            segundo_eje.reactivos[:4],
            2,
            "Proceso institucionalizado",
        )
        add_axis_responses(
            evaluacion_historica,
            capturista_oficial,
            area_oficial,
            primer_eje.reactivos[:2],
            2,
            "Referencia histórica",
        )

        db.session.add(
            EvidenciaEje(
                evaluacion=evaluacion_preliminar,
                eje_version=primer_eje,
                area=area_preliminar,
                usuario=capturista_preliminar,
                version=1,
                archivo_nombre_original="manual-operativo.pdf",
                archivo_guardado="periodos/1/evaluaciones/1/manual-operativo.pdf",
                mime_type="application/pdf",
                tamano_bytes=1024,
                activo=True,
            )
        )

        db.session.commit()
        return app, {
            "current_period_id": periodo_actual.id,
            "previous_period_id": periodo_anterior.id,
            "preliminary_evaluation_id": evaluacion_preliminar.id,
            "official_evaluation_id": evaluacion_oficial.id,
            "historical_evaluation_id": evaluacion_historica.id,
            "preliminary_dependency_id": dependencia_preliminar.id,
            "official_dependency_id": dependencia_oficial.id,
            "axis_id": primer_eje.id,
            "axis_name": primer_eje.nombre,
            "first_reactive_id": primer_eje.reactivos[0].id,
            "preliminary_area_id": area_preliminar.id,
            "questionnaire_version_id": questionnaire.id,
        }


def login(client, email: str, password: str = "secret123"):
    return client.post(
        "/auth/login",
        data={"correo": email, "password": password},
        follow_redirects=True,
    )


def test_admin_report_hub_and_preliminary_report_render_without_mojibake():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")

    response = client.get("/reportes/")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Hub operativo de reportes" in html
    assert "Oficialía Mayor" in html
    assert "Tesorería Municipal" in html
    assert "Centro Ejecutivo" in html
    assert "Ã" not in html

    report = client.get(f"/reportes/evaluaciones/{ids['preliminary_evaluation_id']}")
    report_html = report.get_data(as_text=True)

    assert report.status_code == 200
    assert "Resultado preliminar" in report_html
    assert "Cuestionario respondido" in report_html
    assert "Índice de madurez" in report_html
    assert "Oficialía Mayor" in report_html
    assert "Ã" not in report_html

    pdf = client.get(f"/reportes/evaluaciones/{ids['preliminary_evaluation_id']}/pdf")
    assert pdf.status_code == 200
    assert pdf.mimetype == "application/pdf"

    xlsx = client.get(f"/reportes/evaluaciones/{ids['preliminary_evaluation_id']}/xlsx")
    assert xlsx.status_code == 200
    assert xlsx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(xlsx.data))
    assert "Cuestionario" in workbook.sheetnames
    questionnaire_values = [
        value
        for row in workbook["Cuestionario"].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    ]
    assert any("Avance documentado" in value for value in questionnaire_values)

    word = client.get(f"/reportes/evaluaciones/{ids['preliminary_evaluation_id']}/word")
    assert word.status_code == 200
    assert word.mimetype == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    with zipfile.ZipFile(BytesIO(word.data)) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8")
    assert "Reporte ejecutivo de evaluacion" in document_xml
    assert "Avance documentado" in document_xml


def test_admin_can_open_executive_home_and_see_dependency_and_axis_cards():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get(f"/reportes/ejecutivo?periodo_id={ids['current_period_id']}")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Centro ejecutivo" in html
    assert "Oficialía Mayor" in html
    assert "Tesorería Municipal" in html
    assert ids["axis_name"] in html
    assert "Lectura preliminar" in html
    assert "Ã" not in html


def test_consulta_executive_home_only_sees_official_results():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "consulta@test.local")
    response = client.get(f"/reportes/ejecutivo?periodo_id={ids['current_period_id']}")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Tesorería Municipal" in html
    assert "Oficialía Mayor" not in html


def test_executive_dependency_view_renders_institutional_panels_and_history():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get(
        f"/reportes/ejecutivo/dependencias/{ids['official_dependency_id']}?periodo_id={ids['current_period_id']}"
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Ficha ejecutiva por dependencia" in html
    assert "Matriz de prioridades" in html
    assert "Mapa de brechas" in html
    assert "Evolución de la dependencia" in html
    assert "Periodo Institucional 2025" in html


def test_executive_axis_view_compares_dependencies_and_consulta_sees_only_official():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    admin_response = client.get(
        f"/reportes/ejecutivo/ejes/{ids['axis_id']}?periodo_id={ids['current_period_id']}"
    )
    admin_html = admin_response.get_data(as_text=True)

    assert admin_response.status_code == 200
    assert "Vista ejecutiva por eje" in admin_html
    assert "Dependencias comparadas en el eje" in admin_html
    assert "Oficialía Mayor" in admin_html
    assert "Tesorería Municipal" in admin_html

    client.get("/auth/logout", follow_redirects=True)
    login(client, "consulta@test.local")
    consulta_response = client.get(
        f"/reportes/ejecutivo/ejes/{ids['axis_id']}?periodo_id={ids['current_period_id']}"
    )
    consulta_html = consulta_response.get_data(as_text=True)

    assert consulta_response.status_code == 200
    assert "Tesorería Municipal" in consulta_html
    assert "Oficialía Mayor" not in consulta_html


def test_invalid_executive_routes_and_permissions_are_enforced():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "consulta@test.local")
    assert client.get(
        f"/reportes/ejecutivo/dependencias/{ids['preliminary_dependency_id']}?periodo_id={ids['current_period_id']}"
    ).status_code == 404
    assert client.get("/reportes/").status_code == 403

    client.get("/auth/logout", follow_redirects=True)
    login(client, "admin@test.local")
    assert client.get(
        f"/reportes/ejecutivo/ejes/{ids['axis_id'] + 999}?periodo_id={ids['current_period_id']}"
    ).status_code == 404
    assert client.get(
        f"/reportes/ejecutivo/dependencias/{ids['official_dependency_id'] + 999}?periodo_id={ids['current_period_id']}"
    ).status_code == 404


def test_period_report_separates_operational_and_official_layers_by_role():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    admin_report = client.get(f"/reportes/periodos/{ids['current_period_id']}")
    admin_html = admin_report.get_data(as_text=True)

    assert admin_report.status_code == 200
    assert "Avance operativo" in admin_html
    assert "Resultado oficial" in admin_html
    assert "Oficialía Mayor" in admin_html
    assert "Tesorería Municipal" in admin_html

    client.get("/auth/logout", follow_redirects=True)
    login(client, "consulta@test.local")
    consulta_report = client.get(f"/reportes/periodos/{ids['current_period_id']}")
    consulta_html = consulta_report.get_data(as_text=True)

    assert consulta_report.status_code == 200
    assert "Tesorería Municipal" in consulta_html
    assert "Oficialía Mayor" not in consulta_html


def test_admin_preview_shows_read_only_questionnaire_implementation():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get(f"/admin/evaluaciones/{ids['preliminary_evaluation_id']}/preview")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Vista administrativa de solo lectura" in html
    assert "Recursos Humanos" in html
    assert "Avance documentado" in html
    assert "manual-operativo.pdf" in html
    assert "Sin respuesta" in html
    assert "Ã" not in html


def test_report_hub_and_preview_are_admin_only():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "consulta@test.local")

    assert client.get("/reportes/").status_code == 403
    assert client.get(f"/admin/evaluaciones/{ids['preliminary_evaluation_id']}/preview").status_code == 403


def test_autosave_persists_axis_progress_and_module_comment():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")

    payload = {
        "eje_id": ids["axis_id"],
        "comentario_eje": "Cierre del módulo con evidencia disponible",
        "comentario_eje_area_id": ids["preliminary_area_id"],
        "responses": [
            {
                "reactivo_id": ids["first_reactive_id"],
                "valor": 3,
                "area_id": ids["preliminary_area_id"],
                "comentario": "Fortalecido en autosave",
            }
        ],
    }
    response = client.post(
        f"/evaluaciones/{ids['preliminary_evaluation_id']}/autosave",
        json=payload,
    )

    assert response.status_code == 200
    data = response.get_json()
    assert data["ok"] is True
    assert data["last_saved"]

    with app.app_context():
        comment = ComentarioEje.query.filter_by(
            evaluacion_id=ids["preliminary_evaluation_id"],
            eje_version_id=ids["axis_id"],
        ).first()
        assert comment is not None
        assert comment.comentario == "Cierre del módulo con evidencia disponible"

        updated_response = Respuesta.query.filter_by(
            evaluacion_id=ids["preliminary_evaluation_id"],
            reactivo_version_id=ids["first_reactive_id"],
        ).first()
        assert updated_response is not None
        assert updated_response.valor == 3
        assert updated_response.comentario == "Fortalecido en autosave"


def test_admin_can_open_read_only_evaluator_view_and_monitoring():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    view_response = client.get(f"/admin/evaluaciones/{ids['preliminary_evaluation_id']}/vista-capturista")
    listing_response = client.get("/admin/vista-capturista")
    monitoring_response = client.get("/admin/monitoreo")

    assert view_response.status_code == 200
    assert "modo solo lectura" in view_response.get_data(as_text=True)
    assert listing_response.status_code == 200
    listing_html = listing_response.get_data(as_text=True)
    assert "Vista capturista para administración" in listing_html
    assert "Oficialía Mayor" in listing_html
    assert "Tesorería Municipal" in listing_html
    assert f"/admin/evaluaciones/{ids['historical_evaluation_id']}/vista-capturista" not in listing_html

    assert monitoring_response.status_code == 200
    monitoring_html = monitoring_response.get_data(as_text=True)
    assert "Monitoreo institucional de uso" in monitoring_html
    assert "Acciones por usuario" in monitoring_html


def test_evaluator_views_listing_filters_by_period_and_state():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get(
        f"/admin/vista-capturista?periodo_id={ids['current_period_id']}&estado=en_captura"
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert f"/admin/evaluaciones/{ids['preliminary_evaluation_id']}/vista-capturista" in html
    assert f"/admin/evaluaciones/{ids['official_evaluation_id']}/vista-capturista" not in html


def test_monitoring_is_admin_only():
    app, _ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "consulta@test.local")
    assert client.get("/admin/monitoreo").status_code == 403


def test_admin_catalogs_render_management_screen_and_support_safe_actions():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")

    response = client.get("/admin/catalogos")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Alta de usuario" in html
    assert "Funciones por rol" in html
    assert "Unidad administrativa" in html

    create_response = client.post(
        "/admin/catalogos",
        data={
            "action": "add_dependencia",
            "nombre": "Servicios Públicos",
            "tipo": "Operativa",
            "descripcion": "Nueva dependencia para pruebas",
            "redirect_anchor": "catalogo-dependencias",
        },
        follow_redirects=True,
    )
    assert create_response.status_code == 200
    assert "Dependencia registrada." in create_response.get_data(as_text=True)

    deactivate_area_response = client.post(
        "/admin/catalogos",
        data={
            "action": "toggle_area",
            "area_id": ids["preliminary_area_id"],
            "redirect_anchor": "catalogo-areas",
        },
        follow_redirects=True,
    )
    assert deactivate_area_response.status_code == 200
    assert "Unidad administrativa desactivada." in deactivate_area_response.get_data(as_text=True)

    blocked_delete_response = client.post(
        "/admin/catalogos",
        data={
            "action": "delete_dependencia",
            "dependencia_id": ids["preliminary_dependency_id"],
            "redirect_anchor": "catalogo-dependencias",
        },
        follow_redirects=True,
    )
    assert blocked_delete_response.status_code == 200
    assert "No se puede eliminar la dependencia" in blocked_delete_response.get_data(as_text=True)

    with app.app_context():
        created = Dependencia.query.filter_by(nombre="Servicios Públicos").first()
        assert created is not None
        area = db.session.get(Area, ids["preliminary_area_id"])
        assert area is not None
        assert area.activa is False


def test_admin_can_preview_fill_screen_per_questionnaire_version():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get(f"/admin/cuestionarios/{ids['questionnaire_version_id']}/vista-llenado")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "vista metodológica" in html.lower()
    assert "Comentario condicionado por responsable" in html
    assert "Adjuntar archivos" not in html
    assert "Comentario del reactivo" not in html
    assert "bloque para adjuntar evidencia documental" in html


def test_capture_hides_module_and_point_detail_without_responsable():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get(f"/evaluaciones/{ids['preliminary_evaluation_id']}")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Responsable pendiente" in html
    assert "Responsable requerido para cerrar el módulo" in html
    assert f'data-dependent-block="modulo-{ids["axis_id"]}"' in html


def test_invalid_evaluator_view_redirects_to_admin_listing():
    app, _ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get("/admin/evaluaciones/99999/vista-capturista", follow_redirects=True)
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "La evaluación solicitada no existe todavía" in html
    assert "Vista capturista para administración" in html


def test_admin_dashboard_exposes_report_download_shortcuts():
    app, ids = build_app_with_reporting_data()
    client = app.test_client()

    login(client, "admin@test.local")
    response = client.get("/dashboard/diagnostico")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Reportes listos para descarga" in html
    assert f"/reportes/evaluaciones/{ids['preliminary_evaluation_id']}/pdf" in html
    assert f"/reportes/evaluaciones/{ids['preliminary_evaluation_id']}/xlsx" in html
    assert f"/reportes/evaluaciones/{ids['preliminary_evaluation_id']}/word" in html

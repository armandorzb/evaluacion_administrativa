from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from werkzeug.datastructures import MultiDict

from municipal_diagnostico import create_app
from municipal_diagnostico.extensions import db
from municipal_diagnostico.models import (
    Dependencia,
    Iso9001Asignacion,
    Iso9001Apartado,
    Iso9001Ciclo,
    Iso9001Clausula,
    Iso9001CuestionarioVersion,
    Iso9001Evidencia,
    Iso9001Evaluacion,
    Iso9001Reactivo,
    Iso9001Respuesta,
    Usuario,
)
from municipal_diagnostico.iso9001_seed_data import ISO9001_VERSION
from municipal_diagnostico.services.iso9001 import ISO9001_OPTION_POINTS, ensure_iso9001_catalog, summarize_iso9001_evaluation
from municipal_diagnostico.services.iso9001_exports import (
    _priority_sections,
    _response_distribution,
    _section_evidence_count,
    build_iso9001_pdf,
)


class TestConfig:
    SECRET_KEY = "test"
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = "tests/uploads"
    ALLOWED_EXTENSIONS = {"pdf"}
    AUTO_INIT_DATABASE = True
    BOOTSTRAP_ADMIN_EMAIL = None
    BOOTSTRAP_ADMIN_PASSWORD = None
    BOOTSTRAP_ADMIN_NAME = None


def login(client, email: str, password: str = "secret123"):
    return client.post(
        "/auth/login",
        data={"correo": email, "password": password},
        follow_redirects=True,
    )


def build_app():
    app = create_app(TestConfig)
    with app.app_context():
        dependency_one = Dependencia(nombre="Contraloria", tipo="Administrativa")
        dependency_two = Dependencia(nombre="Secretaria Tecnica", tipo="Administrativa")
        users = [
            Usuario(
                nombre="Admin ISO",
                correo="admin.iso@test.local",
                rol="administrador",
                activo=True,
                acceso_diagnostico=False,
                acceso_bienestar=False,
                acceso_iso9001=True,
            ),
            Usuario(
                nombre="Capturista ISO",
                correo="captura.iso@test.local",
                rol="evaluador",
                activo=True,
                acceso_diagnostico=False,
                acceso_bienestar=False,
                acceso_iso9001=True,
            ),
            Usuario(
                nombre="Revisor ISO",
                correo="revisor.iso@test.local",
                rol="revisor",
                activo=True,
                acceso_diagnostico=False,
                acceso_bienestar=False,
                acceso_iso9001=True,
            ),
            Usuario(
                nombre="Revisor Alterno ISO",
                correo="revisor.alterno.iso@test.local",
                rol="revisor",
                activo=True,
                acceso_diagnostico=False,
                acceso_bienestar=False,
                acceso_iso9001=False,
            ),
            Usuario(
                nombre="Consulta ISO",
                correo="consulta.iso@test.local",
                rol="consulta",
                activo=True,
                acceso_diagnostico=False,
                acceso_bienestar=False,
                acceso_iso9001=True,
            ),
            Usuario(
                nombre="Operativo Disponible",
                correo="operativo.iso@test.local",
                rol="respondente",
                activo=True,
                acceso_diagnostico=True,
                acceso_bienestar=False,
                acceso_iso9001=False,
            ),
            Usuario(
                nombre="Sin ISO",
                correo="sin.iso@test.local",
                rol="consulta",
                activo=True,
                acceso_diagnostico=True,
                acceso_bienestar=False,
                acceso_iso9001=False,
            ),
        ]
        for user in users:
            user.set_password("secret123")
        db.session.add_all([dependency_one, dependency_two, *users])
        db.session.commit()
        return app, {
            "dependency_one_id": dependency_one.id,
            "dependency_two_id": dependency_two.id,
            "admin_id": users[0].id,
            "evaluator_id": users[1].id,
            "reviewer_id": users[2].id,
            "alternate_reviewer_id": users[3].id,
            "respondent_id": users[5].id,
        }


def create_iso_evaluation(dependency_id: int, evaluator_id: int, reviewer_id: int, admin_id: int) -> int:
    version = Iso9001CuestionarioVersion.query.first()
    cycle = Iso9001Ciclo(
        nombre=f"Ciclo ISO {dependency_id}",
        descripcion="Ciclo de prueba",
        estado="activo",
        fecha_inicio=date(2026, 1, 1),
        fecha_cierre=date(2026, 12, 31),
        version=version,
        creado_por_id=admin_id,
    )
    evaluation = Iso9001Evaluacion(
        ciclo=cycle,
        dependencia_id=dependency_id,
        revisor_id=reviewer_id,
        estado="borrador",
    )
    db.session.add_all([cycle, evaluation])
    db.session.flush()
    db.session.add(Iso9001Asignacion(evaluacion=evaluation, usuario_id=evaluator_id, tipo="captura"))
    db.session.commit()
    return evaluation.id


def fill_missing_responses(evaluation_id: int, user_id: int, value: str = "si") -> None:
    evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
    answered = {response.reactivo_id for response in evaluation.respuestas}
    for reactive in Iso9001Reactivo.query.order_by(Iso9001Reactivo.id).all():
        if reactive.id in answered:
            continue
        db.session.add(
            Iso9001Respuesta(
                evaluacion=evaluation,
                reactivo=reactive,
                usuario_id=user_id,
                calificacion=value,
                valor=ISO9001_OPTION_POINTS[value],
            )
        )
    summarize_iso9001_evaluation(evaluation)
    db.session.commit()


def test_iso9001_seed_catalog_matches_source_workbook_counts():
    app, _ids = build_app()

    with app.app_context():
        version = Iso9001CuestionarioVersion.query.first()
        assert version is not None
        assert Iso9001Clausula.query.count() == 7
        assert Iso9001Apartado.query.count() == 56
        assert Iso9001Reactivo.query.count() == 293

        counts = {
            int(clause.numero): sum(len(section.reactivos) for section in clause.apartados)
            for clause in version.clausulas
        }
        assert counts == {4: 23, 5: 27, 6: 17, 7: 62, 8: 112, 9: 36, 10: 16}


def test_iso9001_catalog_backfill_updates_copy_without_touching_responses():
    app, ids = build_app()

    with app.app_context():
        version = Iso9001CuestionarioVersion.query.first()
        first_clause = version.clausulas[0]
        first_section = first_clause.apartados[0]
        first_reactive = first_section.reactivos[0]
        version_id = version.id
        clause_id = first_clause.id
        section_id = first_section.id
        reactive_id = first_reactive.id

        evaluation_id = create_iso_evaluation(
            ids["dependency_one_id"],
            ids["evaluator_id"],
            ids["reviewer_id"],
            ids["admin_id"],
        )
        evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
        response = Iso9001Respuesta(
            evaluacion=evaluation,
            reactivo_id=reactive_id,
            usuario_id=ids["evaluator_id"],
            calificacion="si",
            valor=ISO9001_OPTION_POINTS["si"],
            observacion="Respuesta que debe conservarse.",
        )
        db.session.add(response)
        db.session.flush()
        response_id = response.id

        version.nombre = "Diagnostico de implementacion ISO 9001:2015"
        version.descripcion = "Texto anterior del catalogo."
        first_clause.nombre = "Contexto De La Organización"
        first_section.nombre = "Comprension anterior"
        first_reactive.texto = "Texto anterior del reactivo."
        first_reactive.evidencia_sugerida = "Evidencia anterior."
        first_reactive.criterio_idoneidad = "Criterio anterior."
        db.session.commit()

        ensure_iso9001_catalog()

        expected_clause = ISO9001_VERSION["clausulas"][0]
        expected_section = expected_clause["apartados"][0]
        expected_reactive = expected_section["reactivos"][0]
        restored_version = db.session.get(Iso9001CuestionarioVersion, version_id)
        restored_clause = db.session.get(Iso9001Clausula, clause_id)
        restored_section = db.session.get(Iso9001Apartado, section_id)
        restored_reactive = db.session.get(Iso9001Reactivo, reactive_id)
        restored_response = db.session.get(Iso9001Respuesta, response_id)

        assert restored_version.nombre == ISO9001_VERSION["nombre"]
        assert restored_version.descripcion == ISO9001_VERSION["descripcion"]
        assert restored_clause.nombre == expected_clause["nombre"]
        assert restored_section.nombre == expected_section["nombre"]
        assert restored_reactive.texto == expected_reactive["texto"]
        assert restored_reactive.evidencia_sugerida == expected_reactive["evidencia_sugerida"]
        assert restored_reactive.criterio_idoneidad == expected_reactive["criterio_idoneidad"]
        assert restored_response.calificacion == "si"
        assert restored_response.observacion == "Respuesta que debe conservarse."
        assert Iso9001Clausula.query.count() == 7
        assert Iso9001Apartado.query.count() == 56
        assert Iso9001Reactivo.query.count() == 293


def test_iso9001_scoring_excludes_na_from_denominator():
    app, ids = build_app()

    with app.app_context():
        evaluation_id = create_iso_evaluation(
            ids["dependency_one_id"],
            ids["evaluator_id"],
            ids["reviewer_id"],
            ids["admin_id"],
        )
        evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
        user = db.session.get(Usuario, ids["evaluator_id"])
        reactives = Iso9001Reactivo.query.order_by(Iso9001Reactivo.id).limit(3).all()
        for reactive, value in zip(reactives, ["si", "parcial", "na"]):
            db.session.add(
                Iso9001Respuesta(
                    evaluacion=evaluation,
                    reactivo=reactive,
                    usuario=user,
                    calificacion=value,
                    valor=ISO9001_OPTION_POINTS[value],
                )
            )
        db.session.flush()

        summary = summarize_iso9001_evaluation(evaluation)
        assert summary["answered_questions"] == 3
        assert summary["applicable_questions"] == 2
        assert summary["na_questions"] == 1
        assert summary["points"] == 3
        assert summary["percent"] == 75.0
        assert summary["completion"] == round((3 / 293) * 100, 2)


def test_iso9001_pdf_data_helpers_rank_gaps_and_evidence():
    app, ids = build_app()

    with app.app_context():
        evaluation_id = create_iso_evaluation(
            ids["dependency_one_id"],
            ids["evaluator_id"],
            ids["reviewer_id"],
            ids["admin_id"],
        )
        evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
        user = db.session.get(Usuario, ids["evaluator_id"])
        sections = Iso9001Apartado.query.order_by(Iso9001Apartado.orden).limit(2).all()
        first_section_reactives = sections[0].reactivos[:2]
        second_section_reactives = sections[1].reactivos[:2]
        payload = [
            (first_section_reactives[0], "si"),
            (first_section_reactives[1], "parcial"),
            (second_section_reactives[0], "no"),
            (second_section_reactives[1], "na"),
        ]
        responses = []
        for reactive, value in payload:
            response = Iso9001Respuesta(
                evaluacion=evaluation,
                reactivo=reactive,
                usuario=user,
                calificacion=value,
                valor=ISO9001_OPTION_POINTS[value],
                observacion=f"Observación {value}",
            )
            responses.append(response)
            db.session.add(response)
        db.session.flush()
        db.session.add_all(
            [
                Iso9001Evidencia(
                    respuesta=responses[0],
                    usuario=user,
                    archivo_nombre_original="evidencia-a.pdf",
                    archivo_guardado="iso/evidencia-a.pdf",
                    mime_type="application/pdf",
                    tamano_bytes=20,
                ),
                Iso9001Evidencia(
                    respuesta=responses[0],
                    usuario=user,
                    archivo_nombre_original="evidencia-b.pdf",
                    archivo_guardado="iso/evidencia-b.pdf",
                    mime_type="application/pdf",
                    tamano_bytes=22,
                ),
            ]
        )
        db.session.flush()

        summary = summarize_iso9001_evaluation(evaluation)
        distribution = {row["key"]: row["count"] for row in _response_distribution(summary)}
        priorities = _priority_sections(summary)
        section_map = {section["codigo"]: section for section in summary["sections"]}

        assert distribution == {"no": 1, "parcial": 1, "si": 1, "na": 1}
        assert priorities[0]["codigo"] == sections[1].codigo
        assert priorities[1]["codigo"] == sections[0].codigo
        assert _section_evidence_count(section_map[sections[0].codigo]) == 2
        assert _section_evidence_count(section_map[sections[1].codigo]) == 0


def test_iso9001_pdf_generates_without_completed_capture_and_with_all_na_section():
    app, ids = build_app()

    with app.app_context():
        empty_evaluation_id = create_iso_evaluation(
            ids["dependency_one_id"],
            ids["evaluator_id"],
            ids["reviewer_id"],
            ids["admin_id"],
        )
        empty_evaluation = db.session.get(Iso9001Evaluacion, empty_evaluation_id)
        empty_pdf = build_iso9001_pdf(empty_evaluation)
        assert empty_pdf.getvalue().startswith(b"%PDF")

        na_evaluation_id = create_iso_evaluation(
            ids["dependency_two_id"],
            ids["evaluator_id"],
            ids["reviewer_id"],
            ids["admin_id"],
        )
        na_evaluation = db.session.get(Iso9001Evaluacion, na_evaluation_id)
        user = db.session.get(Usuario, ids["evaluator_id"])
        first_section = Iso9001Apartado.query.order_by(Iso9001Apartado.orden).first()
        for reactive in first_section.reactivos:
            db.session.add(
                Iso9001Respuesta(
                    evaluacion=na_evaluation,
                    reactivo=reactive,
                    usuario=user,
                    calificacion="na",
                    valor=ISO9001_OPTION_POINTS["na"],
                    observacion="No aplica por alcance documentado.",
                )
            )
        db.session.flush()

        na_pdf = build_iso9001_pdf(na_evaluation)
        assert na_pdf.getvalue().startswith(b"%PDF")


def test_iso9001_capture_can_be_assigned_to_any_active_user():
    app, ids = build_app()
    client = app.test_client()

    login(client, "admin.iso@test.local")
    cycle_response = client.post(
        "/iso9001/ciclos",
        data={
            "action": "create_cycle",
            "nombre": "Ciclo ISO Captura Libre",
            "descripcion": "Asignacion a usuario operativo",
            "estado": "activo",
            "fecha_inicio": "2026-01-01",
            "fecha_cierre": "2026-12-31",
        },
        follow_redirects=True,
    )
    assert cycle_response.status_code == 200

    with app.app_context():
        cycle = Iso9001Ciclo.query.filter_by(nombre="Ciclo ISO Captura Libre").first()
        cycle_id = cycle.id

    cycles_page = client.get(f"/iso9001/ciclos?cycle_id={cycle_id}")
    cycles_html = cycles_page.get_data(as_text=True)
    assert cycles_page.status_code == 200
    assert "Operativo Disponible" in cycles_html

    assignment_response = client.post(
        "/iso9001/ciclos",
        data={
            "action": "add_evaluations",
            "cycle_id": str(cycle_id),
            "dependencia_ids": [str(ids["dependency_one_id"])],
            "responsable_id": str(ids["respondent_id"]),
            "revisor_id": str(ids["reviewer_id"]),
        },
        follow_redirects=True,
    )
    assert assignment_response.status_code == 200
    assert "Evaluaciones registradas: 1." in assignment_response.get_data(as_text=True)

    with app.app_context():
        operative = db.session.get(Usuario, ids["respondent_id"])
        assert operative.acceso_iso9001 is True
        evaluation = Iso9001Evaluacion.query.filter_by(ciclo_id=cycle_id, dependencia_id=ids["dependency_one_id"]).first()
        evaluation_id = evaluation.id
        first_section = evaluation.ciclo.version.clausulas[0].apartados[0]
        section_id = first_section.id
        reactive_ids = [reactive.id for reactive in first_section.reactivos]

    client.get("/auth/logout", follow_redirects=True)
    login(client, "operativo.iso@test.local")
    detail = client.get(f"/iso9001/evaluaciones/{evaluation_id}")
    assert detail.status_code == 200
    assert "Captura formal" in detail.get_data(as_text=True)

    section_payload = MultiDict([("apartado_id", str(section_id))])
    for reactive_id in reactive_ids:
        section_payload.add(f"calificacion_{reactive_id}", "si")
        section_payload.add(f"observacion_{reactive_id}", f"Respuesta operativa {reactive_id}")
    save_response = client.post(
        f"/iso9001/evaluaciones/{evaluation_id}",
        data=section_payload,
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert save_response.status_code == 200
    assert "Apartado guardado." in save_response.get_data(as_text=True)

    with app.app_context():
        fill_missing_responses(evaluation_id, ids["respondent_id"])

    submit_response = client.post(f"/iso9001/evaluaciones/{evaluation_id}/enviar", follow_redirects=True)
    assert submit_response.status_code == 200
    assert "Evaluación enviada a revisión." in submit_response.get_data(as_text=True)


def test_iso9001_admin_can_update_existing_evaluation_assignment():
    app, ids = build_app()
    client = app.test_client()

    with app.app_context():
        evaluation_id = create_iso_evaluation(
            ids["dependency_one_id"],
            ids["evaluator_id"],
            ids["reviewer_id"],
            ids["admin_id"],
        )
        evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
        cycle_id = evaluation.ciclo_id

    login(client, "admin.iso@test.local")
    cycles_page = client.get(f"/iso9001/ciclos?cycle_id={cycle_id}")
    cycles_html = cycles_page.get_data(as_text=True)
    assert cycles_page.status_code == 200
    assert f'data-dialog-open="dialog-iso-evaluation-edit-{evaluation_id}"' in cycles_html
    assert "Revisor Alterno ISO" in cycles_html

    invalid_update = client.post(
        "/iso9001/ciclos",
        data={
            "action": "update_evaluation",
            "evaluation_id": str(evaluation_id),
            "responsable_id": "999999",
            "revisor_id": str(ids["alternate_reviewer_id"]),
            "estado": "en_revision",
        },
        follow_redirects=True,
    )
    invalid_html = invalid_update.get_data(as_text=True)
    assert invalid_update.status_code == 200
    assert "Selecciona un responsable de captura activo." in invalid_html

    with app.app_context():
        evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
        alternate_reviewer = db.session.get(Usuario, ids["alternate_reviewer_id"])
        assert evaluation.responsable.id == ids["evaluator_id"]
        assert evaluation.revisor_id == ids["reviewer_id"]
        assert evaluation.estado == "borrador"
        assert alternate_reviewer.acceso_iso9001 is False

    update_response = client.post(
        "/iso9001/ciclos",
        data={
            "action": "update_evaluation",
            "evaluation_id": str(evaluation_id),
            "responsable_id": str(ids["respondent_id"]),
            "revisor_id": str(ids["alternate_reviewer_id"]),
            "estado": "en_revision",
        },
        follow_redirects=False,
    )
    assert update_response.status_code == 302
    assert f"/iso9001/ciclos?cycle_id={cycle_id}" in update_response.headers["Location"]

    updated_page = client.get(update_response.headers["Location"])
    updated_html = updated_page.get_data(as_text=True)
    assert updated_page.status_code == 200
    assert "Operativo Disponible" in updated_html
    assert "Revisor Alterno ISO" in updated_html
    assert "En revisión" in updated_html

    with app.app_context():
        evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
        old_responsible_assignments = [
            assignment
            for assignment in evaluation.asignaciones
            if assignment.tipo == "captura" and assignment.usuario_id == ids["evaluator_id"]
        ]
        respondent = db.session.get(Usuario, ids["respondent_id"])
        alternate_reviewer = db.session.get(Usuario, ids["alternate_reviewer_id"])
        assert old_responsible_assignments == []
        assert evaluation.responsable.id == ids["respondent_id"]
        assert evaluation.revisor_id == ids["alternate_reviewer_id"]
        assert evaluation.estado == "en_revision"
        assert respondent.acceso_iso9001 is True
        assert alternate_reviewer.acceso_iso9001 is True

    clear_response = client.post(
        "/iso9001/ciclos",
        data={
            "action": "update_evaluation",
            "evaluation_id": str(evaluation_id),
            "responsable_id": "",
            "revisor_id": "",
            "estado": "devuelta",
        },
        follow_redirects=True,
    )
    clear_html = clear_response.get_data(as_text=True)
    assert clear_response.status_code == 200
    assert "Sin responsable" in clear_html
    assert "Sin revisor" in clear_html

    with app.app_context():
        evaluation = db.session.get(Iso9001Evaluacion, evaluation_id)
        assert evaluation.responsable is None
        assert evaluation.revisor_id is None
        assert evaluation.estado == "devuelta"


def test_iso9001_capture_review_close_permissions_and_exports():
    app, ids = build_app()
    client = app.test_client()

    login(client, "admin.iso@test.local")
    cycle_response = client.post(
        "/iso9001/ciclos",
        data={
            "action": "create_cycle",
            "nombre": "Ciclo ISO Integral 2026",
            "descripcion": "Diagnóstico anual",
            "estado": "activo",
            "fecha_inicio": "2026-01-01",
            "fecha_cierre": "2026-12-31",
        },
        follow_redirects=True,
    )
    assert cycle_response.status_code == 200
    assert "Ciclo ISO registrado." in cycle_response.get_data(as_text=True)

    with app.app_context():
        cycle = Iso9001Ciclo.query.filter_by(nombre="Ciclo ISO Integral 2026").first()
        cycle_id = cycle.id

    assignment_response = client.post(
        "/iso9001/ciclos",
        data={
            "action": "add_evaluations",
            "cycle_id": str(cycle_id),
            "dependencia_ids": [str(ids["dependency_one_id"])],
            "responsable_id": str(ids["evaluator_id"]),
            "revisor_id": str(ids["reviewer_id"]),
        },
        follow_redirects=True,
    )
    assert assignment_response.status_code == 200
    assert "Evaluaciones registradas: 1." in assignment_response.get_data(as_text=True)

    with app.app_context():
        evaluation = Iso9001Evaluacion.query.filter_by(ciclo_id=cycle_id, dependencia_id=ids["dependency_one_id"]).first()
        evaluation_id = evaluation.id
        first_section = evaluation.ciclo.version.clausulas[0].apartados[0]
        section_id = first_section.id
        reactive_ids = [reactive.id for reactive in first_section.reactivos]

    client.get("/auth/logout", follow_redirects=True)
    login(client, "captura.iso@test.local")
    detail = client.get(f"/iso9001/evaluaciones/{evaluation_id}")
    assert detail.status_code == 200
    assert "Captura formal" in detail.get_data(as_text=True)

    data = MultiDict([("apartado_id", str(section_id))])
    for reactive_id in reactive_ids:
        data.add(f"calificacion_{reactive_id}", "si")
        data.add(f"observacion_{reactive_id}", f"Hallazgo {reactive_id}")
    data.add(f"evidencias_{reactive_ids[0]}", (BytesIO(b"evidencia uno"), "evidencia-uno.pdf"))
    data.add(f"evidencias_{reactive_ids[0]}", (BytesIO(b"evidencia dos"), "evidencia-dos.pdf"))
    save_response = client.post(
        f"/iso9001/evaluaciones/{evaluation_id}",
        data=data,
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert save_response.status_code == 200
    assert "Apartado guardado." in save_response.get_data(as_text=True)

    with app.app_context():
        saved_response = Iso9001Respuesta.query.filter_by(
            evaluacion_id=evaluation_id,
            reactivo_id=reactive_ids[0],
        ).first()
        assert saved_response is not None
        assert len(saved_response.evidencias) == 2
        fill_missing_responses(evaluation_id, ids["evaluator_id"])

    submit_response = client.post(f"/iso9001/evaluaciones/{evaluation_id}/enviar", follow_redirects=True)
    assert submit_response.status_code == 200
    assert "Evaluación enviada a revisión." in submit_response.get_data(as_text=True)

    with app.app_context():
        assert db.session.get(Iso9001Evaluacion, evaluation_id).estado == "en_revision"

    client.get("/auth/logout", follow_redirects=True)
    login(client, "revisor.iso@test.local")
    return_response = client.post(
        f"/iso9001/evaluaciones/{evaluation_id}/revision",
        data={"action": "return", "comentario": "Corregir evidencia del apartado inicial."},
        follow_redirects=True,
    )
    assert return_response.status_code == 200
    assert "Evaluación devuelta." in return_response.get_data(as_text=True)

    client.get("/auth/logout", follow_redirects=True)
    login(client, "captura.iso@test.local")
    correction = MultiDict([("apartado_id", str(section_id))])
    for reactive_id in reactive_ids:
        correction.add(f"calificacion_{reactive_id}", "parcial")
        correction.add(f"observacion_{reactive_id}", f"Corrección {reactive_id}")
    correction_response = client.post(
        f"/iso9001/evaluaciones/{evaluation_id}",
        data=correction,
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert correction_response.status_code == 200
    assert "Apartado guardado." in correction_response.get_data(as_text=True)

    submit_again = client.post(f"/iso9001/evaluaciones/{evaluation_id}/enviar", follow_redirects=True)
    assert submit_again.status_code == 200
    assert "Evaluación enviada a revisión." in submit_again.get_data(as_text=True)

    client.get("/auth/logout", follow_redirects=True)
    login(client, "revisor.iso@test.local")
    close_response = client.post(
        f"/iso9001/evaluaciones/{evaluation_id}/revision",
        data={"action": "close", "comentario": "Cierre oficial del diagnóstico."},
        follow_redirects=True,
    )
    assert close_response.status_code == 200
    assert "Evaluación cerrada como resultado oficial." in close_response.get_data(as_text=True)

    pdf = client.get(f"/iso9001/reportes/{evaluation_id}/pdf")
    assert pdf.status_code == 200
    assert pdf.mimetype == "application/pdf"
    assert pdf.data.startswith(b"%PDF")

    xlsx = client.get(f"/iso9001/reportes/{evaluation_id}/xlsx")
    assert xlsx.status_code == 200
    assert xlsx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(xlsx.data))
    assert "Resumen" in workbook.sheetnames
    assert "Cuestionario" in workbook.sheetnames
    detail_values = [
        value
        for row in workbook["Cuestionario"].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    ]
    assert any("Corrección" in value for value in detail_values)

    with app.app_context():
        open_evaluation = Iso9001Evaluacion(
            ciclo_id=cycle_id,
            dependencia_id=ids["dependency_two_id"],
            estado="borrador",
        )
        db.session.add(open_evaluation)
        db.session.commit()
        open_evaluation_id = open_evaluation.id

    client.get("/auth/logout", follow_redirects=True)
    login(client, "captura.iso@test.local")
    assert client.get(f"/iso9001/evaluaciones/{evaluation_id}").status_code == 200
    assert client.get(f"/iso9001/evaluaciones/{open_evaluation_id}").status_code == 403
    evaluator_report = client.get(f"/iso9001/reportes?cycle_id={cycle_id}")
    evaluator_report_html = evaluator_report.get_data(as_text=True)
    assert evaluator_report.status_code == 200
    assert "Contraloria" in evaluator_report_html
    assert "Secretaria Tecnica" not in evaluator_report_html

    client.get("/auth/logout", follow_redirects=True)
    login(client, "consulta.iso@test.local")
    assert client.get(f"/iso9001/evaluaciones/{evaluation_id}").status_code == 200
    assert client.get(f"/iso9001/evaluaciones/{open_evaluation_id}").status_code == 403

    client.get("/auth/logout", follow_redirects=True)
    login(client, "sin.iso@test.local")
    denied = client.get("/iso9001/")
    assert denied.status_code == 403
    assert "ISO 9001:2015" in denied.get_data(as_text=True)

import zipfile
from datetime import timedelta
from io import BytesIO

from openpyxl import load_workbook

from municipal_diagnostico import create_app
from municipal_diagnostico.extensions import db
from municipal_diagnostico.models import BienestarEncuesta, BienestarPregunta, BienestarRespuesta, Usuario
from municipal_diagnostico.services.wellbeing import build_wellbeing_report_payload
from municipal_diagnostico.timeutils import utcnow


class TestConfig:
    SECRET_KEY = "test"
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = "tests/uploads"
    ALLOWED_EXTENSIONS = {"pdf"}
    AUTO_INIT_DATABASE = True
    BOOTSTRAP_ADMIN_EMAIL = None
    BOOTSTRAP_ADMIN_PASSWORD = None
    BOOTSTRAP_ADMIN_NAME = None


def login(client, email: str, password: str = "secret123"):
    return client.post(
        "/auth/login",
        data={"correo": email, "password": password},
        follow_redirects=True,
    )


def build_app():
    app = create_app(TestConfig)
    with app.app_context():
        admin = Usuario(
            nombre="Admin",
            correo="admin@local.test",
            rol="administrador",
            activo=True,
            acceso_diagnostico=True,
            acceso_bienestar=True,
        )
        admin.set_password("secret123")
        consulta = Usuario(
            nombre="Consulta",
            correo="consulta@local.test",
            rol="consulta",
            activo=True,
            acceso_diagnostico=True,
            acceso_bienestar=True,
        )
        consulta.set_password("secret123")
        db.session.add_all([admin, consulta])
        db.session.commit()
    return app


def test_public_wellbeing_survey_can_complete_and_admin_can_open_dashboard():
    app = build_app()
    client = app.test_client()

    landing = client.get("/bienestar/")
    assert landing.status_code == 200
    assert "Bienestar Policial" in landing.get_data(as_text=True)
    public_link = client.get("/bienestar/publico")
    public_html = public_link.get_data(as_text=True)
    assert public_link.status_code == 200
    assert "Selecciona tu estrato y comienza la encuesta" in public_html
    assert "Iniciar encuesta" in public_html
    assert 'data-wellbeing-start-dialog' in public_html
    assert "URL lista para compartir" not in public_html
    assert "Abrir panel interno" not in public_html

    questions_payload = client.get("/bienestar/api/preguntas").get_json()
    questions = questions_payload["preguntas"]
    assert len(questions) == 45
    assert all(question["txt"].startswith("¿") and question["txt"].endswith("?") for question in questions)
    assert questions[0]["dim"] == "Bienestar Psicológico"
    assert len([question for question in questions if question["tipo_reactivo"] == "perfil"]) == 10
    assert all(question["dim"] == "Situación Socioeconómica" for question in questions[-10:])

    start_payload = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E3"}).get_json()
    folio = start_payload["hash"]
    assert folio

    save_payload = client.post(
        "/bienestar/api/encuesta/guardar",
        json={
            "hash": folio,
            "estado": "completada",
            "ultima_pregunta": len(questions),
            "respuestas": [{"id": question["id"], "dim": question["dim"], "val": 4} for question in questions],
        },
    ).get_json()
    assert save_payload["ok"] is True
    assert save_payload["estado"] == "completada"
    assert save_payload["iibp"] == 100.0
    assert save_payload["ivsp"] == 5.0

    survey_payload = client.get(f"/bienestar/api/encuesta/{folio}").get_json()
    assert survey_payload["estado"] == "completada"
    assert len(survey_payload["respuestas"]) == len(questions)

    survey_page = client.get(f"/bienestar/encuesta?folio={folio}")
    survey_html = survey_page.get_data(as_text=True)
    assert survey_page.status_code == 200
    assert "Menú" not in survey_html
    assert "Acceso interno" not in survey_html

    login(client, "admin@local.test")
    module_home = client.get("/bienestar/")
    assert module_home.status_code == 302
    assert module_home.headers["Location"].endswith("/bienestar/panel")
    dashboard = client.get("/bienestar/panel")
    html = dashboard.get_data(as_text=True)
    assert dashboard.status_code == 200
    assert "Panel interno de Bienestar Policial" in html
    assert "Cubo de datos" in html
    assert "/bienestar/publico" in html
    assert "Detalle de sesiones" in html

    sessions_detail = client.get("/bienestar/sesiones")
    sessions_html = sessions_detail.get_data(as_text=True)
    assert sessions_detail.status_code == 200
    assert "Detalle de sesiones del módulo" in sessions_html
    assert "Bitácora completa de sesiones" in sessions_html

    dashboard_api = client.get("/bienestar/api/dashboard")
    assert dashboard_api.status_code == 200
    dashboard_payload = dashboard_api.get_json()
    assert dashboard_payload["summary"]["total"] >= 1
    assert len(dashboard_payload["survey_rows"]) >= 1
    assert "question_catalog" in dashboard_payload
    assert dashboard_payload["survey_rows"][0]["fecha"].endswith("AM") or dashboard_payload["survey_rows"][0]["fecha"].endswith("PM")

    pdf_export = client.get("/bienestar/exportar/pdf")
    assert pdf_export.status_code == 200
    assert pdf_export.mimetype == "application/pdf"
    assert pdf_export.data.startswith(b"%PDF")

    xlsx_export = client.get("/bienestar/exportar/xlsx")
    assert xlsx_export.status_code == 200
    assert xlsx_export.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(xlsx_export.data))
    assert "Resumen" in workbook.sheetnames
    assert "Estratos" in workbook.sheetnames
    assert "Reactivos" in workbook.sheetnames
    assert "Perfil socioeconómico" in workbook.sheetnames
    assert "Perfil por estrato" in workbook.sheetnames
    summary_values = [
        value
        for row in workbook["Resumen"].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    ]
    assert any("Reporte ejecutivo de Bienestar Policial" in value for value in summary_values)

    word_export = client.get("/bienestar/exportar/word")
    assert word_export.status_code == 200
    assert word_export.mimetype == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    with zipfile.ZipFile(BytesIO(word_export.data)) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8")
    assert "Reporte ejecutivo de Bienestar Policial" in document_xml

    csv_export = client.get("/bienestar/exportar/csv")
    assert csv_export.status_code == 200
    assert csv_export.mimetype == "text/csv"
    assert "pregunta_1" in csv_export.get_data(as_text=True)

    with app.app_context():
        survey_record = BienestarEncuesta.query.filter_by(hash_id=folio).first()
        assert survey_record is not None
        assert survey_record.estado == "completada"
        report = build_wellbeing_report_payload()
        assert len(report["question_rows"]) == BienestarPregunta.query.count()
        assert any(item["stratum"] == "E3" and item["completed"] == 1 for item in report["strata"])
        assert report["question_rows"][0]["by_stratum"]["E3"]["count"] == 1
        assert report["summary"]["history"][0]["fecha"].endswith("AM") or report["summary"]["history"][0]["fecha"].endswith("PM")
        assert report["question_rows"][0]["response_options"][0]["label"] == questions[0]["t_opc"][0]
        assert report["question_rows"][0]["response_options"][0]["count"] == 1
        assert report["question_rows"][0]["response_options"][0]["percent"] == 100.0
        assert len(report["profile_socioeconomico"]["questions"]) == 10
        assert report["profile_socioeconomico"]["questions"][0]["tipo_reactivo"] == "perfil"
        assert report["profile_socioeconomico"]["questions"][0]["response_options"][0]["count"] == 1


def test_admin_can_manage_wellbeing_questions_and_consulta_is_read_only():
    app = build_app()
    client = app.test_client()

    login(client, "admin@local.test")
    response = client.post(
        "/bienestar/preguntas",
        data={
            "action": "add",
            "orden": 99,
            "dimension": "Clima Laboral",
            "texto": "Existe un ambiente colaborativo en su unidad?",
            "tipo_reactivo": "indicador",
            "opcion_1": "Siempre",
            "opcion_2": "Casi siempre",
            "opcion_3": "A veces",
            "opcion_4": "Nunca",
        },
        follow_redirects=True,
    )
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert "Pregunta creada." in html
    assert "Clima Laboral" in html

    with app.app_context():
        question = BienestarPregunta.query.filter_by(orden=99).first()
        assert question is not None
        assert question.texto == "¿Existe un ambiente colaborativo en su unidad?"
        question_id = question.id

    toggle = client.post(
        "/bienestar/preguntas",
        data={"action": "toggle", "question_id": question_id},
        follow_redirects=True,
    )
    assert toggle.status_code == 200
    assert "Pregunta desactivada." in toggle.get_data(as_text=True)

    client.get("/auth/logout", follow_redirects=True)
    login(client, "consulta@local.test")
    assert client.get("/bienestar/preguntas").status_code == 403
    dashboard = client.get("/bienestar/panel")
    assert dashboard.status_code == 200
    assert "Panel interno de Bienestar Policial" in dashboard.get_data(as_text=True)
    assert client.get("/bienestar/exportar/pdf").status_code == 200
    assert client.get("/bienestar/exportar/xlsx").status_code == 200
    assert client.get("/bienestar/exportar/word").status_code == 200


def test_partial_progress_can_be_recovered_by_folio():
    app = build_app()
    client = app.test_client()

    questions = client.get("/bienestar/api/preguntas").get_json()["preguntas"]
    folio = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E2"}).get_json()["hash"]

    save_response = client.post(
        "/bienestar/api/encuesta/guardar",
        json={
            "hash": folio,
            "estado": "en_progreso",
            "ultima_pregunta": 3,
            "respuestas": [{"id": question["id"], "dim": question["dim"], "val": 3} for question in questions[:3]],
        },
    )
    assert save_response.status_code == 200
    payload = save_response.get_json()
    assert payload["estado"] == "en_progreso"

    recovery = client.get(f"/bienestar/api/encuesta/{folio}").get_json()
    assert recovery["estado"] == "en_progreso"
    assert recovery["ultima_pregunta"] == 3
    assert len(recovery["respuestas"]) == 3


def test_admin_can_prune_abandoned_surveys_without_touching_active_or_completed_ones():
    app = build_app()
    client = app.test_client()

    questions = client.get("/bienestar/api/preguntas").get_json()["preguntas"]
    all_answers = [{"id": question["id"], "dim": question["dim"], "val": 4} for question in questions]

    completed_folio = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E1"}).get_json()["hash"]
    abandoned_folio_one = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E2"}).get_json()["hash"]
    abandoned_folio_two = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E3"}).get_json()["hash"]
    progress_folio = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E4"}).get_json()["hash"]

    completed_response = client.post(
        "/bienestar/api/encuesta/guardar",
        json={
            "hash": completed_folio,
            "estado": "completada",
            "ultima_pregunta": len(questions),
            "respuestas": all_answers,
        },
    )
    assert completed_response.status_code == 200

    progress_response = client.post(
        "/bienestar/api/encuesta/guardar",
        json={
            "hash": progress_folio,
            "estado": "en_progreso",
            "ultima_pregunta": 4,
            "respuestas": all_answers[:4],
        },
    )
    assert progress_response.status_code == 200

    login(client, "admin@local.test")
    dashboard = client.get("/bienestar/panel")
    dashboard_html = dashboard.get_data(as_text=True)
    assert dashboard.status_code == 200
    assert "wellbeing-prune-abandoned" in dashboard_html
    assert "Acciones" in dashboard_html

    prune_response = client.post(
        "/bienestar/api/encuestas/depurar",
        json={"folios": [abandoned_folio_one, abandoned_folio_two]},
    )
    assert prune_response.status_code == 200
    prune_payload = prune_response.get_json()
    assert prune_payload["ok"] is True
    assert prune_payload["deleted_count"] == 2
    assert set(prune_payload["deleted_folios"]) == {abandoned_folio_one, abandoned_folio_two}

    dashboard_payload = client.get("/bienestar/api/dashboard").get_json()
    assert dashboard_payload["summary"]["total"] == 2
    assert dashboard_payload["summary"]["abandonadas"] == 0
    remaining_folios = {row["hash"] for row in dashboard_payload["survey_rows"]}
    assert remaining_folios == {completed_folio, progress_folio}

    protected_response = client.post(
        "/bienestar/api/encuestas/depurar",
        json={"folios": [completed_folio, progress_folio]},
    )
    assert protected_response.status_code == 400
    protected_payload = protected_response.get_json()
    assert "Solo puedes depurar sesiones abandonadas" in protected_payload["mensaje"]

    with app.app_context():
        saved_surveys = {survey.hash_id: survey.estado for survey in BienestarEncuesta.query.all()}
        assert saved_surveys == {
            completed_folio: "completada",
            progress_folio: "en_progreso",
        }

    client.get("/auth/logout", follow_redirects=True)
    login(client, "consulta@local.test")
    assert client.post("/bienestar/api/encuestas/depurar", json={"folios": [completed_folio]}).status_code == 403


def test_stale_progress_is_reclassified_as_abandoned_after_8_hours():
    app = build_app()
    client = app.test_client()

    questions = client.get("/bienestar/api/preguntas").get_json()["preguntas"]
    stale_folio = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E1"}).get_json()["hash"]

    save_response = client.post(
        "/bienestar/api/encuesta/guardar",
        json={
            "hash": stale_folio,
            "estado": "en_progreso",
            "ultima_pregunta": 1,
            "respuestas": [{"id": questions[0]["id"], "dim": questions[0]["dim"], "val": 3}],
        },
    )
    assert save_response.status_code == 200
    assert save_response.get_json()["estado"] == "en_progreso"

    with app.app_context():
        survey = BienestarEncuesta.query.filter_by(hash_id=stale_folio).first()
        survey.updated_at = utcnow() - timedelta(hours=9)
        db.session.commit()

    recovery = client.get(f"/bienestar/api/encuesta/{stale_folio}").get_json()
    assert recovery["estado"] == "abandonada"

    login(client, "admin@local.test")
    dashboard_payload = client.get("/bienestar/api/dashboard").get_json()
    stale_row = next(row for row in dashboard_payload["survey_rows"] if row["hash"] == stale_folio)
    assert stale_row["estado"] == "abandonada"
    assert stale_row["estado_label"] == "Abandonada"
    assert dashboard_payload["summary"]["abandonadas"] == 1
    assert dashboard_payload["summary"]["en_progreso"] == 0

    prune_response = client.post("/bienestar/api/encuestas/depurar", json={"folios": [stale_folio]})
    assert prune_response.status_code == 200
    assert prune_response.get_json()["deleted_folios"] == [stale_folio]

    with app.app_context():
        assert BienestarEncuesta.query.filter_by(hash_id=stale_folio).first() is None


def test_profile_questions_do_not_change_iibp_and_legacy_completed_surveys_stay_valid():
    app = build_app()
    client = app.test_client()

    questions = client.get("/bienestar/api/preguntas").get_json()["preguntas"]
    indicator_questions = [question for question in questions if question["tipo_reactivo"] == "indicador"]
    profile_questions = [question for question in questions if question["tipo_reactivo"] == "perfil"]

    folio_a = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E1"}).get_json()["hash"]
    folio_b = client.post("/bienestar/api/encuesta/iniciar", json={"estrato": "E2"}).get_json()["hash"]

    client.post(
        "/bienestar/api/encuesta/guardar",
        json={
            "hash": folio_a,
            "estado": "completada",
            "ultima_pregunta": len(questions),
            "respuestas": (
                [{"id": question["id"], "dim": question["dim"], "val": 4} for question in indicator_questions]
                + [{"id": question["id"], "dim": question["dim"], "val": 4} for question in profile_questions]
            ),
        },
    )
    second_payload = client.post(
        "/bienestar/api/encuesta/guardar",
        json={
            "hash": folio_b,
            "estado": "completada",
            "ultima_pregunta": len(questions),
            "respuestas": (
                [{"id": question["id"], "dim": question["dim"], "val": 4} for question in indicator_questions]
                + [{"id": question["id"], "dim": question["dim"], "val": 1} for question in profile_questions]
            ),
        },
    ).get_json()
    assert second_payload["iibp"] == 100.0
    assert second_payload["ivsp"] == 5.0

    with app.app_context():
        legacy = BienestarEncuesta(
            hash_id="LEGACY35",
            estrato="E4",
            estado="completada",
            iibp=82.5,
            ivsp=22.5,
            ultima_pregunta=35,
        )
        db.session.add(legacy)
        db.session.flush()
        for question in indicator_questions:
            db.session.add(
                BienestarRespuesta(
                    encuesta_id=legacy.id,
                    pregunta_id=question["id"],
                    dimension=question["dim"],
                    valor=3,
                )
            )
        db.session.commit()

        report = build_wellbeing_report_payload()
        legacy_row = next(row for row in report["survey_rows"] if row["hash"] == "LEGACY35")
        assert legacy_row["completion_percent"] == 100.0
        assert len(report["profile_socioeconomico"]["questions"]) == 10
        assert report["profile_socioeconomico"]["questions"][0]["response_options"][0]["count"] >= 1

from __future__ import annotations

from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.graphics.shapes import Drawing, Line, Rect, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from municipal_diagnostico.models import BienestarPregunta
from municipal_diagnostico.services.exports import (
    EXCEL_THEME,
    PDF_THEME,
    WORD_THEME,
    _docx_header_row,
    _docx_page_break,
    _docx_paragraph,
    _docx_table,
    _package_docx,
)
from municipal_diagnostico.services.wellbeing import (
    build_wellbeing_report_payload,
)


THIN_SIDE = Side(style="thin", color=EXCEL_THEME["line"])
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
WELLBEING_PDF_FOOTER = "Bienestar Policial | Ayuntamiento de Hermosillo"


def build_wellbeing_pdf(*, public_url: str | None = None) -> BytesIO:
    report = build_wellbeing_report_payload()
    summary = report["summary"]
    profile_questions = report["profile_socioeconomico"]["questions"]

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.55 * inch,
        rightMargin=0.55 * inch,
        topMargin=0.65 * inch,
        bottomMargin=0.55 * inch,
    )

    styles = _build_pdf_styles()
    story: list = [
        Paragraph("Reporte ejecutivo de Bienestar Policial", styles["title"]),
        Paragraph(
            "Consolidado institucional de captura anónima, métricas de bienestar y hallazgos operativos por estrato.",
            styles["subtitle"],
        ),
    ]
    if public_url:
        story.append(Paragraph(f"Liga pública de respuesta: {escape(public_url)}", styles["note"]))

    story.append(Spacer(1, 0.16 * inch))
    story.append(_build_pdf_metric_cards(summary, styles))
    story.append(Spacer(1, 0.18 * inch))

    story.append(Paragraph("Lectura ejecutiva", styles["section"]))
    for note in report["executive_notes"]:
        story.append(Paragraph(f"• {escape(note)}", styles["bullet"]))

    story.append(Spacer(1, 0.14 * inch))
    story.append(Paragraph("Tablero consolidado por estrato", styles["section"]))
    story.append(
        Paragraph(
            "El siguiente corte resume la muestra completada, los promedios IIBP e IVSP y las principales señales de atención para cada estrato.",
            styles["body"],
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_build_pdf_strata_summary_table(report["strata"]))
    story.append(Spacer(1, 0.12 * inch))

    chart_row = Table(
        [[_build_pdf_completion_chart(report["strata"]), _build_pdf_score_chart(report["strata"])]],
        colWidths=[4.85 * inch, 4.85 * inch],
    )
    chart_row.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(chart_row)
    story.append(Spacer(1, 0.18 * inch))

    story.append(Paragraph("Bitácora reciente", styles["section"]))
    story.append(
        Paragraph(
            "Se muestran las sesiones más recientes para facilitar seguimiento operativo y trazabilidad de la captura.",
            styles["body"],
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_build_pdf_history_table(summary["history"]))

    story.append(PageBreak())
    story.append(Paragraph("Análisis ejecutivo por estrato", styles["section"]))
    story.append(
        Paragraph(
            "Cada estrato incorpora volumen de muestra, promedio institucional y desempeño porcentual por dimensión para orientar decisiones focalizadas.",
            styles["body"],
        )
    )
    for stratum in report["strata"]:
        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph(f"Estrato {escape(stratum['stratum'])}", styles["subsection"]))
        story.append(Paragraph(_build_stratum_summary_text(stratum), styles["body"]))
        story.append(Spacer(1, 0.06 * inch))
        story.append(_build_pdf_stratum_metrics_table(stratum))
        story.append(Spacer(1, 0.06 * inch))
        story.append(_build_pdf_dimension_chart(stratum))

    story.append(PageBreak())
    story.append(Paragraph("Anexo técnico de reactivos", styles["section"]))
    story.append(
        Paragraph(
            "El anexo incorpora todos los reactivos del módulo, con su estado actual y el promedio global y por estrato derivado de la muestra completada.",
            styles["body"],
        )
    )
    for group in report["question_groups"]:
        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph(escape(group["dimension"]), styles["subsection"]))
        story.append(_build_pdf_question_table(group["rows"], report["strata_order"], styles))
        story.append(Spacer(1, 0.08 * inch))
        story.append(
            Paragraph(
                "Distribución de respuesta por reactivo con las opciones reales del cuestionario.",
                styles["body"],
            )
        )
        for row_chunk in _chunked(group["rows"], 2):
            story.append(Spacer(1, 0.05 * inch))
            story.append(_build_pdf_question_panel_row(row_chunk, styles))

    story.append(PageBreak())
    story.append(Paragraph("Composición económica del hogar", styles["section"]))
    story.append(
        Paragraph(
            "Este capítulo consolida los reactivos de perfil socioeconómico. Sus respuestas aportan contexto del hogar y no modifican IIBP ni IVSP.",
            styles["body"],
        )
    )
    for question in profile_questions:
        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph(f"Reactivo {question['orden']}", styles["subsection"]))
        story.append(Paragraph(escape(question["texto"]), styles["body"]))
        story.append(Spacer(1, 0.04 * inch))
        story.append(_build_pdf_profile_distribution_chart(question))
        story.append(Spacer(1, 0.05 * inch))
        story.append(_build_pdf_profile_distribution_table(question, report["strata_order"]))

    document.build(story, onFirstPage=_draw_wellbeing_pdf_chrome, onLaterPages=_draw_wellbeing_pdf_chrome)
    buffer.seek(0)
    return buffer


def build_wellbeing_excel(*, public_url: str | None = None) -> BytesIO:
    report = build_wellbeing_report_payload()
    summary = report["summary"]
    questions = BienestarPregunta.query.order_by(BienestarPregunta.orden.asc()).all()
    profile_questions = report["profile_socioeconomico"]["questions"]
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    _write_title(summary_sheet, "Reporte ejecutivo de Bienestar Policial", "Consolidado general", 4)
    summary_sheet.append([])
    summary_sheet.append(["Indicador", "Valor"])
    summary_sheet.append(["Total de sesiones", summary["total"]])
    summary_sheet.append(["Encuestas completadas", summary["completadas"]])
    summary_sheet.append(["Encuestas en progreso", summary["en_progreso"]])
    summary_sheet.append(["Encuestas abandonadas", summary["abandonadas"]])
    summary_sheet.append(["IIBP promedio", summary["avg_iibp"]])
    summary_sheet.append(["IVSP promedio", summary["avg_ivsp"]])
    summary_sheet.append(["Tasa de finalización", summary["completion_rate"]])
    if public_url:
        summary_sheet.append(["Liga pública", public_url])
    _style_table(summary_sheet, header_row=3)
    _set_widths(summary_sheet, {"A": 28, "B": 42})

    dimensions_sheet = workbook.create_sheet("Dimensiones")
    dimensions_sheet.append(["Dimensión", "Promedio", "Puntaje", "Respuestas"])
    for row in summary["dimensions"]:
        dimensions_sheet.append([row["name"], row["average"], row["percent"], row["count"]])
    _style_table(dimensions_sheet, header_row=1)
    _set_widths(dimensions_sheet, {"A": 34, "B": 12, "C": 12, "D": 12})

    strata_sheet = workbook.create_sheet("Estratos")
    strata_sheet.append(["Estrato", "Completadas", "Participación", "IIBP", "IVSP", "Dimensión fuerte", "Dimensión a atender"])
    for row in report["strata"]:
        strata_sheet.append(
            [
                row["stratum"],
                row["completed"],
                row["share_of_completed"],
                row["avg_iibp"] if row["avg_iibp"] is not None else "",
                row["avg_ivsp"] if row["avg_ivsp"] is not None else "",
                row["strongest_dimension"] or "",
                row["attention_dimension"] or "",
            ]
        )
    _style_table(strata_sheet, header_row=1)
    _set_widths(strata_sheet, {"A": 10, "B": 12, "C": 14, "D": 10, "E": 10, "F": 28, "G": 28})

    reactives_sheet = workbook.create_sheet("Reactivos")
    reactive_header = ["Reactivo", "Estado", "Clasificación", "Dimensión", "Texto", "Global"]
    reactive_header.extend(report["strata_order"])
    reactives_sheet.append(reactive_header)
    for row in report["question_rows"]:
        values = [
            row["orden"],
            row["state_label"],
            row["tipo_reactivo_label"],
            row["dimension"],
            row["texto"],
            _metric_cell_text(row["overall"]),
        ]
        values.extend(_metric_cell_text(row["by_stratum"][stratum]) for stratum in report["strata_order"])
        reactives_sheet.append(values)
    _style_table(reactives_sheet, header_row=1)
    widths = {"A": 10, "B": 12, "C": 14, "D": 24, "E": 54, "F": 14}
    for index in range(7, len(reactive_header) + 1):
        widths[_column_letter(index)] = 11
    _set_widths(reactives_sheet, widths)

    profile_sheet = workbook.create_sheet("Perfil socioeconómico")
    profile_sheet.append(["Reactivo", "Texto", "Opción", "Conteo global", "Porcentaje global"])
    for question in profile_questions:
        for option in question["response_options"]:
            profile_sheet.append(
                [
                    f"R{question['orden']}",
                    question["texto"],
                    option["label"],
                    option["count"],
                    option["percent"],
                ]
            )
    if profile_sheet.max_row == 1:
        profile_sheet.append(["Sin datos", "-", "-", 0, 0])
    _style_table(profile_sheet, header_row=1)
    _set_widths(profile_sheet, {"A": 12, "B": 56, "C": 28, "D": 16, "E": 18})

    profile_strata_sheet = workbook.create_sheet("Perfil por estrato")
    profile_strata_sheet.append(["Reactivo", "Opción", "Estrato", "Conteo", "Porcentaje"])
    for question in profile_questions:
        for option in question["response_options"]:
            for stratum in report["strata_order"]:
                detail = option["by_stratum"].get(stratum, {"count": 0, "percent": 0.0})
                profile_strata_sheet.append(
                    [
                        f"R{question['orden']}",
                        option["label"],
                        stratum,
                        detail["count"],
                        detail["percent"],
                    ]
                )
    if profile_strata_sheet.max_row == 1:
        profile_strata_sheet.append(["Sin datos", "-", "-", 0, 0])
    _style_table(profile_strata_sheet, header_row=1)
    _set_widths(profile_strata_sheet, {"A": 12, "B": 28, "C": 12, "D": 12, "E": 14})

    history_sheet = workbook.create_sheet("Historial")
    history_sheet.append(["Folio", "Fecha", "Estrato", "Estado", "Última pregunta", "IIBP", "IVSP"])
    for row in summary["history"]:
        history_sheet.append(
            [
                row["hash"],
                row["fecha"],
                row["estrato"],
                row["estado_label"],
                row["ultima_pregunta"],
                row["iibp"] if row["iibp"] is not None else "",
                row["ivsp"] if row["ivsp"] is not None else "",
            ]
        )
    _style_table(history_sheet, header_row=1)
    _set_widths(history_sheet, {"A": 16, "B": 22, "C": 10, "D": 16, "E": 16, "F": 12, "G": 12})

    answers_sheet = workbook.create_sheet("Respuestas")
    answer_header = ["Folio", "Estrato", "Estado"]
    answer_header.extend([f"P{question.orden}" for question in questions])
    answers_sheet.append(answer_header)
    question_id_to_order = {question.id: question.orden for question in questions}
    for row in summary["history"]:
        response_values = [""] * len(questions)
        for question_id, value in row["respuestas"].items():
            order = question_id_to_order.get(question_id)
            if order:
                response_values[order - 1] = value
        answers_sheet.append([row["hash"], row["estrato"], row["estado_label"], *response_values])
    _style_table(answers_sheet, header_row=1)
    widths = {"A": 16, "B": 10, "C": 16}
    for index in range(4, len(answer_header) + 1):
        widths[_column_letter(index)] = 8
    _set_widths(answers_sheet, widths)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_wellbeing_word(*, public_url: str | None = None) -> BytesIO:
    report = build_wellbeing_report_payload()
    summary = report["summary"]
    profile_questions = report["profile_socioeconomico"]["questions"]
    body = [
        _docx_paragraph("Reporte ejecutivo de Bienestar Policial", bold=True, size=32, spacing_after=120),
        _docx_paragraph(
            "Consolidado institucional de captura anónima, métricas de bienestar y hallazgos por estrato.",
            size=20,
            color=WORD_THEME["muted"],
        ),
    ]
    if public_url:
        body.append(_docx_paragraph(f"Liga pública de respuesta: {public_url}", size=18, color=WORD_THEME["muted"]))

    body.append(_docx_paragraph("Resumen general", bold=True, size=24, spacing_before=80))
    body.append(
        _docx_table(
            [
                _docx_header_row(["Indicador", "Valor"]),
                ["Total de sesiones", str(summary["total"])],
                ["Encuestas completadas", str(summary["completadas"])],
                ["Encuestas en progreso", str(summary["en_progreso"])],
                ["Encuestas abandonadas", str(summary["abandonadas"])],
                ["IIBP promedio", str(summary["avg_iibp"])],
                ["IVSP promedio", str(summary["avg_ivsp"])],
                ["Tasa de finalización", f'{summary["completion_rate"]}%'],
            ],
            col_widths=[4200, 2200],
        )
    )

    body.append(_docx_paragraph("", size=4, spacing_after=80))
    body.append(_docx_paragraph("Resumen por estrato", bold=True, size=24))
    strata_rows = [_docx_header_row(["Estrato", "Completadas", "Participación", "IIBP", "IVSP"])]
    for row in report["strata"]:
        strata_rows.append(
            [
                row["stratum"],
                str(row["completed"]),
                f'{row["share_of_completed"]}%',
                str(row["avg_iibp"] if row["avg_iibp"] is not None else "-"),
                str(row["avg_ivsp"] if row["avg_ivsp"] is not None else "-"),
            ]
        )
    body.append(_docx_table(strata_rows, col_widths=[1200, 1400, 1400, 1200, 1200]))

    body.append(_docx_page_break())
    body.append(_docx_paragraph("Bitácora reciente", bold=True, size=24))
    history_rows = [_docx_header_row(["Folio", "Fecha", "Estrato", "Estado", "Últ. pregunta", "IIBP", "IVSP"])]
    for row in summary["history"][:40]:
        history_rows.append(
            [
                row["hash"],
                row["fecha"],
                row["estrato"],
                row["estado_label"],
                str(row["ultima_pregunta"]),
                str(row["iibp"] if row["iibp"] is not None else "-"),
                str(row["ivsp"] if row["ivsp"] is not None else "-"),
            ]
        )
    if len(history_rows) == 1:
        history_rows.append(["Sin datos", "-", "-", "-", "-", "-", "-"])
    body.append(_docx_table(history_rows, col_widths=[1400, 1800, 1000, 1400, 1200, 900, 900]))

    body.append(_docx_page_break())
    body.append(_docx_paragraph("Composición económica del hogar", bold=True, size=24))
    body.append(
        _docx_paragraph(
            "Este capítulo consolida los reactivos de perfil socioeconómico para dar contexto del hogar sin alterar IIBP ni IVSP.",
            size=18,
            color=WORD_THEME["muted"],
        )
    )
    for question in profile_questions:
        body.append(_docx_paragraph(f"Reactivo {question['orden']}", bold=True, size=20, spacing_before=60))
        body.append(_docx_paragraph(question["texto"], size=18))
        rows = [_docx_header_row(["Opción", "Global", *report["strata_order"]])]
        for option in question["response_options"]:
            cells = [option["label"], f"{option['count']} | {option['percent']}%"]
            for stratum in report["strata_order"]:
                detail = option["by_stratum"].get(stratum, {"count": 0, "percent": 0.0})
                cells.append(f"{detail['count']} | {detail['percent']}%")
            rows.append(cells)
        body.append(_docx_table(rows, col_widths=[2400, 1400, *([1000] * len(report["strata_order"]))]))

    return _package_docx("".join(body), title="Reporte ejecutivo de Bienestar Policial")


def _build_pdf_styles() -> dict[str, ParagraphStyle]:
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "WellbeingPdfTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            textColor=PDF_THEME["navy"],
            spaceAfter=6,
        ),
        "subtitle": ParagraphStyle(
            "WellbeingPdfSubtitle",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=PDF_THEME["muted"],
            spaceAfter=3,
        ),
        "section": ParagraphStyle(
            "WellbeingPdfSection",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=18,
            textColor=PDF_THEME["navy"],
            spaceBefore=4,
            spaceAfter=6,
        ),
        "subsection": ParagraphStyle(
            "WellbeingPdfSubsection",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=11.5,
            leading=14,
            textColor=PDF_THEME["navy"],
            spaceBefore=2,
            spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "WellbeingPdfBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=PDF_THEME["muted"],
        ),
        "bullet": ParagraphStyle(
            "WellbeingPdfBullet",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=PDF_THEME["muted"],
            leftIndent=10,
            bulletIndent=0,
            spaceAfter=2,
        ),
        "note": ParagraphStyle(
            "WellbeingPdfNote",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=PDF_THEME["muted"],
        ),
        "small_center": ParagraphStyle(
            "WellbeingPdfSmallCenter",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=7,
            leading=8,
            textColor=PDF_THEME["muted"],
            alignment=TA_CENTER,
        ),
        "metric_label": ParagraphStyle(
            "WellbeingPdfMetricLabel",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=PDF_THEME["muted"],
            alignment=TA_CENTER,
            spaceAfter=2,
        ),
        "metric_value": ParagraphStyle(
            "WellbeingPdfMetricValue",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=24,
            leading=26,
            textColor=PDF_THEME["navy"],
            alignment=TA_CENTER,
            spaceAfter=2,
        ),
        "metric_caption": ParagraphStyle(
            "WellbeingPdfMetricCaption",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=PDF_THEME["muted"],
            alignment=TA_CENTER,
        ),
    }


def _build_pdf_metric_cards(summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    cards = [
        ("Sesiones registradas", summary["total"], "Universo acumulado"),
        ("Encuestas completadas", summary["completadas"], "Base analítica del corte"),
        ("IIBP promedio", f'{summary["avg_iibp"]}', "Escala 0 a 100"),
        ("IVSP promedio", f'{summary["avg_ivsp"]}', "Escala 0 a 100"),
        ("En progreso", summary["en_progreso"], "Folios aún abiertos"),
        ("Abandonadas", summary["abandonadas"], "Sesiones sin cierre"),
        ("Finalización", f'{summary["completion_rate"]}%', "Conversión del módulo"),
        ("Estratos activos", sum(1 for value in summary["strata_counts"].values() if value), "Estratos con muestra"),
    ]

    rows = []
    for index in range(0, len(cards), 4):
        chunk = cards[index:index + 4]
        row = []
        for title, value, caption in chunk:
            card = Table(
                [
                    [Paragraph(escape(str(title).upper()), styles["metric_label"])],
                    [Paragraph(escape(str(value)), styles["metric_value"])],
                    [Paragraph(escape(str(caption)), styles["metric_caption"])],
                ],
                colWidths=[2.12 * inch],
                rowHeights=[0.3 * inch, 0.44 * inch, 0.28 * inch],
            )
            card.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 2),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            row.append(card)
        rows.append(row)

    table = Table(rows, colWidths=[2.38 * inch] * 4, rowHeights=[1.15 * inch] * len(rows))
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, PDF_THEME["navy_soft"]]),
                ("BOX", (0, 0), (-1, -1), 0.6, PDF_THEME["line"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]
        )
    )
    return table


def _build_pdf_strata_summary_table(strata: list[dict]) -> Table:
    rows = [
        [
            "Estrato",
            "Completadas",
            "Participación",
            "IIBP",
            "IVSP",
            "Dimensión fuerte",
            "Dimensión a atender",
        ]
    ]
    for row in strata:
        rows.append(
            [
                row["stratum"],
                str(row["completed"]),
                f'{row["share_of_completed"]}%',
                _value_or_dash(row["avg_iibp"]),
                _value_or_dash(row["avg_ivsp"]),
                row["strongest_dimension"] or "Sin muestra",
                row["attention_dimension"] or "Sin muestra",
            ]
        )

    table = Table(
        rows,
        colWidths=[0.7 * inch, 0.9 * inch, 0.95 * inch, 0.75 * inch, 0.75 * inch, 2.15 * inch, 2.3 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PDF_THEME["navy_soft"]]),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 1), (4, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _build_pdf_history_table(history: list[dict]) -> Table:
    rows = [["Folio", "Fecha", "Estrato", "Estado", "Últ. pregunta", "IIBP", "IVSP"]]
    for row in history[:12]:
        rows.append(
            [
                row["hash"],
                row["fecha"],
                row["estrato"],
                row["estado_label"],
                str(row["ultima_pregunta"]),
                _value_or_dash(row["iibp"]),
                _value_or_dash(row["ivsp"]),
            ]
        )
    if len(rows) == 1:
        rows.append(["Sin datos", "-", "-", "-", "-", "-", "-"])

    table = Table(
        rows,
        colWidths=[1.05 * inch, 1.35 * inch, 0.75 * inch, 1.05 * inch, 0.8 * inch, 0.75 * inch, 0.75 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["sage"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PDF_THEME["sage_soft"]]),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _build_pdf_stratum_metrics_table(stratum: dict) -> Table:
    metrics = [
        ["Completadas", str(stratum["completed"])],
        ["Participación", f'{stratum["share_of_completed"]}%'],
        ["IIBP", _value_or_dash(stratum["avg_iibp"])],
        ["IVSP", _value_or_dash(stratum["avg_ivsp"])],
        ["Fortaleza", stratum["strongest_dimension"] or "Sin muestra"],
        ["Atención", stratum["attention_dimension"] or "Sin muestra"],
    ]
    table = Table(metrics, colWidths=[1.1 * inch, 3.2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), PDF_THEME["navy_soft"]),
                ("TEXTCOLOR", (0, 0), (0, -1), PDF_THEME["navy"]),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _build_pdf_completion_chart(strata: list[dict]) -> Drawing:
    drawing = Drawing(350, 180)
    drawing.add(String(175, 164, "Encuestas completadas por estrato", fontName="Helvetica-Bold", fontSize=10, fillColor=PDF_THEME["navy"], textAnchor="middle"))

    plot_x = 38
    plot_y = 32
    plot_width = 274
    plot_height = 102
    drawing.add(Rect(plot_x, plot_y, plot_width, plot_height, fillColor=colors.white, strokeColor=PDF_THEME["line"], strokeWidth=1))

    max_completed = max((row["completed"] for row in strata), default=0)
    if max_completed <= 0:
        drawing.add(String(175, 92, "Sin encuestas completadas para graficar", fontName="Helvetica", fontSize=9, fillColor=PDF_THEME["muted"], textAnchor="middle"))
        return drawing

    ticks = 4
    for tick in range(ticks + 1):
        value = max_completed * tick / ticks
        y = plot_y + (plot_height * tick / ticks)
        drawing.add(Line(plot_x, y, plot_x + plot_width, y, strokeColor=PDF_THEME["line"], strokeWidth=0.5))
        drawing.add(String(plot_x - 8, y - 2, str(int(round(value))), fontName="Helvetica", fontSize=7, fillColor=PDF_THEME["muted"], textAnchor="end"))

    slot = plot_width / max(len(strata), 1)
    bar_width = min(26, slot * 0.55)
    for index, row in enumerate(strata):
        x = plot_x + slot * index + (slot - bar_width) / 2
        bar_height = plot_height * row["completed"] / max_completed
        drawing.add(Rect(x, plot_y, bar_width, bar_height, fillColor=PDF_THEME["sage"], strokeColor=PDF_THEME["sage"], strokeWidth=0))
        drawing.add(String(x + bar_width / 2, plot_y + bar_height + 6, str(row["completed"]), fontName="Helvetica-Bold", fontSize=7, fillColor=PDF_THEME["navy"], textAnchor="middle"))
        drawing.add(String(x + bar_width / 2, plot_y - 14, row["stratum"], fontName="Helvetica-Bold", fontSize=7.5, fillColor=PDF_THEME["muted"], textAnchor="middle"))

    return drawing


def _build_pdf_score_chart(strata: list[dict]) -> Drawing:
    drawing = Drawing(350, 180)
    drawing.add(String(175, 164, "Promedio IIBP e IVSP por estrato", fontName="Helvetica-Bold", fontSize=10, fillColor=PDF_THEME["navy"], textAnchor="middle"))
    drawing.add(Rect(210, 148, 8, 8, fillColor=PDF_THEME["sage"], strokeColor=PDF_THEME["sage"]))
    drawing.add(String(224, 149, "IIBP", fontName="Helvetica", fontSize=7.5, fillColor=PDF_THEME["muted"]))
    drawing.add(Rect(258, 148, 8, 8, fillColor=PDF_THEME["low"], strokeColor=PDF_THEME["low"]))
    drawing.add(String(272, 149, "IVSP", fontName="Helvetica", fontSize=7.5, fillColor=PDF_THEME["muted"]))

    plot_x = 38
    plot_y = 32
    plot_width = 274
    plot_height = 102
    drawing.add(Rect(plot_x, plot_y, plot_width, plot_height, fillColor=colors.white, strokeColor=PDF_THEME["line"], strokeWidth=1))

    for tick in range(5):
        value = 100 * tick / 4
        y = plot_y + (plot_height * tick / 4)
        drawing.add(Line(plot_x, y, plot_x + plot_width, y, strokeColor=PDF_THEME["line"], strokeWidth=0.5))
        drawing.add(String(plot_x - 8, y - 2, str(int(value)), fontName="Helvetica", fontSize=7, fillColor=PDF_THEME["muted"], textAnchor="end"))

    slot = plot_width / max(len(strata), 1)
    bar_width = min(11, slot * 0.23)
    for index, row in enumerate(strata):
        iibp = row["avg_iibp"] or 0
        ivsp = row["avg_ivsp"] or 0
        center = plot_x + slot * index + slot / 2
        iibp_x = center - bar_width - 2
        ivsp_x = center + 2
        drawing.add(Rect(iibp_x, plot_y, bar_width, plot_height * iibp / 100, fillColor=PDF_THEME["sage"], strokeColor=PDF_THEME["sage"], strokeWidth=0))
        drawing.add(Rect(ivsp_x, plot_y, bar_width, plot_height * ivsp / 100, fillColor=PDF_THEME["low"], strokeColor=PDF_THEME["low"], strokeWidth=0))
        drawing.add(String(center, plot_y - 14, row["stratum"], fontName="Helvetica-Bold", fontSize=7.5, fillColor=PDF_THEME["muted"], textAnchor="middle"))

    return drawing


def _build_pdf_dimension_chart(stratum: dict) -> Drawing:
    dimensions = stratum["dimensions"]
    drawing = Drawing(680, 170)
    drawing.add(String(2, 154, "Desempeño porcentual por dimensión", fontName="Helvetica-Bold", fontSize=9.5, fillColor=PDF_THEME["navy"]))

    if not any(row["count"] for row in dimensions):
        drawing.add(String(340, 84, "Sin muestra completada para este estrato", fontName="Helvetica", fontSize=9, fillColor=PDF_THEME["muted"], textAnchor="middle"))
        return drawing

    plot_x = 220
    plot_width = 390
    bar_height = 12
    start_y = 130
    gap = 20
    for index, row in enumerate(dimensions):
        y = start_y - index * gap
        drawing.add(String(4, y + 1, row["name"], fontName="Helvetica", fontSize=8, fillColor=PDF_THEME["muted"]))
        drawing.add(Rect(plot_x, y, plot_width, bar_height, fillColor=PDF_THEME["navy_soft"], strokeColor=PDF_THEME["line"], strokeWidth=0.3))
        if row["percent"] is not None:
            color = _score_color(row["percent"])
            drawing.add(Rect(plot_x, y, plot_width * row["percent"] / 100, bar_height, fillColor=color, strokeColor=color, strokeWidth=0))
            label = f"{row['percent']}% | n={row['count']}"
        else:
            label = "Sin datos"
        drawing.add(String(plot_x + plot_width + 8, y + 1, label, fontName="Helvetica-Bold", fontSize=7.5, fillColor=PDF_THEME["navy"]))

    return drawing


def _build_pdf_question_table(rows: list[dict], strata_order: list[str], styles: dict[str, ParagraphStyle]) -> Table:
    header = ["Reactivo", "Estado", "Pregunta", "Global"]
    header.extend(strata_order)
    table_rows = [header]

    for row in rows:
        body = [
            Paragraph(f"R{row['orden']}", styles["small_center"]),
            Paragraph(row["state_label"], styles["small_center"]),
            Paragraph(escape(row["texto"]), styles["note"]),
            Paragraph(_metric_cell_text(row["overall"]), styles["small_center"]),
        ]
        for stratum in strata_order:
            body.append(Paragraph(_metric_cell_text(row["by_stratum"][stratum]), styles["small_center"]))
        table_rows.append(body)

    widths = [0.52 * inch, 0.7 * inch, 3.55 * inch, 0.72 * inch]
    widths.extend([0.58 * inch] * len(strata_order))

    table = Table(table_rows, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PDF_THEME["navy_soft"]]),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.6),
                ("LEADING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.45, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 1), (1, -1), "CENTER"),
                ("ALIGN", (3, 1), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _build_pdf_question_panel_row(rows: list[dict], styles: dict[str, ParagraphStyle]) -> Table:
    panels = [_build_pdf_question_panel(row, styles) for row in rows]
    while len(panels) < 2:
        panels.append(Spacer(1, 0.01 * inch))

    table = Table([panels], colWidths=[4.85 * inch, 4.85 * inch])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return table


def _build_pdf_question_panel(row: dict, styles: dict[str, ParagraphStyle]):
    title = Paragraph(
        f"<b>R{row['orden']}</b> · {escape(row['texto'])}",
        styles["note"],
    )
    meta = Paragraph(
        (
            f"Estado: <b>{escape(row['state_label'])}</b> | "
            f"Global: <b>{_metric_cell_text(row['overall'])}</b> | "
            f"Respuestas observadas: <b>{row['overall']['count']}</b>"
        ),
        styles["note"],
    )
    distribution = _build_pdf_option_distribution_table(row["response_options"], styles)

    panel = Table(
        [
            [title],
            [meta],
            [distribution],
        ],
        colWidths=[4.72 * inch],
    )
    panel.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.55, PDF_THEME["line"]),
                ("ROUNDEDCORNERS", [10, 10, 10, 10]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return panel


def _build_pdf_option_distribution_table(response_options: list[dict], styles: dict[str, ParagraphStyle]) -> Table:
    if not response_options or not any(option["count"] for option in response_options):
        table = Table([[Paragraph("Sin respuestas completadas para este reactivo.", styles["note"])]], colWidths=[4.52 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), PDF_THEME["navy_soft"]),
                    ("BOX", (0, 0), (-1, -1), 0.4, PDF_THEME["line"]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        return table

    rows = []
    for option in response_options:
        label = Paragraph(
            f"<b>Nivel {option['value']}</b><br/>{escape(option['label'])}",
            styles["note"],
        )
        bar = _build_pdf_percent_bar(option["percent"], _option_color(option["value"]))
        value = Paragraph(
            f"<b>{option['count']}</b> respuestas<br/>{option['percent']}%",
            styles["small_center"],
        )
        rows.append([label, bar, value])

    table = Table(rows, colWidths=[2.08 * inch, 1.42 * inch, 0.92 * inch])
    table.setStyle(
        TableStyle(
            [
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, PDF_THEME["navy_soft"]]),
                ("GRID", (0, 0), (-1, -1), 0.35, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _build_pdf_profile_distribution_chart(question: dict) -> Drawing:
    drawing = Drawing(680, 118)
    drawing.add(String(2, 102, "Distribución global por opción", fontName="Helvetica-Bold", fontSize=9.5, fillColor=PDF_THEME["navy"]))

    options = question["response_options"]
    if not options or not any(option["count"] for option in options):
        drawing.add(String(340, 54, "Sin respuestas completadas para este reactivo", fontName="Helvetica", fontSize=9, fillColor=PDF_THEME["muted"], textAnchor="middle"))
        return drawing

    plot_x = 206
    plot_width = 360
    bar_height = 12
    start_y = 76
    gap = 20
    max_count = max(option["count"] for option in options) or 1
    for index, option in enumerate(options):
        y = start_y - index * gap
        drawing.add(String(4, y + 1, option["label"], fontName="Helvetica", fontSize=8, fillColor=PDF_THEME["muted"]))
        drawing.add(Rect(plot_x, y, plot_width, bar_height, fillColor=PDF_THEME["navy_soft"], strokeColor=PDF_THEME["line"], strokeWidth=0.3))
        if option["count"] > 0:
            color = _option_color(option["value"])
            drawing.add(Rect(plot_x, y, plot_width * option["count"] / max_count, bar_height, fillColor=color, strokeColor=color, strokeWidth=0))
        drawing.add(String(plot_x + plot_width + 8, y + 1, f"{option['count']} respuestas | {option['percent']}%", fontName="Helvetica-Bold", fontSize=7.5, fillColor=PDF_THEME["navy"]))

    return drawing


def _build_pdf_profile_distribution_table(question: dict, strata_order: list[str]) -> Table:
    header = ["Opción", "Global"]
    header.extend(strata_order)
    rows = [header]

    for option in question["response_options"]:
        values = [
            option["label"],
            f"{option['count']} | {option['percent']}%",
        ]
        for stratum in strata_order:
            detail = option["by_stratum"].get(stratum, {"count": 0, "percent": 0.0})
            values.append(f"{detail['count']} | {detail['percent']}%")
        rows.append(values)

    table = Table(rows, colWidths=[3.2 * inch, 1.0 * inch, *([1.0 * inch] * len(strata_order))], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["sage"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PDF_THEME["sage_soft"]]),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.3),
                ("LEADING", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.35, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _build_pdf_percent_bar(percent: float, color) -> Drawing:
    drawing = Drawing(110, 14)
    drawing.add(Rect(0, 2, 110, 10, fillColor=PDF_THEME["navy_soft"], strokeColor=PDF_THEME["line"], strokeWidth=0.4))
    if percent > 0:
        width = max(3, 110 * min(percent, 100) / 100)
        drawing.add(Rect(0, 2, width, 10, fillColor=color, strokeColor=color, strokeWidth=0))
    return drawing


def _build_stratum_summary_text(stratum: dict) -> str:
    strongest = stratum["strongest_dimension"] or "sin una dimensión fuerte identificada"
    attention = stratum["attention_dimension"] or "sin una dimensión crítica por falta de muestra"
    return (
        f"El estrato {escape(stratum['stratum'])} registra <b>{stratum['completed']}</b> encuestas completadas "
        f"({stratum['share_of_completed']}% del universo cerrado). "
        f"Su IIBP promedio es <b>{_value_or_dash(stratum['avg_iibp'])}</b> y su IVSP promedio es "
        f"<b>{_value_or_dash(stratum['avg_ivsp'])}</b>. "
        f"La dimensión con mejor lectura es <b>{escape(strongest)}</b> y la principal prioridad de intervención es "
        f"<b>{escape(attention)}</b>."
    )


def _draw_wellbeing_pdf_chrome(canvas, doc) -> None:
    canvas.saveState()
    width, height = doc.pagesize
    canvas.setFillColor(PDF_THEME["navy"])
    canvas.rect(doc.leftMargin, height - 0.32 * inch, width - doc.leftMargin - doc.rightMargin, 0.1 * inch, fill=1, stroke=0)
    canvas.setFillColor(PDF_THEME["muted"])
    canvas.setFont("Helvetica", 7.5)
    canvas.drawString(doc.leftMargin, 0.28 * inch, WELLBEING_PDF_FOOTER)
    canvas.drawRightString(width - doc.rightMargin, 0.28 * inch, f"Página {canvas.getPageNumber()}")
    canvas.restoreState()


def _value_or_dash(value) -> str:
    return "-" if value is None else str(value)


def _metric_cell_text(metric: dict) -> str:
    if not metric["count"] or metric["percent"] is None:
        return "—"
    return f"{metric['percent']}% ({metric['count']})"


def _chunked(items: list, size: int) -> list[list]:
    return [items[index:index + size] for index in range(0, len(items), size)]


def _option_color(value: int):
    color_map = {
        4: PDF_THEME["optimal"],
        3: PDF_THEME["high"],
        2: PDF_THEME["medium"],
        1: PDF_THEME["low"],
    }
    return color_map.get(value, PDF_THEME["sage"])


def _score_color(percent: float):
    if percent >= 75:
        return PDF_THEME["optimal"]
    if percent >= 60:
        return PDF_THEME["high"]
    if percent >= 45:
        return PDF_THEME["medium"]
    return PDF_THEME["low"]


def _write_title(sheet, title: str, subtitle: str, end_column: int) -> None:
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    sheet["A1"].font = Font(bold=True, size=16, color=EXCEL_THEME["navy"])
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.append([subtitle])
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    sheet["A2"].font = Font(size=11, color=EXCEL_THEME["muted"])
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")


def _style_table(sheet, *, header_row: int) -> None:
    header_fill = PatternFill("solid", fgColor=EXCEL_THEME["navy"])
    zebra_fill = PatternFill("solid", fgColor=EXCEL_THEME["navy_soft"])
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color=EXCEL_THEME["white"])
        cell.fill = header_fill
        cell.alignment = CENTER_ALIGNMENT
    for row_index in range(header_row + 1, sheet.max_row + 1):
        if row_index % 2 == 0:
            for cell in sheet[row_index]:
                cell.fill = zebra_fill


def _set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

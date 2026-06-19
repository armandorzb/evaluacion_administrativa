from __future__ import annotations

import math
import os
from html import escape
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.graphics.shapes import Drawing, Line, Polygon, Rect, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import CondPageBreak, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from municipal_diagnostico.services.iso9001 import format_iso_datetime, summarize_iso9001_evaluation
from municipal_diagnostico.timeutils import utcnow


EXCEL_COLORS = {
    "navy": "B65F1E",
    "soft": "F6E7DA",
    "line": "D8D3CC",
    "white": "FFFFFF",
}

PDF_THEME = {
    "navy": colors.HexColor("#B65F1E"),
    "navy_soft": colors.HexColor("#F6E7DA"),
    "sage": colors.HexColor("#6E7175"),
    "sage_soft": colors.HexColor("#F0F0F0"),
    "line": colors.HexColor("#D8D3CC"),
    "muted": colors.HexColor("#56585C"),
    "low": colors.HexColor("#C62828"),
    "medium": colors.HexColor("#F9A825"),
    "high": colors.HexColor("#F9A825"),
    "optimal": colors.HexColor("#2E7D32"),
    "empty": colors.HexColor("#9A9A9A"),
    "orange_dark": colors.HexColor("#8F4718"),
    "orange_light": colors.HexColor("#FFF3E8"),
    "gray_dark": colors.HexColor("#3F4246"),
    "traffic_red": colors.HexColor("#C62828"),
    "traffic_yellow": colors.HexColor("#F9A825"),
    "traffic_green": colors.HexColor("#2E7D32"),
    "traffic_red_soft": colors.HexColor("#FCE4E4"),
    "traffic_yellow_soft": colors.HexColor("#FFF4CC"),
    "traffic_green_soft": colors.HexColor("#E6F2E7"),
}

ISO_RESPONSE_ORDER = [
    ("no", "No", PDF_THEME["traffic_red"]),
    ("parcial", "Parcial", PDF_THEME["traffic_yellow"]),
    ("si", "Sí", PDF_THEME["traffic_green"]),
    ("na", "N/A", PDF_THEME["muted"]),
]

ISO_MATURITY_GUIDE = [
    ("0%", "Nivel 0 - No iniciado", "low"),
    ("1-20%", "Nivel 1 - Inicial", "low"),
    ("21-40%", "Nivel 2 - En desarrollo", "low"),
    ("41-60%", "Nivel 3 - Definido", "medium"),
    ("61-80%", "Nivel 4 - Gestionado", "medium"),
    ("81-100%", "Nivel 5 - Optimizado", "optimal"),
]

ISO_CLAUSE_GUIDANCE = {
    "4": {
        "focus": "contexto, partes interesadas, alcance y procesos del sistema de gestión de la calidad",
        "evidence": "análisis de contexto, requisitos de partes interesadas, alcance documentado y mapa de procesos",
    },
    "5": {
        "focus": "liderazgo, política de calidad, enfoque al cliente, roles y responsabilidades",
        "evidence": "política vigente, objetivos comunicados, responsabilidades asignadas y evidencia de participación directiva",
    },
    "6": {
        "focus": "riesgos, oportunidades, objetivos de calidad y planeación de cambios",
        "evidence": "matriz de riesgos y oportunidades, objetivos medibles, responsables, indicadores y planes de cambio",
    },
    "7": {
        "focus": "recursos, competencia, toma de conciencia, comunicación e información documentada",
        "evidence": "planes de capacitación, perfiles, registros de competencia, comunicaciones y control documental",
    },
    "8": {
        "focus": "planeación y control operacional, requisitos del servicio, proveedores y salidas no conformes",
        "evidence": "procedimientos operativos, controles de proveedores, validaciones, liberaciones y tratamiento de no conformidades",
    },
    "9": {
        "focus": "seguimiento, medición, satisfacción del usuario, auditorías internas y revisión por la dirección",
        "evidence": "indicadores, resultados de satisfacción, programa de auditorías, hallazgos y actas de revisión directiva",
    },
    "10": {
        "focus": "no conformidades, acciones correctivas y mejora continua",
        "evidence": "acciones correctivas, análisis de causa, verificación de eficacia y cartera de mejoras",
    },
}

ISO_CLAUSE_SUPPORT = {
    "4": {
        "title": "Contexto de la organización",
        "support": "Comprender factores internos y externos, partes interesadas, alcance del sistema y procesos necesarios para el SGC.",
        "evidence": "Análisis de contexto, alcance documentado, requisitos de partes interesadas y mapa o caracterización de procesos.",
    },
    "5": {
        "title": "Liderazgo",
        "support": "Demostrar compromiso directivo, política de calidad, enfoque al usuario/cliente y responsabilidades claras.",
        "evidence": "Política comunicada, roles asignados, evidencias de participación directiva y mecanismos de enfoque al usuario.",
    },
    "6": {
        "title": "Planificación",
        "support": "Planear acciones para riesgos y oportunidades, objetivos de calidad medibles y cambios controlados.",
        "evidence": "Matriz de riesgos, objetivos con indicadores, responsables, metas, planes de acción y control de cambios.",
    },
    "7": {
        "title": "Apoyo",
        "support": "Asegurar recursos, competencia, toma de conciencia, comunicación e información documentada controlada.",
        "evidence": "Perfiles, capacitación, registros de competencia, comunicaciones y control de documentos/registros.",
    },
    "8": {
        "title": "Operación",
        "support": "Planear y controlar la prestación del servicio, requisitos, proveedores externos, liberación y salidas no conformes.",
        "evidence": "Procedimientos operativos, controles de proveedores, validaciones, liberaciones y tratamiento de no conformidades.",
    },
    "9": {
        "title": "Evaluación del desempeño",
        "support": "Medir desempeño, satisfacción, auditorías internas y revisión por la dirección para sostener la eficacia del SGC.",
        "evidence": "Indicadores, encuestas o mediciones, programa de auditoría, hallazgos, acciones y actas de revisión directiva.",
    },
    "10": {
        "title": "Mejora",
        "support": "Atender no conformidades, acciones correctivas y mejora continua con verificación de eficacia.",
        "evidence": "Análisis de causa, acciones correctivas, seguimiento de eficacia, cartera de mejoras y lecciones aprendidas.",
    },
}

ISO_PDF_FOOTER = "Dirección de Recursos Humanos - Ayuntamiento de Hermosillo"


def _register_pdf_fonts() -> tuple[str, str]:
    package_root = Path(__file__).resolve().parents[1]
    regular_env = os.environ.get("ISO9001_PDF_FONT_REGULAR")
    bold_env = os.environ.get("ISO9001_PDF_FONT_BOLD")
    regular_candidates = [
        Path(regular_env) if regular_env else None,
        package_root / "static" / "fonts" / "GoogleSans-Regular.ttf",
        package_root / "static" / "fonts" / "ProductSans-Regular.ttf",
        Path("C:/Windows/Fonts/GoogleSans-Regular.ttf"),
        Path("C:/Windows/Fonts/ProductSans-Regular.ttf"),
        Path("/usr/share/fonts/truetype/google-sans/GoogleSans-Regular.ttf"),
        Path("/usr/share/fonts/truetype/product-sans/ProductSans-Regular.ttf"),
    ]
    bold_candidates = [
        Path(bold_env) if bold_env else None,
        package_root / "static" / "fonts" / "GoogleSans-Bold.ttf",
        package_root / "static" / "fonts" / "ProductSans-Bold.ttf",
        Path("C:/Windows/Fonts/GoogleSans-Bold.ttf"),
        Path("C:/Windows/Fonts/ProductSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/google-sans/GoogleSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/product-sans/ProductSans-Bold.ttf"),
    ]
    regular_path = next((path for path in regular_candidates if path and path.exists()), None)
    bold_path = next((path for path in bold_candidates if path and path.exists()), regular_path)
    if regular_path:
        try:
            pdfmetrics.registerFont(TTFont("GoogleSans", str(regular_path)))
            pdfmetrics.registerFont(TTFont("GoogleSans-Bold", str(bold_path or regular_path)))
            return "GoogleSans", "GoogleSans-Bold"
        except Exception:
            pass
    return "Helvetica", "Helvetica-Bold"


PDF_FONT_REGULAR, PDF_FONT_BOLD = _register_pdf_fonts()


def build_iso9001_excel(evaluation) -> BytesIO:
    summary = summarize_iso9001_evaluation(evaluation)
    workbook = Workbook()

    resumen = workbook.active
    resumen.title = "Resumen"
    _write_title(resumen, "Diagnóstico ISO 9001:2015", evaluation.dependencia.nombre, end_column=7)
    resumen.append([])
    resumen.append(["Estado", summary["state_label"]])
    resumen.append(["Avance", summary["completion"]])
    resumen.append(["Cumplimiento", summary["percent"] if summary["percent"] is not None else "Sin aplicables"])
    resumen.append(["Madurez", summary["maturity_label"]])
    resumen.append([])
    resumen.append(["Cláusula", "Nombre", "Reactivos", "Respondidos", "Aplicables", "Pts.", "% Cumpl.", "Madurez"])
    header_row = resumen.max_row
    for clause in summary["clauses"]:
        resumen.append(
            [
                clause["numero"],
                clause["nombre"],
                clause["total"],
                clause["answered"],
                clause["applicable"],
                clause["points"],
                clause["percent"] if clause["percent"] is not None else "",
                clause["maturity_label"],
            ]
        )
    _style_table(resumen, header_row)
    _set_widths(resumen, {"A": 14, "B": 34, "C": 12, "D": 14, "E": 12, "F": 10, "G": 12, "H": 28})

    detalle = workbook.create_sheet("Cuestionario")
    _write_title(detalle, "Detalle de respuestas ISO 9001:2015", evaluation.ciclo.nombre, end_column=10)
    detalle.append([])
    detalle.append(
        [
            "Cláusula",
            "Apartado",
            "N",
            "Reactivo",
            "Calificación",
            "Pts.",
            "Observación / hallazgo",
            "Evidencia sugerida",
            "Criterio de idoneidad",
            "Archivos",
        ]
    )
    detail_header = detalle.max_row
    for clause in summary["clauses"]:
        for section in clause["sections"]:
            for row in section["questions"]:
                reactive = row["reactivo"]
                detalle.append(
                    [
                        clause["numero"],
                        section["codigo"],
                        reactive.numero,
                        reactive.texto,
                        row["selected_label"],
                        "" if row["is_na"] or not row["answered"] else row["points"],
                        row["observacion"],
                        reactive.evidencia_sugerida,
                        reactive.criterio_idoneidad,
                        len(row["evidence"]),
                    ]
                )
    _style_table(detalle, detail_header)
    _set_widths(detalle, {"A": 10, "B": 12, "C": 8, "D": 50, "E": 15, "F": 8, "G": 36, "H": 36, "I": 44, "J": 10})

    guia = workbook.create_sheet("Guia")
    _write_title(guia, "Guia de madurez ISO 9001:2015", evaluation.dependencia.nombre, end_column=5)
    guia.append([])
    guia.append(["Lectura global", _global_maturity_comment(summary)])
    guia.append([])
    guia.append(["Clausula", "Madurez", "% Cumpl.", "Comentario guia", "Siguiente paso"])
    guide_header = guia.max_row
    for clause in summary["clauses"]:
        guide_row = _clause_guidance_row(clause)
        guia.append(
            [
                guide_row["clause"],
                guide_row["maturity"],
                clause["percent"] if clause["percent"] is not None else "",
                guide_row["reading"],
                guide_row["next_step"],
            ]
        )
    guia.append([])
    guia.append(["Apartado prioritario", "Clausula", "% Cumpl.", "Comentario guia"])
    priority_header = guia.max_row
    for section in _priority_sections(summary):
        guia.append(
            [
                f"{section['codigo']} {section['nombre']}",
                section["clausula"],
                section["percent"] if section["percent"] is not None else "",
                _section_maturity_comment(section),
            ]
        )
    _style_table(guia, guide_header)
    _style_table(guia, priority_header)
    _set_widths(guia, {"A": 28, "B": 26, "C": 12, "D": 62, "E": 62})

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A4"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_iso9001_pdf(evaluation) -> BytesIO:
    summary = summarize_iso9001_evaluation(evaluation)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.55 * inch,
        rightMargin=0.55 * inch,
        topMargin=0.62 * inch,
        bottomMargin=0.55 * inch,
    )

    styles = _build_iso_pdf_styles()
    story: list = []
    story.extend(_build_iso_cover_story(evaluation, summary, styles))
    story.extend(_build_report_index_story(evaluation, summary, styles))

    story.append(Paragraph("Resumen ejecutivo", styles["section"]))
    if not summary["is_final"]:
        story.append(
            Paragraph(
                "Resultado preliminar. El autodiagnóstico puede cambiar mientras la captura o revisión siga abierta.",
                styles["warning"],
            )
        )
        story.append(Spacer(1, 0.08 * inch))
    story.append(_build_iso_metric_cards(summary, styles))
    story.append(Spacer(1, 0.14 * inch))
    story.append(_build_methodology_panel(styles))
    story.append(Spacer(1, 0.16 * inch))

    story.append(Paragraph("Gráficas de autodiagnóstico", styles["section"]))
    story.append(_build_chart_pair(summary))
    story.append(Spacer(1, 0.12 * inch))
    story.append(_build_progress_comparison_chart(summary))
    story.append(Spacer(1, 0.12 * inch))
    story.append(_build_iso_radar_chart(summary))
    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph(f"<b>Lectura guía:</b> {_paragraph_escape(_global_maturity_comment(summary))}", styles["body"]))
    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph("Soporte normativo de interpretación", styles["subsection"]))
    story.append(
        Paragraph(
            "Contenido de apoyo en paráfrasis. Para auditoría formal debe consultarse la norma ISO 9001:2015 oficial.",
            styles["note"],
        )
    )
    story.append(Spacer(1, 0.06 * inch))
    story.append(_build_normative_support_table(summary, styles))

    story.append(CondPageBreak(3.2 * inch))
    story.append(Paragraph("Diagnóstico accionable", styles["section"]))
    story.append(
        Paragraph(
            "La lectura prioriza brechas por cumplimiento, avance de captura y soporte documental disponible.",
            styles["body"],
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_build_clause_heatmap_table(summary, styles))
    story.append(Spacer(1, 0.14 * inch))
    story.append(Paragraph("Subapartados de atención prioritaria", styles["subsection"]))
    story.append(_build_priority_sections_table(_priority_sections(summary), styles))
    story.append(Spacer(1, 0.14 * inch))
    story.append(Paragraph("Cobertura de evidencia por cláusula", styles["subsection"]))
    story.append(_build_evidence_coverage_table(summary, styles))

    story.append(CondPageBreak(3.2 * inch))
    story.append(Paragraph("Guía de madurez por cláusula", styles["section"]))
    story.append(
        Paragraph(
            "Los comentarios son orientativos: traducen la calificación del autodiagnóstico en señales de madurez y próximos pasos para fortalecer el SGC.",
            styles["body"],
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_build_clause_guidance_table(summary, styles))

    for clause in summary["clauses"]:
        story.append(CondPageBreak(3.7 * inch))
        story.extend(_build_clause_story(clause, styles))

    story.append(CondPageBreak(3.2 * inch))
    story.append(Paragraph("Anexo consultable de hallazgos y evidencias", styles["section"]))
    story.append(
        Paragraph(
            "Se listan reactivos con respuesta No, Parcial o N/A, así como reactivos con observación o evidencia cargada.",
            styles["body"],
        )
    )
    story.append(Spacer(1, 0.08 * inch))
    story.append(_build_findings_appendix(summary, styles))

    doc.build(story, onFirstPage=_draw_iso_pdf_chrome, onLaterPages=_draw_iso_pdf_chrome)
    buffer.seek(0)
    return buffer


def _build_iso_pdf_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "cover_kicker": ParagraphStyle(
            "IsoCoverKicker",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=9.8,
            leading=12,
            textColor=PDF_THEME["sage"],
            alignment=TA_CENTER,
            spaceAfter=8,
        ),
        "cover_title": ParagraphStyle(
            "IsoCoverTitle",
            parent=base["Title"],
            fontName=PDF_FONT_BOLD,
            fontSize=31,
            leading=35,
            textColor=PDF_THEME["navy"],
            alignment=TA_CENTER,
            spaceAfter=10,
        ),
        "cover_subject": ParagraphStyle(
            "IsoCoverSubject",
            parent=base["Heading2"],
            fontName=PDF_FONT_BOLD,
            fontSize=17,
            leading=21,
            textColor=PDF_THEME["navy"],
            alignment=TA_CENTER,
            spaceAfter=16,
        ),
        "cover_org": ParagraphStyle(
            "IsoCoverOrg",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=12,
            leading=15,
            alignment=TA_CENTER,
            textColor=PDF_THEME["gray_dark"],
            spaceAfter=2,
        ),
        "cover_department": ParagraphStyle(
            "IsoCoverDepartment",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=9.5,
            leading=12,
            alignment=TA_CENTER,
            textColor=PDF_THEME["navy"],
            spaceAfter=8,
        ),
        "cover_note": ParagraphStyle(
            "IsoCoverNote",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=10,
            leading=13,
            alignment=TA_CENTER,
            textColor=PDF_THEME["muted"],
        ),
        "cover_meta": ParagraphStyle(
            "IsoCoverMeta",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            textColor=PDF_THEME["muted"],
        ),
        "index_lead": ParagraphStyle(
            "IsoIndexLead",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=9.2,
            leading=12.2,
            textColor=PDF_THEME["muted"],
            spaceAfter=6,
        ),
        "section": ParagraphStyle(
            "IsoSection",
            parent=base["Heading2"],
            fontName=PDF_FONT_BOLD,
            fontSize=14,
            leading=18,
            textColor=PDF_THEME["navy"],
            spaceBefore=4,
            spaceAfter=6,
        ),
        "subsection": ParagraphStyle(
            "IsoSubsection",
            parent=base["Heading3"],
            fontName=PDF_FONT_BOLD,
            fontSize=11.2,
            leading=14,
            textColor=PDF_THEME["navy"],
            spaceBefore=2,
            spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "IsoBody",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=9,
            leading=12,
            textColor=PDF_THEME["muted"],
        ),
        "note": ParagraphStyle(
            "IsoNote",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=7.6,
            leading=9.2,
            textColor=PDF_THEME["muted"],
        ),
        "small": ParagraphStyle(
            "IsoSmall",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=7.8,
            leading=9.5,
            textColor=PDF_THEME["gray_dark"],
        ),
        "small_center": ParagraphStyle(
            "IsoSmallCenter",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=7.8,
            leading=9.5,
            alignment=TA_CENTER,
            textColor=PDF_THEME["gray_dark"],
        ),
        "traffic_label_light": ParagraphStyle(
            "IsoTrafficLabelLight",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=7.8,
            leading=9.5,
            alignment=TA_CENTER,
            textColor=colors.white,
        ),
        "traffic_label_dark": ParagraphStyle(
            "IsoTrafficLabelDark",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=7.8,
            leading=9.5,
            alignment=TA_CENTER,
            textColor=PDF_THEME["gray_dark"],
        ),
        "table_header": ParagraphStyle(
            "IsoTableHeader",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=7.5,
            leading=9,
            alignment=TA_CENTER,
            textColor=colors.white,
        ),
        "metric_label": ParagraphStyle(
            "IsoMetricLabel",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=7.3,
            leading=9,
            alignment=TA_CENTER,
            textColor=PDF_THEME["muted"],
        ),
        "metric_value": ParagraphStyle(
            "IsoMetricValue",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=18,
            leading=20,
            alignment=TA_CENTER,
            textColor=PDF_THEME["navy"],
        ),
        "metric_caption": ParagraphStyle(
            "IsoMetricCaption",
            parent=base["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=7.1,
            leading=8.4,
            alignment=TA_CENTER,
            textColor=PDF_THEME["muted"],
        ),
        "warning": ParagraphStyle(
            "IsoWarning",
            parent=base["BodyText"],
            fontName=PDF_FONT_BOLD,
            fontSize=8.5,
            leading=11,
            textColor=PDF_THEME["medium"],
            backColor=colors.HexColor("#F8E0C5"),
            borderPadding=7,
        ),
    }


def _build_iso_cover_story(evaluation, summary: dict, styles: dict[str, ParagraphStyle]) -> list:
    status = "OFICIAL" if summary["is_final"] else "PRELIMINAR"
    status_color = PDF_THEME["gray_dark"] if summary["is_final"] else PDF_THEME["medium"]
    institutional_header = Table(
        [
            [Paragraph("H. Ayuntamiento de Hermosillo", styles["cover_org"])],
            [Paragraph("Dirección de Recursos Humanos", styles["cover_department"])],
        ],
        colWidths=[7.05 * inch],
        rowHeights=[0.28 * inch, 0.24 * inch],
    )
    institutional_header.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), PDF_THEME["navy_soft"]),
                ("LINEBELOW", (0, 1), (-1, 1), 1.6, PDF_THEME["navy"]),
                ("BOX", (0, 0), (-1, -1), 0.45, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    status_badge = Table(
        [[Paragraph(status, styles["table_header"])]],
        colWidths=[1.55 * inch],
        rowHeights=[0.32 * inch],
    )
    status_badge.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), status_color),
                ("BOX", (0, 0), (-1, -1), 0.4, status_color),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )

    story = [
        Spacer(1, 0.42 * inch),
        institutional_header,
        Spacer(1, 1.25 * inch),
        Paragraph("Sistema de gestión de la calidad", styles["cover_kicker"]),
        Paragraph("Autodiagnóstico ISO 9001:2015", styles["cover_title"]),
        Paragraph(_paragraph_escape(evaluation.dependencia.nombre), styles["cover_subject"]),
        status_badge,
        Spacer(1, 0.42 * inch),
        Paragraph(
            "Informe ejecutivo de implementación, evidencia documental y madurez institucional por cláusula auditable.",
            styles["cover_note"],
        ),
        Spacer(1, 0.22 * inch),
        Paragraph(
            f"Ciclo: {_paragraph_escape(evaluation.ciclo.nombre)} | Estado: {_paragraph_escape(summary['state_label'])}",
            styles["cover_meta"],
        ),
        Spacer(1, 2.25 * inch),
        PageBreak(),
    ]
    return story


def _build_report_index_story(evaluation, summary: dict, styles: dict[str, ParagraphStyle]) -> list:
    sections = [
        ("1", "Resumen ejecutivo", "Indicadores clave, escala de cálculo y lectura global del autodiagnóstico."),
        ("2", "Gráficas de autodiagnóstico", "Barras por cláusula, distribución de respuestas, avance comparado y araña de madurez."),
        ("3", "Diagnóstico accionable", "Matriz de madurez, subapartados prioritarios y cobertura de evidencia."),
        ("4", "Guía de madurez por cláusula", "Interpretación por requisito auditable y siguiente paso sugerido."),
        ("5", "Detalle por cláusula", "Resumen de cada cláusula ISO 9001:2015 y sus subapartados evaluados."),
        ("6", "Anexo consultable", "Reactivos con brecha, N/A, observación o evidencia para revisión posterior."),
    ]
    rows = [["Orden", "Sección", "Qué consultar"]]
    rows.extend(sections)
    story = [
        Paragraph("Índice del reporte", styles["section"]),
        Paragraph(
            "Este documento se organiza como un autodiagnóstico consultable para revisión ejecutiva, "
            "seguimiento operativo y trazabilidad de hallazgos.",
            styles["index_lead"],
        ),
        _build_maturity_index_panel(evaluation, summary, styles),
        Spacer(1, 0.16 * inch),
        _pdf_table(
            rows,
            styles,
            col_widths=[0.62 * inch, 1.8 * inch, 4.65 * inch],
        ),
        Spacer(1, 0.16 * inch),
        Paragraph(
            "Índice de madurez ISO: el porcentaje global resume los puntos obtenidos contra el máximo posible "
            "de los reactivos aplicables. Los N/A se documentan y se excluyen del denominador.",
            styles["note"],
        ),
        PageBreak(),
    ]
    return story


def _build_maturity_index_panel(evaluation, summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    maturity = _format_percent(summary["percent"])
    completion = _format_percent(summary["completion"])
    responsible = evaluation.responsable.nombre if evaluation.responsable else "Sin responsable"
    reviewer = evaluation.revisor.nombre if evaluation.revisor else "Sin revisor"
    generated = format_iso_datetime(utcnow())
    cycle_dates = f"{_date_or_dash(evaluation.ciclo.fecha_inicio)} al {_date_or_dash(evaluation.ciclo.fecha_cierre)}"
    rows = [
        [
            Paragraph("Dependencia", styles["table_header"]),
            Paragraph("Ciclo", styles["table_header"]),
            Paragraph("Índice de madurez", styles["table_header"]),
            Paragraph("Estado", styles["table_header"]),
        ],
        [
            Paragraph(_paragraph_escape(evaluation.dependencia.nombre), styles["small"]),
            Paragraph(_paragraph_escape(evaluation.ciclo.nombre), styles["small"]),
            Paragraph(_paragraph_escape(f"{maturity} - {summary['maturity_label']}"), styles["small_center"]),
            Paragraph(_paragraph_escape(summary["state_label"]), styles["small_center"]),
        ],
        [
            Paragraph("Periodo", styles["table_header"]),
            Paragraph("Responsable", styles["table_header"]),
            Paragraph("Revisor", styles["table_header"]),
            Paragraph("Descarga", styles["table_header"]),
        ],
        [
            Paragraph(_paragraph_escape(cycle_dates), styles["small_center"]),
            Paragraph(_paragraph_escape(responsible), styles["small"]),
            Paragraph(_paragraph_escape(reviewer), styles["small"]),
            Paragraph(_paragraph_escape(generated), styles["small_center"]),
        ],
        [
            Paragraph("Captura", styles["table_header"]),
            Paragraph("Versión", styles["table_header"]),
            Paragraph("Reactivos aplicables", styles["table_header"]),
            Paragraph("Evidencias", styles["table_header"]),
        ],
        [
            Paragraph(_paragraph_escape(completion), styles["small_center"]),
            Paragraph(_paragraph_escape(evaluation.ciclo.version.nombre), styles["small"]),
            Paragraph(_paragraph_escape(summary["applicable_questions"]), styles["small_center"]),
            Paragraph(_paragraph_escape(summary["evidence_count"]), styles["small_center"]),
        ],
    ]
    table = Table(rows, colWidths=[2.15 * inch, 2.05 * inch, 1.65 * inch, 1.2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("BACKGROUND", (0, 2), (-1, 2), PDF_THEME["navy"]),
                ("BACKGROUND", (0, 4), (-1, 4), PDF_THEME["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("TEXTCOLOR", (0, 2), (-1, 2), colors.white),
                ("TEXTCOLOR", (0, 4), (-1, 4), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), PDF_THEME["orange_light"]),
                ("BACKGROUND", (0, 3), (-1, 3), colors.white),
                ("BACKGROUND", (0, 5), (-1, 5), PDF_THEME["orange_light"]),
                ("GRID", (0, 0), (-1, -1), 0.45, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table

def _build_iso_metric_cards(summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    maturity_level = _maturity_level_prefix(summary["maturity_label"])
    cards = [
        ("Avance", _format_percent(summary["completion"]), f"{summary['answered_questions']} de {summary['total_questions']} reactivos"),
        ("Cumplimiento", _format_percent(summary["percent"]), "Puntos sobre aplicables"),
        ("Madurez", _compact_maturity_label(summary["maturity_label"]), f"{maturity_level} ISO" if maturity_level else "Nivel global ISO"),
        ("Evidencias", str(summary["evidence_count"]), "Archivos activos"),
        ("Aplicables", str(summary["applicable_questions"]), "Base del denominador"),
        ("N/A", str(summary["na_questions"]), "Excluidos del cálculo"),
        ("Puntos", str(summary["points"]), "Escala 0, 1, 2"),
        ("Estado", summary["state_label"], "Flujo formal"),
    ]
    rows = []
    for start in range(0, len(cards), 4):
        cells = []
        for label, value, caption in cards[start:start + 4]:
            card = Table(
                [
                    [Paragraph(_paragraph_escape(label.upper()), styles["metric_label"])],
                    [Paragraph(_paragraph_escape(value), styles["metric_value"])],
                    [Paragraph(_paragraph_escape(caption), styles["metric_caption"])],
                ],
                colWidths=[1.55 * inch],
                rowHeights=[0.22 * inch, 0.43 * inch, 0.27 * inch],
            )
            card.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 2),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            cells.append(card)
        rows.append(cells)

    table = Table(rows, colWidths=[1.76 * inch] * 4, rowHeights=[0.98 * inch] * len(rows))
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, PDF_THEME["navy_soft"]]),
                ("BOX", (0, 0), (-1, -1), 0.55, PDF_THEME["line"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.55, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
            ]
        )
    )
    return table


def _build_methodology_panel(styles: dict[str, ParagraphStyle]) -> Table:
    method = Paragraph(
        "Método de cálculo: No=0, Parcial=1, Sí=2. N/A cuenta como reactivo respondido para avance de captura, "
        "pero se excluye del denominador de cumplimiento. Cumplimiento = puntos obtenidos / (reactivos aplicables x 2).",
        styles["body"],
    )
    layout = Table(
        [[method, _build_maturity_legend_table(styles)]],
        colWidths=[4.18 * inch, 2.85 * inch],
    )
    layout.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.55, PDF_THEME["line"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.45, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return layout


def _build_maturity_legend_table(styles: dict[str, ParagraphStyle]) -> Table:
    rows = [[Paragraph("Rango", styles["table_header"]), Paragraph("Nivel", styles["table_header"])]]
    for range_label, label, slug in ISO_MATURITY_GUIDE:
        range_style = _traffic_label_style_for_slug(slug, styles)
        rows.append(
            [
                Paragraph(_paragraph_escape(range_label), range_style),
                Paragraph(_paragraph_escape(label), styles["small"]),
            ]
        )
    table = Table(rows, colWidths=[0.65 * inch, 1.85 * inch], repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
        ("GRID", (0, 0), (-1, -1), 0.4, PDF_THEME["line"]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for row_index, (_range, _label, slug) in enumerate(ISO_MATURITY_GUIDE, start=1):
        style_commands.append(("BACKGROUND", (0, row_index), (0, row_index), PDF_THEME[slug]))
        style_commands.append(("TEXTCOLOR", (0, row_index), (0, row_index), colors.white))
    table.setStyle(TableStyle(style_commands))
    return table


def _build_chart_pair(summary: dict) -> Table:
    table = Table(
        [[_build_clause_bar_chart(summary), _build_response_distribution_chart(summary)]],
        colWidths=[4.65 * inch, 2.6 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return table


def _build_clause_bar_chart(summary: dict) -> Drawing:
    drawing = Drawing(330, 205)
    drawing.add(String(165, 188, "Cumplimiento por cláusula", fontName=PDF_FONT_BOLD, fontSize=10, fillColor=PDF_THEME["navy"], textAnchor="middle"))
    plot_x = 36
    plot_y = 36
    plot_width = 268
    plot_height = 122
    drawing.add(Rect(plot_x, plot_y, plot_width, plot_height, fillColor=colors.white, strokeColor=PDF_THEME["line"], strokeWidth=0.8))

    for tick in range(5):
        value = 25 * tick
        y = plot_y + (plot_height * value / 100)
        drawing.add(Line(plot_x, y, plot_x + plot_width, y, strokeColor=PDF_THEME["line"], strokeWidth=0.45))
        drawing.add(String(plot_x - 7, y - 2, str(value), fontName=PDF_FONT_REGULAR, fontSize=6.5, fillColor=PDF_THEME["muted"], textAnchor="end"))

    clauses = summary["clauses"]
    slot = plot_width / max(len(clauses), 1)
    bar_width = min(20, slot * 0.55)
    for index, clause in enumerate(clauses):
        percent = clause["percent"] or 0
        x = plot_x + slot * index + (slot - bar_width) / 2
        height = plot_height * min(percent, 100) / 100
        color = _traffic_color_for_percent(clause["percent"])
        drawing.add(Rect(x, plot_y, bar_width, height, fillColor=color, strokeColor=color, strokeWidth=0))
        drawing.add(String(x + bar_width / 2, plot_y + height + 5, _format_percent(clause["percent"]), fontName=PDF_FONT_BOLD, fontSize=6.4, fillColor=PDF_THEME["navy"], textAnchor="middle"))
        drawing.add(String(x + bar_width / 2, plot_y - 13, clause["numero"], fontName=PDF_FONT_BOLD, fontSize=7.2, fillColor=PDF_THEME["muted"], textAnchor="middle"))
    drawing.add(String(plot_x + plot_width / 2, 10, "Cláusulas auditables 4 a 10", fontName=PDF_FONT_REGULAR, fontSize=7.2, fillColor=PDF_THEME["muted"], textAnchor="middle"))
    return drawing


def _build_response_distribution_chart(summary: dict) -> Drawing:
    rows = _response_distribution(summary)
    total = sum(row["count"] for row in rows)
    drawing = Drawing(185, 205)
    drawing.add(String(92, 188, "Distribución de respuestas", fontName=PDF_FONT_BOLD, fontSize=10, fillColor=PDF_THEME["navy"], textAnchor="middle"))
    bar_x = 14
    bar_y = 150
    bar_width = 156
    bar_height = 16
    drawing.add(Rect(bar_x, bar_y, bar_width, bar_height, fillColor=PDF_THEME["navy_soft"], strokeColor=PDF_THEME["line"], strokeWidth=0.7))
    if total <= 0:
        drawing.add(String(92, 112, "Sin respuestas registradas", fontName=PDF_FONT_REGULAR, fontSize=8.5, fillColor=PDF_THEME["muted"], textAnchor="middle"))
        return drawing

    offset = bar_x
    for row in rows:
        width = bar_width * row["count"] / total
        if width > 0:
            drawing.add(Rect(offset, bar_y, width, bar_height, fillColor=row["color"], strokeColor=row["color"], strokeWidth=0))
        offset += width

    start_y = 121
    for index, row in enumerate(rows):
        y = start_y - index * 23
        drawing.add(Rect(14, y, 9, 9, fillColor=row["color"], strokeColor=row["color"]))
        drawing.add(String(30, y + 1, row["label"], fontName=PDF_FONT_BOLD, fontSize=7.4, fillColor=PDF_THEME["navy"]))
        drawing.add(String(171, y + 1, f"{row['count']} | {_format_percent(row['percent'])}", fontName=PDF_FONT_REGULAR, fontSize=7.2, fillColor=PDF_THEME["muted"], textAnchor="end"))
    return drawing


def _build_progress_comparison_chart(summary: dict) -> Drawing:
    drawing = Drawing(520, 72)
    drawing.add(String(2, 58, "Avance de captura vs cumplimiento", fontName=PDF_FONT_BOLD, fontSize=9.5, fillColor=PDF_THEME["navy"]))
    rows = [
        ("Avance captura", summary["completion"], _traffic_color_for_percent(summary["completion"])),
        ("Cumplimiento ISO", summary["percent"] or 0, _traffic_color_for_percent(summary["percent"])),
    ]
    for index, (label, value, color) in enumerate(rows):
        y = 35 - index * 23
        drawing.add(String(4, y + 2, label, fontName=PDF_FONT_REGULAR, fontSize=7.8, fillColor=PDF_THEME["muted"]))
        drawing.add(Rect(120, y, 330, 11, fillColor=PDF_THEME["navy_soft"], strokeColor=PDF_THEME["line"], strokeWidth=0.4))
        if value > 0:
            drawing.add(Rect(120, y, 330 * min(value, 100) / 100, 11, fillColor=color, strokeColor=color, strokeWidth=0))
        drawing.add(String(463, y + 2, _format_percent(value if label == "Avance captura" else summary["percent"]), fontName=PDF_FONT_BOLD, fontSize=7.6, fillColor=PDF_THEME["navy"]))
    return drawing


def _build_iso_radar_chart(summary: dict) -> Drawing:
    clauses = summary["clauses"]
    drawing = Drawing(520, 300)
    center_x = 260
    center_y = 130
    radius = 92
    label_radius = radius + 26
    drawing.add(String(260, 288, "Araña de madurez por cláusula", fontName=PDF_FONT_BOLD, fontSize=11, fillColor=PDF_THEME["navy"], textAnchor="middle"))
    drawing.add(String(260, 272, "Índice comparativo de cumplimiento 0-100; N/A excluido del denominador", fontName=PDF_FONT_REGULAR, fontSize=7.8, fillColor=PDF_THEME["muted"], textAnchor="middle"))

    if not clauses:
        drawing.add(String(center_x, center_y, "Sin cláusulas disponibles", fontName=PDF_FONT_REGULAR, fontSize=8.5, fillColor=PDF_THEME["muted"], textAnchor="middle"))
        return drawing

    angles = [math.pi / 2 - (2 * math.pi * index / len(clauses)) for index in range(len(clauses))]
    scale_labels = []
    for level in (25, 50, 75, 100):
        scale = level / 100
        points = []
        for angle in angles:
            points.extend([center_x + math.cos(angle) * radius * scale, center_y + math.sin(angle) * radius * scale])
        drawing.add(Polygon(points, fillColor=None, strokeColor=PDF_THEME["line"], strokeWidth=0.65))
        scale_labels.append((level, center_x + 18, center_y + radius * scale - 2))

    value_points = []
    for angle, clause in zip(angles, clauses):
        axis_x = center_x + math.cos(angle) * radius
        axis_y = center_y + math.sin(angle) * radius
        drawing.add(Line(center_x, center_y, axis_x, axis_y, strokeColor=PDF_THEME["line"], strokeWidth=0.5))
        label_x = center_x + math.cos(angle) * label_radius
        label_y = center_y + math.sin(angle) * label_radius
        anchor = "middle"
        if label_x < center_x - 10:
            anchor = "end"
        elif label_x > center_x + 10:
            anchor = "start"
        drawing.add(String(label_x, label_y - 3, f"C{clause['numero']}", fontName=PDF_FONT_BOLD, fontSize=7.8, fillColor=PDF_THEME["gray_dark"], textAnchor=anchor))
        percent = min(clause["percent"] or 0, 100)
        value_points.extend([center_x + math.cos(angle) * radius * percent / 100, center_y + math.sin(angle) * radius * percent / 100])

    if any((clause["percent"] or 0) > 0 for clause in clauses):
        radar_color = _traffic_color_for_percent(summary["percent"])
        drawing.add(Polygon(value_points, fillColor=_traffic_soft_color_for_percent(summary["percent"]), strokeColor=radar_color, strokeWidth=1.8))
        for point_index in range(0, len(value_points), 2):
            drawing.add(Rect(value_points[point_index] - 2, value_points[point_index + 1] - 2, 4, 4, fillColor=radar_color, strokeColor=radar_color))
    else:
        drawing.add(String(center_x, center_y - 4, "Sin cumplimiento registrado", fontName=PDF_FONT_REGULAR, fontSize=8.2, fillColor=PDF_THEME["muted"], textAnchor="middle"))

    for level, label_x, label_y in scale_labels:
        drawing.add(Rect(label_x - 7, label_y - 3, 24, 10, fillColor=colors.white, strokeColor=colors.white, strokeWidth=0))
        drawing.add(String(label_x + 5, label_y, str(level), fontName=PDF_FONT_BOLD, fontSize=6.7, fillColor=PDF_THEME["gray_dark"], textAnchor="middle"))

    legend_rows = [
        ("0-40", "Atención prioritaria", PDF_THEME["low"]),
        ("41-80", "Implementación parcial", PDF_THEME["medium"]),
        ("81-100", "Madurez alta", PDF_THEME["traffic_green"]),
    ]
    legend_x = 392
    legend_y = 150
    for index, (range_label, label, color) in enumerate(legend_rows):
        y = legend_y - index * 18
        drawing.add(Rect(legend_x, y, 8, 8, fillColor=color, strokeColor=color))
        drawing.add(String(legend_x + 14, y + 1, f"{range_label}: {label}", fontName=PDF_FONT_REGULAR, fontSize=7.3, fillColor=PDF_THEME["muted"]))
    return drawing


def _build_clause_heatmap_table(summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    headers = ["Cláusula", "0%", "1-20", "21-40", "41-60", "61-80", "81-100", "Cumpl."]
    rows = [[Paragraph(header, styles["table_header"]) for header in headers]]
    heat_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
        ("GRID", (0, 0), (-1, -1), 0.45, PDF_THEME["line"]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_index, clause in enumerate(summary["clauses"], start=1):
        bucket = _maturity_bucket(clause["percent"])
        cells = [
            Paragraph(f"{clause['numero']} {escape(clause['nombre'])}", styles["small"]),
            "",
            "",
            "",
            "",
            "",
            "",
            Paragraph(_format_percent(clause["percent"]), styles["small_center"]),
        ]
        if bucket is not None:
            cells[bucket + 1] = Paragraph("OK", _traffic_label_style_for_percent(clause["percent"], styles))
            color = _traffic_color_for_percent(clause["percent"])
            heat_styles.extend(
                [
                    ("BACKGROUND", (bucket + 1, row_index), (bucket + 1, row_index), color),
                    ("TEXTCOLOR", (bucket + 1, row_index), (bucket + 1, row_index), colors.white),
                    ("FONTNAME", (bucket + 1, row_index), (bucket + 1, row_index), PDF_FONT_BOLD),
                ]
            )
        rows.append(cells)
    table = Table(
        rows,
        colWidths=[2.45 * inch, 0.52 * inch, 0.58 * inch, 0.62 * inch, 0.62 * inch, 0.62 * inch, 0.7 * inch, 0.72 * inch],
        repeatRows=1,
    )
    table.setStyle(TableStyle(heat_styles))
    return table


def _build_priority_sections_table(sections: list[dict], styles: dict[str, ParagraphStyle]) -> Table:
    rows = [[
        "Apartado",
        "Cláusula",
        "Avance",
        "Cumpl.",
        "Aplicables",
        "Evid.",
        "Prioridad",
    ]]
    if not sections:
        rows.append(["Sin apartados aplicables para priorizar", "-", "-", "-", "-", "-", "-"])
    for section in sections[:10]:
        rows.append(
            [
                f"{section['codigo']} {section['nombre']}",
                section["clausula"],
                _format_percent(section["completion"]),
                _format_percent(section["percent"]),
                str(section["applicable"]),
                str(_section_evidence_count(section)),
                _priority_label(section["percent"]),
            ]
        )
    return _pdf_table(
        rows,
        styles,
        col_widths=[2.55 * inch, 0.6 * inch, 0.74 * inch, 0.74 * inch, 0.75 * inch, 0.55 * inch, 1.05 * inch],
        center_from_col=1,
    )


def _build_evidence_coverage_table(summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    rows = [["Cláusula", "Reactivos con evidencia", "Archivos", "Cobertura", "Lectura"]]
    for row in _evidence_coverage(summary):
        rows.append(
            [
                row["clause"],
                str(row["questions_with_evidence"]),
                str(row["files"]),
                _format_percent(row["coverage"]),
                row["label"],
            ]
        )
    return _pdf_table(
        rows,
        styles,
        col_widths=[2.5 * inch, 1.35 * inch, 0.8 * inch, 0.9 * inch, 1.45 * inch],
        center_from_col=1,
    )


def _build_normative_support_table(summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    rows = [["Cláusula", "Soporte ISO 9001:2015 (paráfrasis)", "Evidencia de soporte"]]
    for clause in summary["clauses"]:
        support = _clause_support_row(clause)
        rows.append(
            [
                f"{clause['numero']} {support['title']}",
                support["support"],
                support["evidence"],
            ]
        )
    return _pdf_table(
        rows,
        styles,
        col_widths=[1.45 * inch, 3.0 * inch, 2.6 * inch],
    )


def _build_clause_guidance_table(summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    rows = [["Cláusula", "Madurez", "Comentario guía", "Siguiente paso"]]
    for clause in summary["clauses"]:
        guidance = _clause_guidance_row(clause)
        rows.append(
            [
                guidance["clause"],
                guidance["maturity"],
                guidance["reading"],
                guidance["next_step"],
            ]
        )
    return _pdf_table(
        rows,
        styles,
        col_widths=[0.95 * inch, 1.15 * inch, 2.95 * inch, 2.15 * inch],
    )


def _build_clause_story(clause: dict, styles: dict[str, ParagraphStyle]) -> list:
    attention = [
        section
        for section in clause["sections"]
        if section["applicable"] > 0 and section["percent"] is not None and section["percent"] < 80
    ]
    attention.sort(key=lambda item: (item["percent"], item["completion"], item["codigo"]))
    guidance = _clause_guidance_row(clause)
    support = _clause_support_row(clause)
    story: list = [
        Paragraph(f"Cláusula {clause['numero']} - {_paragraph_escape(clause['nombre'])}", styles["section"]),
        _build_clause_metrics_table(clause),
        Spacer(1, 0.1 * inch),
        Paragraph("Comentario guía", styles["subsection"]),
        Paragraph(
            f"{_paragraph_escape(guidance['reading'])} <b>Siguiente paso:</b> {_paragraph_escape(guidance['next_step'])}",
            styles["body"],
        ),
        Spacer(1, 0.08 * inch),
        Paragraph("Soporte normativo de referencia", styles["subsection"]),
        Paragraph(
            f"<b>{_paragraph_escape(support['title'])}:</b> {_paragraph_escape(support['support'])} "
            f"<b>Evidencia:</b> {_paragraph_escape(support['evidence'])}",
            styles["body"],
        ),
        Spacer(1, 0.1 * inch),
        Paragraph("Subapartados evaluados", styles["subsection"]),
        _build_clause_sections_table(clause, styles),
        Spacer(1, 0.12 * inch),
        Paragraph("Áreas de atención", styles["subsection"]),
    ]
    if attention:
        for section in attention[:5]:
            story.append(
                Paragraph(
                    f"- {section['codigo']} {_paragraph_escape(section['nombre'])}: "
                    f"{_paragraph_escape(_section_maturity_comment(section))}",
                    styles["body"],
                )
            )
    else:
        story.append(Paragraph("Sin apartados críticos identificados con la información capturada.", styles["body"]))
    return story


def _build_clause_metrics_table(clause: dict) -> Table:
    table = Table(
        [
            ["Reactivos", "Respondidos", "Aplicables", "N/A", "Cumplimiento", "Madurez"],
            [
                str(clause["total"]),
                str(clause["answered"]),
                str(clause["applicable"]),
                str(clause["na"]),
                _format_percent(clause["percent"]),
                clause["maturity_label"],
            ],
        ],
        colWidths=[0.9 * inch, 0.95 * inch, 0.95 * inch, 0.62 * inch, 1.1 * inch, 2.1 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy_soft"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.45, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _build_clause_sections_table(clause: dict, styles: dict[str, ParagraphStyle]) -> Table:
    rows = [["Apartado", "React.", "Avance", "Aplicables", "N/A", "Cumpl.", "Evid.", "Madurez"]]
    for section in clause["sections"]:
        rows.append(
            [
                f"{section['codigo']} {section['nombre']}",
                str(section["total"]),
                _format_percent(section["completion"]),
                str(section["applicable"]),
                str(section["na"]),
                _format_percent(section["percent"]),
                str(_section_evidence_count(section)),
                section["maturity_label"],
            ]
        )
    return _pdf_table(
        rows,
        styles,
        col_widths=[2.45 * inch, 0.62 * inch, 0.72 * inch, 0.72 * inch, 0.46 * inch, 0.72 * inch, 0.46 * inch, 1.0 * inch],
        center_from_col=1,
    )


def _build_findings_appendix(summary: dict, styles: dict[str, ParagraphStyle]) -> Table:
    rows = [["Cláusula", "Apartado", "Reactivo", "Respuesta", "Observación / evidencia", "Evid."]]
    findings = _relevant_questions(summary)
    if not findings:
        rows.append(["Sin hallazgos", "-", "No hay reactivos con brecha, observación o evidencia registrada.", "-", "-", "-"])
    for item in findings:
        notes = item["observacion"] or "Sin observación"
        if item["evidence_count"]:
            notes = f"{notes} | Evidencias cargadas: {item['evidence_count']}"
        rows.append(
            [
                item["clause"],
                item["section"],
                item["question"],
                item["answer"],
                notes,
                str(item["evidence_count"]),
            ]
        )
    return _pdf_table(
        rows,
        styles,
        col_widths=[0.62 * inch, 0.72 * inch, 2.35 * inch, 0.8 * inch, 2.15 * inch, 0.45 * inch],
        center_from_col=3,
    )


def _pdf_table(
    rows: list[list],
    styles: dict[str, ParagraphStyle],
    *,
    col_widths: list | None = None,
    repeat_rows: int = 1,
    header: bool = True,
    center_from_col: int | None = None,
) -> Table:
    table_rows = []
    for row_index, row in enumerate(rows):
        style = styles["table_header"] if header and row_index == 0 else styles["small"]
        table_rows.append([_as_paragraph(value, style) for value in row])
    table = Table(table_rows, colWidths=col_widths, repeatRows=repeat_rows if header else 0, hAlign="LEFT")
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.4, PDF_THEME["line"]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1 if header else 0), (-1, -1), [colors.white, colors.HexColor("#F5F2EF")]),
    ]
    if header:
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ]
        )
    else:
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (0, -1), PDF_THEME["navy_soft"]),
                ("BACKGROUND", (2, 0), (2, -1), PDF_THEME["navy_soft"]),
            ]
        )
    if center_from_col is not None:
        commands.append(("ALIGN", (center_from_col, 1 if header else 0), (-1, -1), "CENTER"))
    table.setStyle(TableStyle(commands))
    return table


def _response_distribution(summary: dict) -> list[dict]:
    counts = {key: 0 for key, _label, _color in ISO_RESPONSE_ORDER}
    for section in summary["sections"]:
        for row in section["questions"]:
            selected = row["selected"]
            if selected in counts:
                counts[selected] += 1
    total = sum(counts.values())
    distribution = []
    for key, label, color in ISO_RESPONSE_ORDER:
        count = counts[key]
        percent = round((count / total) * 100, 2) if total else 0.0
        distribution.append({"key": key, "label": label, "count": count, "percent": percent, "color": color})
    return distribution


def _global_maturity_comment(summary: dict) -> str:
    profile = _maturity_profile(summary["percent"])
    return (
        f"{profile['reading']} La lectura global debe revisarse con enfoque de procesos, liderazgo, evidencia "
        "objetiva y mejora continua; use los comentarios por cláusula para convertir brechas en acciones verificables."
    )


def _clause_guidance_row(clause: dict) -> dict:
    guide = ISO_CLAUSE_GUIDANCE.get(
        str(clause["numero"]),
        {"focus": "requisitos del sistema de gestión de la calidad", "evidence": "evidencia objetiva del cumplimiento"},
    )
    profile = _maturity_profile(clause["percent"])
    counts = _clause_response_counts(clause)
    evidence_files = sum(_section_evidence_count(section) for section in clause["sections"])
    reading = (
        f"{profile['reading']} En esta cláusula el foco es {guide['focus']}. "
        f"El corte muestra {counts['no']} No, {counts['parcial']} Parcial, {counts['si']} Sí, "
        f"{counts['na']} N/A y {evidence_files} evidencias activas."
    )
    next_step = f"{profile['action']} Evidencia guía: {guide['evidence']}."
    return {
        "clause": f"{clause['numero']} {clause['nombre']}",
        "maturity": clause["maturity_label"],
        "reading": _clip(reading, 430),
        "next_step": _clip(next_step, 300),
    }


def _clause_support_row(clause: dict) -> dict:
    support = ISO_CLAUSE_SUPPORT.get(
        str(clause["numero"]),
        {
            "title": clause.get("nombre", "Requisito ISO 9001:2015"),
            "support": "Revisar la conformidad del requisito aplicable mediante evidencia objetiva y trazable.",
            "evidence": "Registros, controles, resultados y documentos vigentes relacionados con el alcance declarado.",
        },
    )
    return {
        "title": support["title"],
        "support": support["support"],
        "evidence": support["evidence"],
    }


def _section_maturity_comment(section: dict) -> str:
    guide = ISO_CLAUSE_GUIDANCE.get(
        str(section["clausula"]),
        {"focus": "requisitos del sistema de gestión de la calidad", "evidence": "evidencia objetiva del cumplimiento"},
    )
    profile = _maturity_profile(section["percent"])
    counts = _section_response_counts(section)
    evidence_count = _section_evidence_count(section)
    if section["completion"] < 100:
        capture_note = f"Complete la captura: avance {_format_percent(section['completion'])}."
    else:
        capture_note = "Captura completa."
    comment = (
        f"{profile['reading']} {capture_note} Para este apartado revise {guide['focus']}; "
        f"hay {counts['no']} No, {counts['parcial']} Parcial, {counts['na']} N/A y {evidence_count} evidencias. "
        f"Siguiente paso: {profile['action']}"
    )
    return _clip(comment, 500)


def _maturity_profile(percent: float | None) -> dict[str, str]:
    if percent is None:
        return {
            "reading": "Sin base de cálculo por ausencia de reactivos aplicables.",
            "action": "Valide la justificación de los N/A y confirme que el alcance del SGC esté correctamente delimitado.",
        }
    if percent == 0:
        return {
            "reading": "Madurez no iniciada: no se observa evidencia de implementación para los reactivos aplicables.",
            "action": "Defina responsables, controles mínimos, registros requeridos y fecha de arranque del plan de implementación.",
        }
    if percent <= 20:
        return {
            "reading": "Madurez inicial: existen señales aisladas, pero el control todavía depende de acciones reactivas o informales.",
            "action": "Formalice el método de trabajo, asigne responsables y capture evidencia objetiva suficiente.",
        }
    if percent <= 40:
        return {
            "reading": "Madurez en desarrollo: hay prácticas parciales, pero falta consistencia, despliegue o trazabilidad documental.",
            "action": "Convierta las prácticas parciales en procedimientos, registros, indicadores y seguimiento periódico.",
        }
    if percent <= 60:
        return {
            "reading": "Madurez definida: el requisito está identificado y opera parcialmente, aunque aún requiere mayor evidencia y control.",
            "action": "Cierre brechas de evidencia, verifique eficacia y conecte el apartado con riesgos, objetivos e indicadores.",
        }
    if percent <= 80:
        return {
            "reading": "Madurez gestionada: el requisito funciona con evidencia razonable, pero todavía puede fortalecerse la medición.",
            "action": "Use auditorías internas, análisis de datos y revisión directiva para sostener y mejorar el desempeño.",
        }
    return {
        "reading": "Madurez optimizada: el requisito está integrado al trabajo diario y cuenta con evidencia sólida.",
        "action": "Mantenga seguimiento, comparta buenas prácticas y documente mejoras preventivas o innovaciones del proceso.",
    }


def _clause_response_counts(clause: dict) -> dict[str, int]:
    counts = {"no": 0, "parcial": 0, "si": 0, "na": 0}
    for section in clause["sections"]:
        section_counts = _section_response_counts(section)
        for key in counts:
            counts[key] += section_counts[key]
    return counts


def _section_response_counts(section: dict) -> dict[str, int]:
    counts = {"no": 0, "parcial": 0, "si": 0, "na": 0}
    for row in section["questions"]:
        selected = row["selected"]
        if selected in counts:
            counts[selected] += 1
    return counts


def _priority_sections(summary: dict, limit: int = 10) -> list[dict]:
    sections = [
        section
        for section in summary["sections"]
        if section["applicable"] > 0 and section["percent"] is not None
    ]
    sections.sort(key=lambda section: (section["percent"], section["completion"], section["codigo"]))
    return sections[:limit]


def _section_evidence_count(section: dict) -> int:
    return sum(len(row["evidence"]) for row in section["questions"])


def _section_questions_with_evidence(section: dict) -> int:
    return sum(1 for row in section["questions"] if row["evidence"])


def _evidence_coverage(summary: dict) -> list[dict]:
    rows = []
    for clause in summary["clauses"]:
        files = sum(_section_evidence_count(section) for section in clause["sections"])
        questions = sum(_section_questions_with_evidence(section) for section in clause["sections"])
        coverage = round((questions / clause["applicable"]) * 100, 2) if clause["applicable"] else None
        if coverage is None:
            label = "Sin aplicables"
        elif coverage >= 60:
            label = "Soporte amplio"
        elif coverage >= 25:
            label = "Soporte parcial"
        else:
            label = "Reforzar evidencia"
        rows.append(
            {
                "clause": f"{clause['numero']} {clause['nombre']}",
                "questions_with_evidence": questions,
                "files": files,
                "coverage": coverage,
                "label": label,
            }
        )
    return rows


def _relevant_questions(summary: dict) -> list[dict]:
    findings = []
    for clause in summary["clauses"]:
        for section in clause["sections"]:
            for row in section["questions"]:
                evidence_count = len(row["evidence"])
                observation = (row["observacion"] or "").strip()
                selected = row["selected"]
                if selected not in {"no", "parcial", "na"} and not observation and not evidence_count:
                    continue
                reactive = row["reactivo"]
                findings.append(
                    {
                        "clause": clause["numero"],
                        "section": section["codigo"],
                        "question": _clip(f"{reactive.codigo} {reactive.texto}", 290),
                        "answer": row["selected_label"],
                        "observacion": _clip(observation, 230),
                        "evidence_count": evidence_count,
                    }
                )
    return findings


def _priority_label(percent: float | None) -> str:
    if percent is None:
        return "Sin base"
    if percent <= 40:
        return "Alta"
    if percent <= 70:
        return "Media"
    return "Seguimiento"


def _maturity_bucket(percent: float | None) -> int | None:
    if percent is None:
        return None
    if percent == 0:
        return 0
    if percent <= 20:
        return 1
    if percent <= 40:
        return 2
    if percent <= 60:
        return 3
    if percent <= 80:
        return 4
    return 5


def _traffic_color_for_percent(percent: float | int | None):
    if percent is None:
        return PDF_THEME["empty"]
    value = float(percent)
    if value <= 40:
        return PDF_THEME["traffic_red"]
    if value <= 80:
        return PDF_THEME["traffic_yellow"]
    return PDF_THEME["traffic_green"]


def _traffic_soft_color_for_percent(percent: float | int | None):
    if percent is None:
        return PDF_THEME["sage_soft"]
    value = float(percent)
    if value <= 40:
        return PDF_THEME["traffic_red_soft"]
    if value <= 80:
        return PDF_THEME["traffic_yellow_soft"]
    return PDF_THEME["traffic_green_soft"]


def _traffic_label_style_for_percent(percent: float | int | None, styles: dict[str, ParagraphStyle]) -> ParagraphStyle:
    if percent is None:
        return styles["traffic_label_dark"]
    value = float(percent)
    if value <= 40 or value > 80:
        return styles["traffic_label_light"]
    return styles["traffic_label_dark"]


def _traffic_label_style_for_slug(slug: str, styles: dict[str, ParagraphStyle]) -> ParagraphStyle:
    if slug in {"low", "optimal"}:
        return styles["traffic_label_light"]
    return styles["traffic_label_dark"]


def _compact_maturity_label(label: str | None) -> str:
    text = str(label or "-")
    if " - " in text:
        return text.split(" - ", 1)[1]
    return text


def _maturity_level_prefix(label: str | None) -> str:
    text = str(label or "")
    if " - " in text:
        return text.split(" - ", 1)[0]
    return ""


def _format_percent(value: float | int | None) -> str:
    if value is None:
        return "-"
    numeric = round(float(value), 2)
    text = f"{numeric:.2f}".rstrip("0").rstrip(".")
    return f"{text}%"


def _date_or_dash(value) -> str:
    return value.strftime("%d/%m/%Y") if value else "-"


def _paragraph_escape(value) -> str:
    return escape(str(value or ""), quote=False)


def _as_paragraph(value, style: ParagraphStyle):
    if isinstance(value, Paragraph):
        return value
    return Paragraph(_paragraph_escape(value), style)


def _clip(value: str, limit: int) -> str:
    value = " ".join(str(value or "").split())
    if len(value) <= limit:
        return value
    return f"{value[: max(limit - 3, 0)].rstrip()}..."


def _draw_iso_pdf_chrome(canvas, doc) -> None:
    canvas.saveState()
    width, height = doc.pagesize
    canvas.setFillColor(PDF_THEME["navy"])
    canvas.rect(doc.leftMargin, height - 0.32 * inch, width - doc.leftMargin - doc.rightMargin, 0.1 * inch, fill=1, stroke=0)
    canvas.setFillColor(PDF_THEME["muted"])
    canvas.setFont(PDF_FONT_REGULAR, 7.2)
    canvas.drawString(doc.leftMargin, 0.28 * inch, ISO_PDF_FOOTER)
    canvas.drawRightString(width - doc.rightMargin, 0.28 * inch, f"Pagina {canvas.getPageNumber()}")
    canvas.restoreState()


def _write_title(sheet, title: str, subtitle: str, *, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=18, color=EXCEL_COLORS["white"])
    sheet["A1"].fill = PatternFill("solid", fgColor=EXCEL_COLORS["navy"])
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    sheet["A2"] = subtitle
    sheet["A2"].fill = PatternFill("solid", fgColor=EXCEL_COLORS["soft"])
    sheet["A2"].alignment = Alignment(horizontal="center")


def _style_table(sheet, header_row: int) -> None:
    border = Border(
        left=Side(style="thin", color=EXCEL_COLORS["line"]),
        right=Side(style="thin", color=EXCEL_COLORS["line"]),
        top=Side(style="thin", color=EXCEL_COLORS["line"]),
        bottom=Side(style="thin", color=EXCEL_COLORS["line"]),
    )
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == header_row:
                cell.fill = PatternFill("solid", fgColor=EXCEL_COLORS["navy"])
                cell.font = Font(color=EXCEL_COLORS["white"], bold=True)


def _set_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 24


def _table(rows: list[list], repeat_rows: int = 0) -> Table:
    table = Table(rows, repeatRows=repeat_rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_THEME["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("GRID", (0, 0), (-1, -1), 0.35, PDF_THEME["line"]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F2EF")]),
            ]
        )
    )
    return table
